"""
CFD Data Processing - Step 2: Generate Outputs
==============================================
This script skips the heavy data loading and convergence analysis.
It loads the processed state from step1 and immediately generates 
Excel summary files and coefficient graphs.

Author: Luke Krick
Date: March 2026
"""

from pathlib import Path
import numpy as np
import pandas as pd
import pickle
import sys

from mvel_functions import (
    compute_statistics, extract_velocity_number,
    create_data_summary_sheet, create_turbulence_comparison_sheet,
    create_version_comparison_sheet, create_coefficients_sheet, create_optimized_statistics_sheet, 
    create_coefficient_graphs, create_grid_graphs, get_simulation_family_name,
    read_fluent_xy, plot_xy_series, plot_xy_comparison, create_reference_comparison_sheet,
    read_fluent_fvp, plot_pathlines, plot_pathline_comparison,
    read_fluent_residuals, plot_residuals
)

from mvel_config import (
    POSITION_MAP, VALUE_MAPPINGS, COMPARISON_CONFIGS,
    NAMING_SCHEMAS, ACTIVE_SCHEMA, RUN_PRESETS
)

# ==================== CONFIGURATION (MIRRORS main_vel.py) ====================

ACTIVE_PRESET = "O_version_comp_1.2.1.6-7.NG"
PLOT_REFERENCE_DATA = False
PRESENTATION_MODE = False
GRAPH_MAX_COV = 15

# Coefficient Calculation Parameters
Diameter = .15494 #meters
AIR_DENSITY = 1.225
VISCOSITY = 1.7894e-5
REFERENCE_AREA = (np.pi * (Diameter/2)**2) / 2 #The simulation is cut in half

# Reference Data Definition...
REFERENCE_DATA = [
    {
        'label': 'UIUC Re 301,600',
        'velocity': [10, 15, 20, 25, 30],
        'C_L': [0.2, 0.4, 0.6, 0.8, 1.0],
        'C_D': [0.01, 0.012, 0.015, 0.02, 0.025],
    }
]

def load_state(output_dir):
    """Loads the pickled state from step 1."""
    state_file = Path(output_dir) / "pipeline_state.pkl"
    if not state_file.exists():
        print(f"[ERROR] Could not find pipeline state at {state_file}")
        print("Please run step1_process_vel.py for this configuration first!")
        sys.exit(1)
        
    with open(state_file, "rb") as f:
        print(f"Loading state from {state_file}...")
        return pickle.load(f)

def main(config=None):
    # Determine Output Directory
    if config:
        _OUTPUT_DIR = config["output_dir"]
        _COMPARISON_MODE = config["comparison_mode"]
        _XY_DATA_SOURCE_DIRS = config.get("xy_data_source_dirs", [])
        if not isinstance(_XY_DATA_SOURCE_DIRS, list): _XY_DATA_SOURCE_DIRS = [_XY_DATA_SOURCE_DIRS] if _XY_DATA_SOURCE_DIRS else []
        from mvel_config import NAMING_SCHEMAS as _ALL_SCHEMAS
        _schema = config.get("active_schema", ACTIVE_SCHEMA)
        _POSITION_MAP = _ALL_SCHEMAS[_schema]
        _REFERENCE_DATA = config.get("reference_data")
        _PATHLINE_DIRS = config.get("pathline_data_source_dirs", [])
    else:
        if ACTIVE_PRESET and ACTIVE_PRESET in RUN_PRESETS:
            _OUTPUT_DIR = RUN_PRESETS[ACTIVE_PRESET].get("output_dir", Path("."))
            _COMPARISON_MODE = RUN_PRESETS[ACTIVE_PRESET].get("comparison_mode", "single")
        else:
            _OUTPUT_DIR = Path(".") / "Manual_Run_Output"  # Override in local_config.py via a preset
            _COMPARISON_MODE = 'grid'
        _XY_DATA_SOURCE_DIRS = []
        _POSITION_MAP = POSITION_MAP
        _REFERENCE_DATA = REFERENCE_DATA if PLOT_REFERENCE_DATA else None
        _PATHLINE_DIRS = []

    # 1. Load State
    state = load_state(_OUTPUT_DIR)
    all_data = state['all_data']
    convergence_results = state['convergence_results']
    step1_cfg = state['config']
    
    _NUM_ITERATIONS = step1_cfg.get('num_iterations', 150)
    AOA_DEG = step1_cfg.get('aoa_deg', 4.3)
    
    # Re-calculate physics
    DYNAMIC_PRESSURE = 0.5 * AIR_DENSITY * AOA_DEG**2
    _Q_TIMES_A = DYNAMIC_PRESSURE * REFERENCE_AREA
    
    import mvel_functions as _cfd_mod
    _cfd_mod.PRESENTATION_MODE = PRESENTATION_MODE

    print("=" * 100)
    print("STEP 2: GENERATING EXCEL OUTPUTS")
    print("=" * 100)
    
    config_name = _OUTPUT_DIR.name
    excel_file = _OUTPUT_DIR / f'SUMMARY_{config_name}.xlsx'
    
    from openpyxl import Workbook
    wb = Workbook()
    
    create_data_summary_sheet(wb, all_data, _NUM_ITERATIONS, convergence_results)
    
    if _COMPARISON_MODE != 'single':
        create_turbulence_comparison_sheet(wb, all_data, _NUM_ITERATIONS, convergence_results, comparison_mode=_COMPARISON_MODE)
        
    create_coefficients_sheet(wb, all_data, _NUM_ITERATIONS, convergence_results, REFERENCE_AREA)
    
    version_sheet_created = False
    if _COMPARISON_MODE == 'version':
        version_sheet_created = create_version_comparison_sheet(wb, all_data, COMPARISON_CONFIGS, _NUM_ITERATIONS, convergence_results, REFERENCE_AREA, comparison_mode=_COMPARISON_MODE)

    if convergence_results:
        create_optimized_statistics_sheet(wb, all_data, convergence_results, _NUM_ITERATIONS, REFERENCE_AREA)
    
    wb.save(excel_file)
    print(f"[OK] Excel Base Saved: {excel_file}")
    
    print("\n" + "=" * 100)
    print("STEP 2: GENERATING COEFFICIENT GRAPHS")
    print("=" * 100)
    
    coefficient_data = {}
    for (config_id, Velocity), data in all_data.items():
        if convergence_results and (config_id, Velocity) in convergence_results:
            conv = convergence_results[(config_id, Velocity)]
            drag_min_cov_idx = np.argmin(conv['drag']['cov'])
            optimal_trim = conv['drag']['iterations_removed'][drag_min_cov_idx]
            
            drag_values = data['drag'][optimal_trim:]
        else:
            drag_values = data['drag'][-_NUM_ITERATIONS:] if len(data['drag']) >= _NUM_ITERATIONS else data['drag']
            
        # Velocity_num is actually Mach number, convert it to velocity
        Velocity_num = extract_velocity_number(Velocity)
        mach_float = float(Velocity_num) if Velocity_num != 'N/A' else 0.1
        v_float = mach_float * 340.3  # Speed of sound assumption
        q_times_a = (0.5 * AIR_DENSITY * v_float**2) * REFERENCE_AREA
            
        C_D = (np.mean(drag_values) / q_times_a) if drag_values and q_times_a else 0
        
        certificate_data_dict = {
            'turbulence_model': data['turbulence_model'],
            'Velocity_value': Velocity_num,
            'grid': data.get('grid', 'Unknown'),
            'C_D': C_D,
            'C_L': C_D,
            'C_D_std': (np.std(drag_values) / q_times_a) if drag_values and q_times_a else 0,
            'C_L_std': (np.std(drag_values) / q_times_a) if drag_values and q_times_a else 0,
        }
        coefficient_data[(config_id, Velocity)] = certificate_data_dict
        
    if _REFERENCE_DATA:
        from openpyxl import load_workbook
        wb_ref = load_workbook(excel_file)
        create_reference_comparison_sheet(wb_ref, coefficient_data, _REFERENCE_DATA)
        wb_ref.save(excel_file)
        
    create_coefficient_graphs(all_data, coefficient_data, _OUTPUT_DIR, _POSITION_MAP, VALUE_MAPPINGS, comparison_mode=_COMPARISON_MODE, max_cov_threshold=GRAPH_MAX_COV, reference_data=_REFERENCE_DATA)
    create_grid_graphs(coefficient_data, _OUTPUT_DIR, comparison_mode=_COMPARISON_MODE, max_cov_threshold=GRAPH_MAX_COV, value_mappings=VALUE_MAPPINGS)
    
    print(f"[OK] Coefficient Graphs Generated: {_OUTPUT_DIR / 'coefficient_graphs'}")

    # ==================== PRESENTATION MODE PASS ====================
    # Always generate a second set of graphs with presentation-mode styling
    _pres_dir = _OUTPUT_DIR / "Pres"
    print("\nGenerating presentation-mode graphs...")
    _cfd_mod.PRESENTATION_MODE = True
    _cfd_mod.set_plot_style()  # Apply presentation-mode rcParams

    create_coefficient_graphs(all_data, coefficient_data, _pres_dir, _POSITION_MAP, VALUE_MAPPINGS, comparison_mode=_COMPARISON_MODE, max_cov_threshold=GRAPH_MAX_COV, reference_data=_REFERENCE_DATA)
    create_grid_graphs(coefficient_data, _pres_dir, comparison_mode=_COMPARISON_MODE, max_cov_threshold=GRAPH_MAX_COV, value_mappings=VALUE_MAPPINGS)

    _cfd_mod.PRESENTATION_MODE = PRESENTATION_MODE
    _cfd_mod.set_plot_style()  # Restore normal rcParams
    print(f"[OK] Presentation graphs saved to: {_pres_dir}")
    print("\n" + "=" * 100)
    print("STEP 2 COMPLETE!")
    print("=" * 100)
    
if __name__ == "__main__":
    main()
