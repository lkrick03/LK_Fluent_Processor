"""
CFD Data Processing - Main Execution Script
Consolidates all 4 notebooks into a single executable workflow.

This script:
1. Loads and processes CFD simulation data
2. (Optional) Runs convergence analysis
3. Generates Excel summary files
4. Creates coefficient graphs


Author: Luke Krick
Date: February 2026
"""

from pathlib import Path
import numpy as np
import pandas as pd
from cfd_functions import (
    load_lift_drag_data, compute_statistics, extract_aoa_number,
    analyze_convergence, plot_convergence_analysis, plot_convergence_summary, create_data_summary_sheet, create_turbulence_comparison_sheet,
    create_version_comparison_sheet, create_coefficients_sheet, create_optimized_statistics_sheet, apply_excel_formatting,
    create_coefficient_graphs, create_grid_graphs, apply_data_manipulations, get_simulation_family_name,
    read_fluent_xy, plot_xy_series, plot_xy_comparison, create_reference_comparison_sheet,
    read_fluent_fvp, plot_pathlines, plot_pathline_comparison,
    read_fluent_residuals, plot_residuals
)
from config import (
    POSITION_MAP, VALUE_MAPPINGS, COMPARISON_CONFIGS, DATA_MANIPULATIONS,
    NAMING_SCHEMAS, ACTIVE_SCHEMA, RUN_PRESETS
)


# ==================== USER CONFIGURATION ====================

# Set ACTIVE_PRESET to the key of the preset you want to run, 
# or None to use manual settings below.
ACTIVE_PRESET = "single_4.3.2.NG"

if ACTIVE_PRESET and ACTIVE_PRESET in RUN_PRESETS:
    preset = RUN_PRESETS[ACTIVE_PRESET]
    print(f"Loading Configuration Preset: '{preset.get('name', ACTIVE_PRESET)}'")
    DATA_SOURCES = preset.get("data_sources", [])
    OUTPUT_DIR = preset.get("output_dir", Path("."))
    COMPARISON_MODE = preset.get("comparison_mode", "single")
    AOA_FILTER = preset.get("aoa_filter", [])
else:
    print("Loading Manual Configuration (ACTIVE_PRESET is None or Invalid)")
    
    # Input/Output Directories
    # DATA_SOURCES acts as a priority list. If duplicates exist, the one with the higher Version number wins.
    # If versions are identical, they are treated as duplicates.
    DATA_SOURCES = [
        # Example: Path(r"C:\Users\lukek\OneDrive\Documents\Thesis\NACA_2414_2D\Fluent\Directories\2414.6.5.6\5.6.1.1.NG"),
    ]
    OUTPUT_DIR = Path(r"C:\Users\lukek\OneDrive\Documents\Thesis\NACA_2414_2D\Fluent\Directories\Processed_Data\Comparisons\Manual_Run")

    # Comparison Mode
    # Options: 'single', 'turbulence', 'grid', 'mesh', 'version', 'mixed', 'family_grid'
    COMPARISON_MODE = 'grid'

    # AoA Filter: Set to a list of angles (e.g., [0, 2, 4]) to only process those AoAs.
    # Set to [] or None to process all.
    AOA_FILTER = []

# Configuration Extraction Method
CONFIG_EXTRACTION_METHOD = 'case_file'  # Options: 'case_file' or 'folder'

# ==================== REFERENCE DATA (Experimental / Published) ====================
# Set PLOT_REFERENCE_DATA = True to overlay reference data on coefficient graphs.
# Each entry in the list is one reference dataset with a label, AoA, C_L, and C_D arrays.
# You can add multiple datasets (e.g., different experiments, different Re).
PLOT_REFERENCE_DATA = False       # Toggle on/off

# Path to folder containing .xy files (if not in source folders)
# Auto-discovered: scans each DATA_SOURCES path for a 'y_plus_pressure_data' subfolder
XY_DATA_SOURCE_DIR = [
    src / "y_plus_pressure_data" for src in DATA_SOURCES
    if (src / "y_plus_pressure_data").exists()
]

# Path to folder(s) containing .fvp pathline files (exported by jou_post_exporter.py)
# Set to [] or None to skip pathline processing
PROCESS_PATHLINES = False  # Toggle on/off
PATHLINE_DATA_SOURCE_DIRS = [
    # Path(r"C:\...\pathline_data"),
]

# Presentation Mode: larger fonts, thicker lines, bigger markers for slides/projectors
# NOTE: Presentation-mode graphs are now always generated in a 'Presentation' subfolder.
#       This flag only controls whether the PRIMARY output also uses presentation styling.
PRESENTATION_MODE = False

# Processing Parameters
NUM_ITERATIONS = 150  # Number of last iterations to use for statistics only if convergenc anaylsis is turned off
RUN_CONVERGENCE_ANALYSIS = True  # Set to False to skip convergence analysis
CONVERGENCE_MAX_TRIM = 0.8  # Maximum fraction of data to trim (0.8 = 80%)
CONVERGENCE_NUM_TESTS = 20  # Number of trim amounts to test
GRAPH_MAX_COV = 15  # Data points with COV > 5% will be excluded from graphs

# Coefficient Calculation Parameters
SPAN = 0.85344 # [m] Test width is 2.8 feet
CHORD = 0.3048 # [m] Test chord is 1 foot
AIR_DENSITY = 1.225 # [kg/m^3] Standard sea level air density
#Velocity is based on config.py

# Auto-detect velocity from the first DATA_SOURCES folder name using config.py mappings
if DATA_SOURCES:
    _vel_idx = int(DATA_SOURCES[0].name.split('.')[POSITION_MAP['velocity']])
    VELOCITY = float(VALUE_MAPPINGS['velocity'][_vel_idx])
    print(f"Auto-detected velocity: {VELOCITY} m/s (from index {_vel_idx})")
else:
    VELOCITY = 0.0
    print("[WARN] No DATA_SOURCES configured — VELOCITY set to 0.0")

REFERENCE_AREA = SPAN * CHORD
DYNAMIC_PRESSURE = 0.5 * AIR_DENSITY * VELOCITY**2
Q_TIMES_A = DYNAMIC_PRESSURE * REFERENCE_AREA

# Reynolds Number Calculation
VISCOSITY = 1.7894e-5  # [kg/(m·s)] Dynamic viscosity of air at sea level (15°C)
REYNOLDS_NUMBER = (AIR_DENSITY * VELOCITY * CHORD) / VISCOSITY
print(f"Reynolds Number: {REYNOLDS_NUMBER:,.0f}")



REFERENCE_DATA = [
    {
        'label': 'UIUC Re 301,600',  # Legend label
        'aoa':  [-6.56,-4.98,-3.56,-1.91,-0.37,1.18,2.74,4.25,5.77,7.29,8.83],  # Angles of attack [deg]
        'C_L':  [-0.4, -0.252, -0.086, 0.092, 0.230, 0.432, 0.596, 0.743, 0.874, 0.980, 1.061],  # Lift coefficients
        'C_D':  [0.0285, 0.0153, 0.0114, 0.0080, 0.0071, 0.0080, 0.0095, 0.0118, 0.0155, 0.0224, 0.0322],  # Drag coefficients
    },
    # Add more datasets by copying the block above, for example:
    # {
    #     'label': 'Wind Tunnel Data',
    #     'aoa':  [0, 5, 10, 15, 20],
    #     'C_L':  [0.22, 0.70, 1.05, 1.22, 0.90],
    #     'C_D':  [0.011, 0.020, 0.042, 0.075, 0.160],
    # },
]

# ==================== MAIN WORKFLOW ====================

def export_force_data(filepath, data, metadata, force_type):
    """Helper to export force data to text file."""
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(f"# Configuration: {metadata['config']} | AoA: {metadata['aoa']} | Turbulence: {metadata['turbulence_model']}\n")
        f.write(f"# Geometry: {metadata['geometry']} | Mesh: {metadata['mesh']} | Grid: {metadata['grid']}\n")
        f.write(f"# Total Points: {len(data)}\n#\n")
        np.savetxt(f, data, fmt='%.6f')

def main(config=None):
    """Main execution function.
    
    Args:
        config: Optional dict of settings (from GUI). When None, uses
                the module-level constants defined above, so running
                ``python main.py`` directly is unchanged.
    """
    # -- Config override: GUI passes a dict, CLI uses module-level globals --
    _config_dict = config  # Save before loop variables shadow 'config'
    if config:
        _DATA_SOURCES           = config["data_sources"]
        _OUTPUT_DIR             = config["output_dir"]
        _CONFIG_EXTRACTION_METHOD = config["config_extraction_method"]
        _COMPARISON_MODE        = config["comparison_mode"]
        _AOA_FILTER             = config.get("aoa_filter")
        _xy_val                 = config.get("xy_data_source_dirs", XY_DATA_SOURCE_DIR)
        _XY_DATA_SOURCE_DIRS    = _xy_val if isinstance(_xy_val, list) else ([_xy_val] if _xy_val else [])
        _NUM_ITERATIONS         = config["num_iterations"]
        _RUN_CONVERGENCE        = config["run_convergence_analysis"]
        _CONVERGENCE_MAX_TRIM   = config["convergence_max_trim"]
        _CONVERGENCE_NUM_TESTS  = config["convergence_num_tests"]
        _GRAPH_MAX_COV          = config["graph_max_cov"]
        _Q_TIMES_A              = config["q_times_a"]
        _REYNOLDS_NUMBER        = config.get("reynolds_number", REYNOLDS_NUMBER)
        # config.py settings (may be overridden by schema choice)
        from config import NAMING_SCHEMAS as _ALL_SCHEMAS
        _schema = config.get("active_schema", ACTIVE_SCHEMA)
        _POSITION_MAP           = _ALL_SCHEMAS[_schema]
        _VALUE_MAPPINGS         = VALUE_MAPPINGS  # shared, not overridden
        _COMPARISON_CONFIGS     = COMPARISON_CONFIGS
        _DATA_MANIPULATIONS     = DATA_MANIPULATIONS
        _REFERENCE_DATA         = config.get("reference_data")  # None = off
    else:
        _DATA_SOURCES           = DATA_SOURCES
        _OUTPUT_DIR             = OUTPUT_DIR
        _CONFIG_EXTRACTION_METHOD = CONFIG_EXTRACTION_METHOD
        _COMPARISON_MODE        = COMPARISON_MODE
        _AOA_FILTER             = AOA_FILTER
        _XY_DATA_SOURCE_DIRS    = XY_DATA_SOURCE_DIR if isinstance(XY_DATA_SOURCE_DIR, list) else ([XY_DATA_SOURCE_DIR] if XY_DATA_SOURCE_DIR else [])
        _NUM_ITERATIONS         = NUM_ITERATIONS
        _RUN_CONVERGENCE        = RUN_CONVERGENCE_ANALYSIS
        _CONVERGENCE_MAX_TRIM   = CONVERGENCE_MAX_TRIM
        _CONVERGENCE_NUM_TESTS  = CONVERGENCE_NUM_TESTS
        _GRAPH_MAX_COV          = GRAPH_MAX_COV
        _Q_TIMES_A              = Q_TIMES_A
        _REYNOLDS_NUMBER        = REYNOLDS_NUMBER
        _POSITION_MAP           = POSITION_MAP
        _VALUE_MAPPINGS         = VALUE_MAPPINGS
        _COMPARISON_CONFIGS     = COMPARISON_CONFIGS
        _DATA_MANIPULATIONS     = DATA_MANIPULATIONS
        _REFERENCE_DATA         = REFERENCE_DATA if PLOT_REFERENCE_DATA else None
    
    # Apply presentation mode to cfd_functions before any plots are generated
    import cfd_functions as _cfd_mod
    _cfd_mod.PRESENTATION_MODE = config.get("presentation_mode", PRESENTATION_MODE) if config else PRESENTATION_MODE
    
    print("=" * 100)
    print("CFD DATA PROCESSING - CONSOLIDATED WORKFLOW")
    print("=" * 100)
    
    # ==================== PART 1: LOAD AND PROCESS DATA ====================
    print("\n" + "=" * 100)
    print("PART 1: LOADING AND PROCESSING DATA")
    print("=" * 100)
    
    print(f"\nLoading data from {len(_DATA_SOURCES)} sources...")
    for src in _DATA_SOURCES:
        print(f"  - {src}")

    all_data, validation_report = load_lift_drag_data(_DATA_SOURCES, _CONFIG_EXTRACTION_METHOD, _POSITION_MAP, _VALUE_MAPPINGS, comparison_mode=_COMPARISON_MODE)
    
    # Apply Quick AoA Filter if set
    if _AOA_FILTER:
        print(f"\nApplying AoA Filter: {_AOA_FILTER}")
        initial_count = len(all_data)
        filtered_data = {}
        for key, value in all_data.items():
            # key is (config, aoa_str)
            # We need to extract number from aoa_str (e.g. "AoA_0" -> 0)
            if extract_aoa_number(key[1]) in _AOA_FILTER:
                 filtered_data[key] = value
        
        all_data = filtered_data
        print(f"   -> Filtered down to {len(all_data)} simulations (removed {initial_count - len(all_data)})")
        
        # Apply the exact same filter to REFERENCE_DATA
        if _REFERENCE_DATA:
            for ref in _REFERENCE_DATA:
                filtered_ref_aoa, filtered_ref_cl, filtered_ref_cd = [], [], []
                for a, cl, cd in zip(ref['aoa'], ref['C_L'], ref['C_D']):
                    if a in _AOA_FILTER:
                        filtered_ref_aoa.append(a)
                        filtered_ref_cl.append(cl)
                        filtered_ref_cd.append(cd)
                ref['aoa'] = filtered_ref_aoa
                ref['C_L'] = filtered_ref_cl
                ref['C_D'] = filtered_ref_cd
    
    # Print validation report
    print("\n" + "-" * 100)
    print("DATA VALIDATION REPORT")
    print("-" * 100)
    print(f"[OK] Total folders scanned: {validation_report['total_folders_found']}")
    print(f"[OK] Valid candidates found: {validation_report['valid_folders_scanned']}")
    print(f"[OK] Unique simulations processed: {len(all_data)}")
    print(f"[INFO] Old versions suppressed: {validation_report['versions_suppressed']}")
    print(f"[ERROR] Skipped folders (errors): {validation_report['skipped_folders']}")
    
    if validation_report['issues']:
        print(f"\nIssues found ({len(validation_report['issues'])}):")
        for folder_path, issue in validation_report['issues']:
            # Extract just the parent path for context
            path_obj = Path(folder_path)
            display_name = f".../{path_obj.parent.name}/{path_obj.name}"
            print(f"  [WARN] {display_name}: {issue}")
    print("-" * 100)

    # Apply optional data manipulations (e.g., NG/G ratios)
    derived_entries, manipulation_reports = apply_data_manipulations(all_data, _DATA_MANIPULATIONS, _VALUE_MAPPINGS)
    if derived_entries:
        all_data.update(derived_entries)

    if manipulation_reports:
        print("\n" + "-" * 100)
        print("DATA MANIPULATIONS")
        print("-" * 100)
        for report in manipulation_reports:
            note = report.get('note')
            if note:
                print(f"  [WARN] {report['name']}: {note}")
            else:
                print(f"  - {report['name']}: created {report['created']} derived series (missing pairs: {report['missing_pairs']})")
        print("-" * 100)
    
    print(f"\n[OK] Loaded data for {len(all_data)} configuration-AoA combinations:")
    for (config, aoa), data in sorted(list(all_data.items())[:5]):  # Show first 5
        print(f"  {config} @ {aoa}: {len(data['lift'])} points - {data['turbulence_model']}")
    if len(all_data) > 5:
        print(f"  ... and {len(all_data) - 5} more")
    
    # Export to text files
    processed_data_dir = _OUTPUT_DIR / "processed_data"
    processed_data_dir.mkdir(parents=True, exist_ok=True)
    
    for (config, aoa), data in all_data.items():
        # Create AoA specific folder
        aoa_dir = processed_data_dir / aoa
        aoa_dir.mkdir(exist_ok=True)

        # Export lift
        export_force_data(
            aoa_dir / f"{config}_lift.txt", 
            data['lift'], 
            {**data, 'config': config, 'aoa': aoa}, 
            "Lift"
        )
        
        # Export drag
        export_force_data(
            aoa_dir / f"{config}_drag.txt", 
            data['drag'], 
            {**data, 'config': config, 'aoa': aoa}, 
            "Drag"
        )
    
    print(f"[OK] Exported {len(all_data) * 2} text files to: {processed_data_dir}")
    
    # Create summary statistics text file
    config_name = _OUTPUT_DIR.name
    summary_file = _OUTPUT_DIR / f"SUMMARY_{config_name}.txt"
    
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write("=" * 100 + "\n")
        f.write(f"DATA PROCESSING SUMMARY - Last {_NUM_ITERATIONS} Iterations\n")
        f.write("=" * 100 + "\n\n")
        f.write(f"Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Data Sources:\n")
        for src in _DATA_SOURCES:
            f.write(f"  - {src}\n")
        f.write(f"Output Directory: {_OUTPUT_DIR}\n")
        f.write(f"Extraction Method: {_CONFIG_EXTRACTION_METHOD}\n\n")
        f.write("=" * 100 + "\n\n")
        
        sorted_data = sorted(all_data.items(), key=lambda x: (get_simulation_family_name(x[0][0]), extract_aoa_number(x[0][1])))
        
        for (config, aoa), data in sorted_data:
            lift_last_n = data['lift'][-_NUM_ITERATIONS:] if len(data['lift']) >= _NUM_ITERATIONS else data['lift']
            drag_last_n = data['drag'][-_NUM_ITERATIONS:] if len(data['drag']) >= _NUM_ITERATIONS else data['drag']
            
            lift_mean, lift_cov = compute_statistics(lift_last_n) if lift_last_n else (0, 0)
            drag_mean, drag_cov = compute_statistics(drag_last_n) if drag_last_n else (0, 0)
            
            f.write(f"Configuration: {config}\n")
            f.write(f"  Turbulence Model: {data['turbulence_model']}\n")
            f.write(f"  Geometry: {data['geometry']}\n")
            f.write(f"  Mesh: {data['mesh']}\n")
            f.write(f"  Grid: {data['grid']}\n")
            f.write(f"  Angle of Attack: {aoa}\n")
            f.write(f"  Total Data Points: {len(data['lift'])}\n")
            f.write(f"  Points Used: {len(lift_last_n)}\n")
            f.write("-" * 100 + "\n")
            f.write(f"  Lift Mean:  {lift_mean:12.4f} N\n")
            f.write(f"  Lift COV:   {lift_cov:12.2f} %\n")
            f.write(f"  Drag Mean:  {drag_mean:12.4f} N\n")
            f.write(f"  Drag COV:   {drag_cov:12.2f} %\n")
            f.write("=" * 100 + "\n\n")
    
    print(f"[OK] Summary statistics text file: {summary_file}")
    
    # ==================== PART 2: CONVERGENCE ANALYSIS (OPTIONAL) ====================
    convergence_results = {}
    
    if _RUN_CONVERGENCE:
        print("\n" + "=" * 100)
        print("PART 2: CONVERGENCE ANALYSIS")
        print("=" * 100)
        
        print(f"\nAnalyzing convergence for {len(all_data)} configurations...")
        print(f"Max trim: {_CONVERGENCE_MAX_TRIM * 100}% of data, Tests: {_CONVERGENCE_NUM_TESTS}")
        
        for idx, ((config, aoa), data) in enumerate(all_data.items(), 1):
            if len(data['lift']) < _CONVERGENCE_NUM_TESTS + 10 or len(data['drag']) < _CONVERGENCE_NUM_TESTS + 10:
                print(f"\n  [{idx}/{len(all_data)}] Skipping: {config} - {aoa} (Insufficient data points: {len(data['lift'])})")
                continue

            print(f"\n  [{idx}/{len(all_data)}] Analyzing: {config} - {aoa}")
            
            # Create convergence plots and analyze data
            lift_results, drag_results, plot_path = plot_convergence_analysis(
                config, aoa,
                data['lift'],
                data['drag'],
                _OUTPUT_DIR,
                _CONVERGENCE_MAX_TRIM,
                _CONVERGENCE_NUM_TESTS
            )
            
            convergence_results[(config, aoa)] = {
                'lift': lift_results,
                'drag': drag_results,
                'plot': plot_path
            }
            
            # Print optimization recommendations with confidence info
            print(f"    [OK] Plot saved: {plot_path}")
            
            # Lift recommendation
            if lift_results['trim_recommendation'] is not None:
                print(f"    [OK] Lift - {lift_results['trim_reason']}")
            else:
                print(f"    [WARN]  Lift - No clear recommendation")
                for warning in lift_results.get('warnings', []):
                    print(f"       {warning}")
            
            # Drag recommendation
            if drag_results['trim_recommendation'] is not None:
                print(f"    [OK] Drag - {drag_results['trim_reason']}")
            else:
                print(f"    [WARN]  Drag - No clear recommendation")
                for warning in drag_results.get('warnings', []):
                    print(f"       {warning}")

        # Create Summary Charts
        print("\n  Generating convergence summary plots...")
        plot_convergence_summary(convergence_results, all_data, _OUTPUT_DIR, comparison_mode=_COMPARISON_MODE)
        
        print(f"\n[OK] Convergence analysis complete")
        
        # Export convergence analysis text file
        convergence_dir = _OUTPUT_DIR / "convergence_analysis"
        convergence_dir.mkdir(parents=True, exist_ok=True)
        
        convergence_text_file = convergence_dir / "Convergence_Analysis_Results.txt"
        
        with open(convergence_text_file, 'w') as f:
            f.write("CONVERGENCE ANALYSIS RESULTS\n")
            f.write("=" * 120 + "\n\n")
            
            sorted_convergence = sorted(convergence_results.items(), key=lambda x: (get_simulation_family_name(x[0][0]), extract_aoa_number(x[0][1])))
            
            for (config, aoa), results in sorted_convergence:
                lift_results = results['lift']
                drag_results = results['drag']
                
                lift_min_cov_idx = np.argmin(lift_results['cov'])
                drag_min_cov_idx = np.argmin(drag_results['cov'])
                
                f.write(f"Configuration: {config} | AoA: {aoa}\n")
                f.write(f"Total Iterations: {len(all_data[(config, aoa)]['lift'])}\n")
                f.write("-" * 120 + "\n\n")
                
                f.write("LIFT CONVERGENCE:\n")
                f.write(f"{'Iterations_Removed':<20} {'Iterations_Used':<20} {'Mean':<15} {'StdDev':<15} {'COV(%)':<10}\n")
                f.write("-" * 120 + "\n")
                for i in range(len(lift_results['iterations_removed'])):
                    marker = " <-- MIN COV" if i == lift_min_cov_idx else ""
                    f.write(f"{lift_results['iterations_removed'][i]:<20} "
                           f"{lift_results['iterations_used'][i]:<20} "
                           f"{lift_results['mean'][i]:<15.6f} "
                           f"{lift_results['std_dev'][i]:<15.6f} "
                           f"{lift_results['cov'][i]:<10.2f}{marker}\n")
                
                f.write("\n")
                
                f.write("DRAG CONVERGENCE:\n")
                f.write(f"{'Iterations_Removed':<20} {'Iterations_Used':<20} {'Mean':<15} {'StdDev':<15} {'COV(%)':<10}\n")
                f.write("-" * 120 + "\n")
                for i in range(len(drag_results['iterations_removed'])):
                    marker = " <-- MIN COV" if i == drag_min_cov_idx else ""
                    f.write(f"{drag_results['iterations_removed'][i]:<20} "
                           f"{drag_results['iterations_used'][i]:<20} "
                           f"{drag_results['mean'][i]:<15.6f} "
                           f"{drag_results['std_dev'][i]:<15.6f} "
                           f"{drag_results['cov'][i]:<10.2f}{marker}\n")
                
                f.write("\n" + "=" * 120 + "\n\n")
        
        print(f"[OK] Convergence text file: {convergence_text_file}")
        
        # Export optimized data to text files
        postprocessed_dir = convergence_dir / "optimized_data"
        postprocessed_dir.mkdir(parents=True, exist_ok=True)
        
        for (config, aoa), conv_data in sorted_convergence:
            data = all_data[(config, aoa)]
            
            # Create AoA specific folder
            aoa_dir = postprocessed_dir / aoa
            aoa_dir.mkdir(exist_ok=True)

            lift_min_cov_idx = np.argmin(conv_data['lift']['cov'])
            drag_min_cov_idx = np.argmin(conv_data['drag']['cov'])
            
            optimal_lift_trim = conv_data['lift']['iterations_removed'][lift_min_cov_idx]
            optimal_drag_trim = conv_data['drag']['iterations_removed'][drag_min_cov_idx]
            optimal_trim = max(optimal_lift_trim, optimal_drag_trim)
            
            optimized_lift = data['lift'][optimal_trim:]
            optimized_drag = data['drag'][optimal_trim:]
            
            # Export optimized lift
            export_force_data(
                aoa_dir / f"{config}_lift_optimized.txt",
                optimized_lift,
                {**data, 'config': config, 'aoa': aoa},
                f"Optimized Lift (Trimmed {optimal_trim})"
            )
            
            # Export optimized drag
            export_force_data(
                aoa_dir / f"{config}_drag_optimized.txt",
                optimized_drag,
                {**data, 'config': config, 'aoa': aoa},
                f"Optimized Drag (Trimmed {optimal_trim})"
            )
        
        print(f"[OK] Optimized data files: {postprocessed_dir}")
        print(f"  ({len(convergence_results) * 2} files created)")
    else:
        print("\n[WARN] Skipping convergence analysis (RUN_CONVERGENCE_ANALYSIS = False)")
    
    # ==================== PART 3: EXCEL OUTPUTS ====================
    print("\n" + "=" * 100)
    print("PART 3: GENERATING EXCEL OUTPUTS")
    print("=" * 100)
    
    # Extract config name from OUTPUT_DIR
    config_name = _OUTPUT_DIR.name
    excel_file = _OUTPUT_DIR / f'SUMMARY_{config_name}.xlsx'
    
    # Create workbook
    from openpyxl import Workbook
    wb = Workbook()
    
    # Sheet 1: Data Summary
    print("\n  Creating sheet: Data Summary")
    create_data_summary_sheet(wb, all_data, _NUM_ITERATIONS, convergence_results)
    
    # Sheet 2: Turbulence Comparison
    # Sheet 2: Turbulence / Grid Comparison
    if _COMPARISON_MODE != 'single':
        print(f"  Creating sheet: {_COMPARISON_MODE.capitalize()} Comparison")
        create_turbulence_comparison_sheet(wb, all_data, _NUM_ITERATIONS, convergence_results, comparison_mode=_COMPARISON_MODE)
    else:
        print("  Skipping comparison sheet (Mode: single)")
    
    # Sheet 3: Coefficients
    print("  Creating sheet: Coefficients")
    create_coefficients_sheet(wb, all_data, _NUM_ITERATIONS, convergence_results, _Q_TIMES_A)
    
    version_sheet_created = False
    # Only use COMPARISON_CONFIGS if explicitly in 'version' mode
    # For 'single', 'turbulence', 'grid', we rely on their specific logic.
    should_run_version_comparison = (_COMPARISON_MODE == 'version')
    
    if should_run_version_comparison:
        print("  Creating sheet: Version_Comparison")
        version_sheet_created = create_version_comparison_sheet(
            wb,
            all_data,
            _COMPARISON_CONFIGS,
            _NUM_ITERATIONS,
            convergence_results,
            _Q_TIMES_A,
            comparison_mode=_COMPARISON_MODE
        )
        if not version_sheet_created:
            print("  [WARN] Version comparison sheet skipped (no valid pairs or multi-version families)")

    # Sheet 4: Optimized Statistics (if convergence was run)
    if convergence_results:
        print("  Creating sheet: Optimized_Statistics")
        create_optimized_statistics_sheet(wb, all_data, convergence_results, _NUM_ITERATIONS, _Q_TIMES_A)
    
    # Save workbook
    wb.save(excel_file)
    
    sheet_count = 3
    if convergence_results:
        sheet_count += 1
    if version_sheet_created:
        sheet_count += 1
    print(f"\n[OK] Excel file created with {sheet_count} sheets")
    print(f"[OK] Saved to: {excel_file}")
    
    # ==================== PART 4: COEFFICIENT GRAPHS ====================
    print("\n" + "=" * 100)
    print("PART 4: GENERATING COEFFICIENT GRAPHS")
    print("=" * 100)
    
    # Calculate coefficients
    print("\nCalculating coefficients...")
    coefficient_data = {}
    
    for (config, aoa), data in all_data.items():
        # Get optimized or fixed iteration data
        if convergence_results and (config, aoa) in convergence_results:
            conv = convergence_results[(config, aoa)]
            lift_min_cov_idx = np.argmin(conv['lift']['cov'])
            drag_min_cov_idx = np.argmin(conv['drag']['cov'])
            
            optimal_lift_trim = conv['lift']['iterations_removed'][lift_min_cov_idx]
            optimal_drag_trim = conv['drag']['iterations_removed'][drag_min_cov_idx]
            optimal_trim = max(optimal_lift_trim, optimal_drag_trim)
            
            lift_values = data['lift'][optimal_trim:]
            drag_values = data['drag'][optimal_trim:]
        else:
            lift_values = data['lift'][-_NUM_ITERATIONS:] if len(data['lift']) >= _NUM_ITERATIONS else data['lift']
            drag_values = data['drag'][-_NUM_ITERATIONS:] if len(data['drag']) >= _NUM_ITERATIONS else data['drag']
        
        lift_mean = np.mean(lift_values) if lift_values else 0
        drag_mean = np.mean(drag_values) if drag_values else 0
        lift_std = np.std(lift_values) if lift_values else 0
        drag_std = np.std(drag_values) if drag_values else 0
        
        C_L = lift_mean / _Q_TIMES_A if _Q_TIMES_A != 0 else 0
        C_D = drag_mean / _Q_TIMES_A if _Q_TIMES_A != 0 else 0
        C_L_std = lift_std / _Q_TIMES_A if _Q_TIMES_A != 0 else 0
        C_D_std = drag_std / _Q_TIMES_A if _Q_TIMES_A != 0 else 0
        
        coefficient_data[(config, aoa)] = {
            'turbulence_model': data['turbulence_model'],
            'aoa_degrees': extract_aoa_number(aoa),
            'grid': data.get('grid', 'Unknown'), # Pass grid status for grid graphs
            'C_L': C_L,
            'C_D': C_D,
            'C_L_std': C_L_std,
            'C_D_std': C_D_std,
        }
    
    print(f"[OK] Coefficients calculated for {len(coefficient_data)} configurations")
    
    # Add Reference Comparison sheet to Excel (needs coefficient_data from above)
    if _REFERENCE_DATA:
        from openpyxl import load_workbook
        wb_ref = load_workbook(excel_file)
        print("  Creating sheet: Reference_Comparison")
        create_reference_comparison_sheet(wb_ref, coefficient_data, _REFERENCE_DATA)
        wb_ref.save(excel_file)
        print(f"  [OK] Reference comparison sheet added to {excel_file.name}")
    
    # Create graphs
    print("\nGenerating graphs...")
    create_coefficient_graphs(all_data, coefficient_data, _OUTPUT_DIR, _POSITION_MAP, _VALUE_MAPPINGS, 
                              comparison_mode=_COMPARISON_MODE, max_cov_threshold=_GRAPH_MAX_COV,
                              reference_data=_REFERENCE_DATA)

    # Create Grid Graphs (Efficiency Ratio)
    print("Generating grid graphs...")
    create_grid_graphs(coefficient_data, _OUTPUT_DIR, comparison_mode=_COMPARISON_MODE, max_cov_threshold=_GRAPH_MAX_COV, value_mappings=_VALUE_MAPPINGS)
    
    graphs_dir = _OUTPUT_DIR / "coefficient_graphs"
    print(f"\n[OK] Graphs saved to: {graphs_dir}")
    print("[OK] Organization: coefficient_graphs / turbulence_model / config /")
    print("[OK] Each config contains: C_L_vs_AoA.png, C_D_vs_AoA.png, C_L_C_D_Combined.png")

    # ==================== PRESENTATION MODE PASS ====================
    # Always generate a second set of graphs with presentation-mode styling
    _pres_dir = _OUTPUT_DIR / "Pres"
    _saved_pres = _cfd_mod.PRESENTATION_MODE
    if not _saved_pres:
        print("\n" + "-" * 100)
        print("Generating presentation-mode graphs...")
        print("-" * 100)
        _cfd_mod.PRESENTATION_MODE = True
        _cfd_mod.set_plot_style()  # Apply presentation-mode rcParams

        # Re-generate coefficient graphs
        create_coefficient_graphs(all_data, coefficient_data, _pres_dir, _POSITION_MAP, _VALUE_MAPPINGS,
                                  comparison_mode=_COMPARISON_MODE, max_cov_threshold=_GRAPH_MAX_COV,
                                  reference_data=_REFERENCE_DATA)
        create_grid_graphs(coefficient_data, _pres_dir, comparison_mode=_COMPARISON_MODE, max_cov_threshold=_GRAPH_MAX_COV, value_mappings=_VALUE_MAPPINGS)

        # Re-generate convergence summary if applicable
        if convergence_results:
            plot_convergence_summary(convergence_results, all_data, _pres_dir, comparison_mode=_COMPARISON_MODE)

        # Restore original mode
        _cfd_mod.PRESENTATION_MODE = _saved_pres
        _cfd_mod.set_plot_style()  # Restore normal rcParams
        print(f"[OK] Presentation graphs saved to: {_pres_dir}")
    
    # ==================== PART 5: GENERATING EXPORTED PLOTS (Cp, Y+) ====================
    if _COMPARISON_MODE != 'single' and not _XY_DATA_SOURCE_DIRS:
        print("\n[WARN] Skipping Extra Plots (Cp, Y+, Cf) — no XY_DATA_SOURCE_DIR configured for comparison mode.")
    elif _COMPARISON_MODE != 'single' and _XY_DATA_SOURCE_DIRS:
        # --- COMPARISON MODE: overlay XY data from multiple directories ---
        print("\n" + "=" * 100)
        print("PART 5: GENERATING COMPARISON XY PLOTS (Cp, Y+, Cf)")
        print("=" * 100)

        import re as _re
        from collections import defaultdict as _ddict

        # Collect all .xy files from each source, tagged with a label
        # Label = parent folder name (e.g. "5.6.1.1.NG")
        tagged_files = []  # list of (label, filepath)
        for xy_dir in _XY_DATA_SOURCE_DIRS:
            if not xy_dir.exists():
                print(f"  [WARN] XY source dir does not exist: {xy_dir}")
                continue
            label = xy_dir.parent.name  # e.g. "5.6.1.1.NG" from ".../5.6.1.1.NG/y_plus_pressure_data"
            for f in xy_dir.glob("*.xy"):
                tagged_files.append((label, f))
            print(f"  Scanned {xy_dir.name} ({label}): {len(list(xy_dir.glob('*.xy')))} .xy files")

        if not tagged_files:
            print("  [WARN] No .xy files found in any XY_DATA_SOURCE_DIR.")
        else:
            # Classify each file by type (cp, yplus, cf) and extract AoA
            # Group: type_groups[type][aoa_str] = [(label, filepath), ...]
            type_groups = _ddict(lambda: _ddict(list))

            for label, f in tagged_files:
                fname_lower = f.name.lower()

                # Determine type
                if 'cp' in fname_lower or 'pressure' in fname_lower:
                    ftype = 'Cp'
                elif 'yplus' in fname_lower or 'y-plus' in fname_lower or 'y_plus' in fname_lower:
                    ftype = 'Y+'
                elif 'skin' in fname_lower or 'friction' in fname_lower:
                    ftype = 'Cf'
                else:
                    continue  # unknown type, skip

                # Extract AoA from filename: last numeric segment(s)
                # e.g. "5.6.1.1.NG.10.Cp.xy" -> AoA=10
                # e.g. "5.6.1.1.NG.5.5.yplus.xy" -> AoA=5.5
                parts = f.stem.split('.')
                aoa_str = None
                for i in range(len(parts) - 1, -1, -1):
                    if parts[i].lstrip('-').isdigit():
                        # Check for decimal AoA
                        if i >= 2 and parts[i-1].lstrip('-').isdigit() and not parts[i-2].lstrip('-').isdigit():
                            aoa_str = f"{parts[i-1]}.{parts[i]}"
                        else:
                            aoa_str = parts[i]
                        break

                if aoa_str is None:
                    continue

                type_groups[ftype][aoa_str].append((label, f))

            # Now generate comparison plots
            plot_counts = {"Cp": 0, "Y+": 0, "Cf": 0}
            base_plot_dir = _OUTPUT_DIR / "Extra_Plots" / "Comparison"

            type_config = {
                'Cp': {'ylabel': 'Pressure Coefficient ($C_p$)', 'title_prefix': 'Pressure Coefficient', 'subdir': 'pressure_coefficient', 'invert_y': True},
                'Y+': {'ylabel': 'Wall Y+', 'title_prefix': 'Wall Y+ Distribution', 'subdir': 'y_plus', 'invert_y': False},
                'Cf': {'ylabel': 'Skin Friction Coefficient ($C_f$)', 'title_prefix': 'Skin Friction Coefficient', 'subdir': 'skin_friction', 'invert_y': False},
            }

            for ftype, aoa_dict in sorted(type_groups.items()):
                cfg = type_config[ftype]
                for aoa_str, entries in sorted(aoa_dict.items(), key=lambda x: float(x[0])):
                    if len(entries) < 2:
                        continue  # need at least 2 sources to compare

                    # Apply AoA filter if set
                    try:
                        aoa_num = float(aoa_str)
                    except ValueError:
                        continue
                    if _AOA_FILTER and aoa_num not in _AOA_FILTER:
                        continue

                    series_list = []
                    for label, filepath in entries:
                        df = read_fluent_xy(filepath)
                        if not df.empty:
                            series_list.append({'df': df, 'label': label})

                    if len(series_list) < 2:
                        continue

                    out_dir = base_plot_dir / cfg['subdir']
                    out_dir.mkdir(parents=True, exist_ok=True)
                    output_path = out_dir / f"{ftype}_AoA_{aoa_str}.png"

                    plot_xy_comparison(
                        series_list,
                        title=f"{cfg['title_prefix']} Comparison - AoA {aoa_str}°",
                        xlabel="X Position (m)",
                        ylabel=cfg['ylabel'],
                        output_path=output_path,
                        invert_y=cfg['invert_y']
                    )
                    plot_counts[ftype] += 1

            print(f"\n[OK] Generated {plot_counts['Cp']} Cp, {plot_counts['Y+']} Y+, {plot_counts['Cf']} Cf comparison plots")
            print(f"[OK] Saved to: {base_plot_dir}")
    else:
        print("\n" + "=" * 100)
        print("PART 5: GENERATING EXPORTED PLOTS (Cp, Y+, Residuals)")
        print("=" * 100)
    
        plot_counts = {"Cp": 0, "Y+": 0, "Cf": 0, "Residuals": 0}
    
        for (cfg, aoa), data in all_data.items():
            source_dir = data.get('source_dir')
            if not source_dir:
                continue
            
            # Look for data files in y_plus_pressure_data (per-AoA or parent level)
            xy_dir = source_dir / "y_plus_pressure_data"
            if not xy_dir.exists():
                xy_dir = source_dir.parent / "y_plus_pressure_data"
            if not xy_dir.exists():
                print(f"DEBUG: xy_dir {xy_dir} does not exist for {cfg} @ {aoa}")
                continue
            
            all_xy_files = list(xy_dir.glob("*.xy")) + list(xy_dir.glob("*.csv"))
            if not all_xy_files:
                print(f"DEBUG: No .xy or .csv files found in {xy_dir}")
                continue

            # When using parent-level folder, filter to only this AoA's files
            # Files follow naming: CONFIG.AOA.type.ext (e.g. 4.3.1.3.NG.5.Cp.xy)
            # The 'aoa' variable here is usually a string like "AoA_10" or "AoA_5.5"
            aoa_str = str(aoa).replace("AoA_", "") if str(aoa).startswith("AoA_") else str(aoa)
            
            # Since the naming is {cfg}.{aoa}.{type}.xy, let's use exact match instead of generic '.'
            # Check specifically for f"{cfg}.{aoa_str}." at the start of the filename
            prefix_marker = f"{cfg}.{aoa_str}."
            all_xy_files = [f for f in all_xy_files if f.name.startswith(prefix_marker)]
            
            if not all_xy_files:
                continue
            
            print(f"\n  {cfg} @ {aoa_str}: found {len(all_xy_files)} data files")
            
            # Classify files by type based on filename keywords
            cp_files = [f for f in all_xy_files if 'cp' in f.name.lower() or 'pressure' in f.name.lower()]
            yplus_files = [f for f in all_xy_files if 'yplus' in f.name.lower() or 'y-plus' in f.name.lower() or 'y_plus' in f.name.lower()]
            cf_files = [f for f in all_xy_files if 'skin' in f.name.lower() or 'friction' in f.name.lower()]
            
            # Setup output directories
            base_plot_dir = _OUTPUT_DIR / "Extra_Plots"
            cp_out_dir = base_plot_dir / "pressure_coefficient" / data['turbulence_model'] / cfg
            yplus_out_dir = base_plot_dir / "y_plus" / data['turbulence_model'] / cfg
            cf_out_dir = base_plot_dir / "skin_friction" / data['turbulence_model'] / cfg
            
            # Process Cp Files
            for f in cp_files:
                cp_out_dir.mkdir(parents=True, exist_ok=True)
                df = read_fluent_xy(f)
                if not df.empty:
                    output_path = cp_out_dir / f"Cp_{cfg}_{aoa}.png"
                    plot_xy_series(
                        df, 
                        title=f"Pressure Coefficient - {cfg} - {aoa}",
                        xlabel="X Position (m)",
                        ylabel="Pressure Coefficient ($C_p$)",
                        output_path=output_path,
                        invert_y=True
                    )
                    plot_counts["Cp"] += 1
                    
            # Process Y+ Files
            for f in yplus_files:
                yplus_out_dir.mkdir(parents=True, exist_ok=True)
                df = read_fluent_xy(f)
                if not df.empty:
                    output_path = yplus_out_dir / f"Yplus_{cfg}_{aoa}.png"
                    plot_xy_series(
                        df,
                        title=f"Wall Y+ Distribution - {cfg} - {aoa}",
                        xlabel="X Position (m)",
                        ylabel="Wall Y+",
                        output_path=output_path,
                        invert_y=False
                    )
                    plot_counts["Y+"] += 1
                    
            # Process Cf Files
            for f in cf_files:
                cf_out_dir.mkdir(parents=True, exist_ok=True)
                df = read_fluent_xy(f)
                if not df.empty:
                    output_path = cf_out_dir / f"Cf_{cfg}_{aoa}.png"
                    plot_xy_series(
                        df,
                        title=f"Skin Friction Coefficient - {cfg} - {aoa}",
                        xlabel="X Position (m)",
                        ylabel="Skin Friction Coefficient ($C_f$)",
                        output_path=output_path,
                        invert_y=False
                    )
                    plot_counts["Cf"] += 1

            # Process Residual Files
            residual_files = [f for f in all_xy_files if 'residual' in f.name.lower()]
            for f in residual_files:
                res_out_dir = base_plot_dir / "residuals" / data['turbulence_model'] / cfg
                res_out_dir.mkdir(parents=True, exist_ok=True)
                res_data = read_fluent_residuals(f)
                
                if res_data:
                    # Read the *processed* lift data exact length to sync residual plots
                    max_iters = None
                    processed_lift_file = _OUTPUT_DIR / "processed_data" / str(aoa) / f"{cfg}_lift.txt"
                    if processed_lift_file.exists():
                        try:
                            # Count non-header lines to get exact iteration count of clean data
                            with open(processed_lift_file, 'r') as pf:
                                max_iters = sum(1 for line in pf if line.strip() and not line.startswith('#'))
                        except Exception as e:
                            print(f"DEBUG: Failed to read processed lift file for length: {e}")
                    
                    if max_iters == 0: max_iters = None
                    
                    output_path = res_out_dir / f"Residuals_{cfg}_{aoa}.png"
                    plot_residuals(
                        res_data,
                        title=f"Residual Convergence - {cfg} - {aoa}",
                        output_path=output_path,
                        max_iters=max_iters
                    )
                    plot_counts["Residuals"] += 1

        print(f"\n[OK] Generated {plot_counts['Cp']} Cp, {plot_counts['Y+']} Y+, {plot_counts['Cf']} Cf, {plot_counts['Residuals']} Residual plots")
        print(f"[OK] Saved to: {_OUTPUT_DIR / 'Extra_Plots'}")    

    # ==================== PART 6: PATHLINE DATA ====================
    _PROCESS_PATHLINES = _config_dict.get("process_pathlines", PROCESS_PATHLINES) if _config_dict else PROCESS_PATHLINES
    _PATHLINE_DIRS = _config_dict.get("pathline_data_source_dirs", PATHLINE_DATA_SOURCE_DIRS) if _config_dict else PATHLINE_DATA_SOURCE_DIRS
    if not _PATHLINE_DIRS:
        _PATHLINE_DIRS = []

    if _PROCESS_PATHLINES and _PATHLINE_DIRS:
        print("\n" + "=" * 100)
        print("PART 6: PROCESSING PATHLINE DATA")
        print("=" * 100)

        import re as _re
        from collections import defaultdict as _ddict

        # Collect all .fvp files from each source, tagged with a label
        tagged_fvp = []  # list of (label, filepath)
        for pl_dir in _PATHLINE_DIRS:
            pl_path = Path(pl_dir)
            if not pl_path.exists():
                print(f"  [WARN] Pathline source dir does not exist: {pl_path}")
                continue
            label = pl_path.parent.name  # e.g. "5.6.1.1.G"
            fvp_files = list(pl_path.glob("*.fvp"))
            for f in fvp_files:
                tagged_fvp.append((label, f))
            print(f"  Scanned {pl_path.name} ({label}): {len(fvp_files)} .fvp files")

        if not tagged_fvp:
            print("  [WARN] No .fvp pathline files found in any PATHLINE_DATA_SOURCE_DIRS.")
        else:
            # Group by AoA for comparison and individual plots
            # Key structure: aoa_groups[aoa_str] = [(label, filepath), ...]
            aoa_groups = _ddict(list)

            for label, f in tagged_fvp:
                # Extract AoA from filename:
                # e.g. "5.6.1.1.G.10.pathline.velocity_magnitude.fvp" -> AoA = 10
                parts = f.stem.split('.')
                aoa_str = None
                for i in range(len(parts) - 1, -1, -1):
                    # Walk backwards looking for the AoA number
                    # Skip known non-AoA tokens
                    if parts[i] in ('pathline', 'fvp') or parts[i].startswith('var_'):
                        continue
                    if any(c.isalpha() for c in parts[i].replace('-', '')):
                        continue
                    if parts[i].lstrip('-').replace('.', '', 1).isdigit():
                        # Check for decimal AoA
                        if i >= 2 and parts[i-1].lstrip('-').isdigit() and not parts[i-2].lstrip('-').isdigit():
                            aoa_str = f"{parts[i-1]}.{parts[i]}"
                        else:
                            aoa_str = parts[i]
                        break

                if aoa_str is None:
                    print(f"  [WARN] Could not extract AoA from: {f.name}")
                    continue

                # Apply AoA filter
                try:
                    aoa_num = float(aoa_str)
                except ValueError:
                    continue
                if _AOA_FILTER and aoa_num not in _AOA_FILTER:
                    continue

                aoa_groups[aoa_str].append((label, f))

            # Generate plots
            pl_plot_count = 0
            base_pl_dir = _OUTPUT_DIR / "Extra_Plots"

            if _COMPARISON_MODE != 'single' and len(_PATHLINE_DIRS) >= 2:
                # Comparison mode: overlay multiple configs per AoA
                comp_dir = base_pl_dir / "Comparison" / "pathlines"
                comp_dir.mkdir(parents=True, exist_ok=True)

                for aoa_str, entries in sorted(aoa_groups.items(), key=lambda x: float(x[0])):
                    if len(entries) < 2:
                        continue

                    datasets = []
                    for label, filepath in entries:
                        data = read_fluent_fvp(filepath)
                        if data:
                            datasets.append({'data': data, 'label': label})

                    if len(datasets) >= 2:
                        output_path = comp_dir / f"Pathlines_AoA_{aoa_str}.png"
                        plot_pathline_comparison(
                            datasets,
                            title=f"Pathline Comparison - AoA {aoa_str}°",
                            output_path=output_path
                        )
                        pl_plot_count += 1

            # Individual plots per (label, AoA)
            for aoa_str, entries in sorted(aoa_groups.items(), key=lambda x: float(x[0])):
                for label, filepath in entries:
                    data = read_fluent_fvp(filepath)
                    if not data:
                        continue

                    # Determine color_by from first particle's columns
                    first_df = next(iter(data.values()))
                    scalar_cols = [c for c in first_df.columns if c not in ('x', 'y', 'z')]
                    color_col = scalar_cols[0] if scalar_cols else None

                    out_dir = base_pl_dir / "pathlines" / label
                    out_dir.mkdir(parents=True, exist_ok=True)
                    output_path = out_dir / f"Pathlines_{label}_AoA_{aoa_str}.png"

                    plot_pathlines(
                        data,
                        title=f"Pathlines - {label} - AoA {aoa_str}°",
                        output_path=output_path,
                        color_by=color_col
                    )
                    pl_plot_count += 1

            print(f"\n[OK] Generated {pl_plot_count} pathline plots")
            print(f"[OK] Saved to: {base_pl_dir}")
    elif _PROCESS_PATHLINES:
        print("\n[WARN] Skipping pathline processing (no PATHLINE_DATA_SOURCE_DIRS configured)")

    # ==================== FINAL SUMMARY ====================
    print("\n" + "=" * 100)
    print("WORKFLOW COMPLETE!")
    print("=" * 100)
    print(f"\nOutputs saved to: {_OUTPUT_DIR}")
    print(f"  |-- processed_data.pkl")
    if convergence_results:
        print(f"  |-- convergence_results.pkl")
    print(f"  |-- SUMMARY_Statistics.xlsx ({sheet_count} sheets)")
    print(f"  |-- processed_data/ ({len(all_data) * 2} text files)")
    print(f"  \-- coefficient_graphs/ (organized by turbulence model)")
    print("\n[OK] All processing complete!")


if __name__ == "__main__":
    main()
