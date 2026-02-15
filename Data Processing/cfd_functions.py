"""
CFD Data Processing Functions
Contains all reusable functions for processing CFD simulation data.
"""

import os
from pathlib import Path
import glob
import numpy as np
import pandas as pd
from collections import defaultdict
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ==================== PLOTTING CONSTANTS & STYLING ====================

def set_plot_style():
    """Configure professional academic plot styles for Matplotlib."""
    plt.rcParams.update({
        # High-res output
        'figure.dpi': 300,
        'savefig.dpi': 300,
        
        # Typography
        'font.family': 'sans-serif',
        'font.sans-serif': ['Arial', 'Helvetica', 'DejaVu Sans'],
        'font.size': 11,
        'axes.titlesize': 14,
        'axes.titleweight': 'bold',
        'axes.labelsize': 12,
        'legend.fontsize': 10,
        'xtick.labelsize': 10,
        'ytick.labelsize': 10,
        
        # Grid & Ticks
        'axes.grid': True,
        'axes.grid.which': 'both',
        'grid.alpha': 0.3,
        'grid.linestyle': '--',
        'grid.linewidth': 0.5,
        'axes.axisbelow': True,
        'xtick.direction': 'in',
        'ytick.direction': 'in',
        'xtick.top': True,
        'ytick.right': True,
        
        # Spine aesthetics
        'axes.linewidth': 1.0,
        'axes.spines.top': True,
        'axes.spines.right': True,
        
        # Markers & Lines
        'lines.linewidth': 1.8,
        'lines.markersize': 7,
        'errorbar.capsize': 3,
    })

# Academic color palette
ACADEMIC_COLORS = [
    '#004c6d', # Dark blue
    '#c33c54', # Muted red
    '#254441', # Deep teal
    '#ef7a1a', # Vibrant orange
    '#432371', # Deep purple
    '#7b9e89', # Sage green
    '#d4a373', # Sandy brown
    '#588157', # Forest green
    '#ffb703', # Gold
    '#fb8500', # Bright orange
]

ACADEMIC_MARKERS = ['o', 's', '^', 'D', 'v', 'p', 'P', '*', 'h', 'X']


# ==================== DATA VALIDATION FUNCTIONS ====================

def read_fluent_xy(filepath):
    """
    Parses an ANSYS Fluent .xy file.
    
    Format:
    (title "...")
    (labels "..." "...")
    ((xy/key/label "surface_name")
    x1  y1
    x2  y2
    ...
    )
    ((xy/key/label "another_surface")
    ...
    )
    
    Returns:
        pd.DataFrame with columns ['x', 'y'] containing all points combined.
    """
    data_points = []
    
    try:
        with open(filepath, 'r') as f:
            in_data_block = False
            for line in f:
                line = line.strip()
                if not line:
                    continue
                
                # Check for start of data block
                if line.startswith('((xy/key/label'):
                    in_data_block = True
                    continue
                
                # Check for end of data block
                if line == ')' and in_data_block:
                    in_data_block = False
                    continue
                
                # Check for headers or metadata to ignore
                if line.startswith('('):
                    continue
                
                # Parse data points if inside a block
                if in_data_block:
                    parts = line.split()
                    if len(parts) >= 2:
                        try:
                            x = float(parts[0])
                            y = float(parts[1])
                            data_points.append({'x': x, 'y': y})
                        except ValueError:
                            continue
                            
        if not data_points:
            return pd.DataFrame(columns=['x', 'y'])
            
        return pd.DataFrame(data_points)
        
    except Exception as e:
        print(f"Error reading XY file {filepath}: {e}")
        return pd.DataFrame(columns=['x', 'y'])


def plot_xy_series(df, title, xlabel, ylabel, output_path, invert_y=False):
    """
    Generates and saves a single XY plot for Cp or Y+ data.
    """
    if df.empty:
        print(f"  Warning: Empty dataframe for {title}, skipping plot.")
        return

    plt.figure(figsize=(10, 6))
    
    # Scatter plot for data points (handles disjoint surfaces better than line plot)
    plt.plot(df['x'], df['y'], 'o', markersize=2, alpha=0.6, color='black', label='Data Points')
    
    plt.title(title, fontsize=14, fontweight='bold')
    plt.xlabel(xlabel, fontsize=12)
    plt.ylabel(ylabel, fontsize=12)
    plt.grid(True, alpha=0.3)
    
    if invert_y:
        plt.gca().invert_yaxis()
        
    plt.tight_layout()
    try:
        plt.savefig(output_path, dpi=300)
    except Exception as e:
        print(f"  Error saving plot to {output_path}: {e}")
    plt.close()



# ==================== PLOTTING CONSTANTS & STYLING ====================

def set_plot_style():
    """Configure professional academic plot styles for Matplotlib."""
    plt.rcParams.update({
        # High-res output
        'figure.dpi': 300,
        'savefig.dpi': 300,
        
        # Typography
        'font.family': 'sans-serif',
        'font.sans-serif': ['Arial', 'Helvetica', 'DejaVu Sans'],
        'font.size': 11,
        'axes.titlesize': 14,
        'axes.titleweight': 'bold',
        'axes.labelsize': 12,
        'legend.fontsize': 10,
        'xtick.labelsize': 10,
        'ytick.labelsize': 10,
        
        # Grid & Ticks
        'axes.grid': True,
        'axes.grid.which': 'both',
        'grid.alpha': 0.3,
        'grid.linestyle': '--',
        'grid.linewidth': 0.5,
        'axes.axisbelow': True,
        'xtick.direction': 'in',
        'ytick.direction': 'in',
        'xtick.top': True,
        'ytick.right': True,
        
        # Spine aesthetics
        'axes.linewidth': 1.0,
        'axes.spines.top': True,
        'axes.spines.right': True,
        
        # Markers & Lines
        'lines.linewidth': 1.8,
        'lines.markersize': 7,
        'errorbar.capsize': 3,
    })

# Academic color palette
ACADEMIC_COLORS = [
    '#004c6d', # Dark blue
    '#c33c54', # Muted red
    '#254441', # Deep teal
    '#ef7a1a', # Vibrant orange
    '#432371', # Deep purple
    '#7b9e89', # Sage green
    '#d4a373', # Sandy brown
    '#588157', # Forest green
    '#ffb703', # Gold
    '#fb8500', # Bright orange
]

ACADEMIC_MARKERS = ['o', 's', '^', 'D', 'v', 'p', 'P', '*', 'h', 'X']


# ==================== DATA VALIDATION FUNCTIONS ====================

def validate_aoa_folder(dirpath, filenames):
    """
    Validate that an AoA folder has required files with no duplicates or issues.
    
    Args:
        dirpath: Path to the AoA folder
        filenames: List of filenames in the folder
    
    Returns:
        Tuple of (is_valid, lift_file, drag_file, case_file, error_msg)
        - is_valid (bool): True if folder passes all checks
        - lift_file (str): Validated lift filename (None if invalid)
        - drag_file (str): Validated drag filename (None if invalid)
        - case_file (str): Validated case filename (None if invalid)
        - error_msg (str): Description of any issues (empty if valid)
    """
    errors = []
    
    # Helper to pick best file
    def pick_best_file(file_list, file_type):
        if not file_list:
            return None, f"No {file_type} file found"
        
        if len(file_list) == 1:
            return file_list[0], None
            
        # Multiple files: Sort by Last Modified (Newest First), then Size
        def get_sort_key(fname):
            full_path = dirpath / fname
            try:
                # Primary: Modification Time, Secondary: File Size
                return (full_path.stat().st_mtime, full_path.stat().st_size)
            except OSError:
                return (0, 0)
                
        sorted_files = sorted(file_list, key=get_sort_key, reverse=True)
        winner = sorted_files[0]
        
        # Log warning for multiple files
        print(f"    [WARNING] [Auto-Select] Multiple {file_type} files in {dirpath.name}. Using newest: {winner}")
        return winner, None

    # Find lift files
    lift_candidates = [f for f in filenames if 'lift_force' in f and f.endswith('.txt')]
    lift_file, lift_err = pick_best_file(lift_candidates, "lift_force_*.txt")
    if lift_err: errors.append(lift_err)
    
    # Find drag files
    drag_candidates = [f for f in filenames if 'drag_force' in f and f.endswith('.txt')]
    drag_file, drag_err = pick_best_file(drag_candidates, "drag_force_*.txt")
    if drag_err: errors.append(drag_err)
    
    # Find case file
    case_files = [f for f in filenames if f.endswith('.cas') or f.endswith('.cas.h5')]
    if len(case_files) == 0:
        errors.append("No case file (.cas or .cas.h5) found")
    elif len(case_files) > 1:
        errors.append(f"Multiple case files found: {case_files}")
    
    # Compile results
    is_valid = len(errors) == 0
    case_file = case_files[0] if len(case_files) >= 1 else None
    error_msg = " | ".join(errors) if errors else ""
    
    return is_valid, lift_file, drag_file, case_file, error_msg


# ==================== DATA LOADING FUNCTIONS ====================

def find_simulation_folders(root_dir):
    """Generator that yields paths to folders containing simulation data (case + force files)."""
    for dirpath, _, filenames in os.walk(root_dir):
        # Look for case files to identify simulation folders
        has_case_file = any(f.endswith('.cas') or f.endswith('.cas.h5') for f in filenames)
        if has_case_file:
            yield Path(dirpath), filenames

# Legacy alias for backward compatibility
def find_aoa_folders(root_dir):
    """Deprecated: Use find_simulation_folders instead."""
    return find_simulation_folders(root_dir)

def parse_configuration(dirpath, case_file, config_extraction_method, position_map, value_mappings):
    """
    Extracts configuration metadata from case filename.
    
    Supports two naming conventions:
    1. _AoA_ format: 4.3.1.2_AoA_10.cas.h5 → config=4.3.1.2, AoA=10
    2. Dot format: 4.3.1.3.NG.5.cas.h5 → config=4.3.1.3.NG, AoA=5
    """
    import re
    
    if not case_file:
        return None, "No case file provided"
    
    # Strip file extension
    base_name = case_file.replace('.cas.h5', '').replace('.cas', '')
    
    # Try to extract config and AoA
    config = None
    aoa_number = None
    
    # Format 1: Contains _AoA_ (e.g., 4.3.1.2_AoA_10)
    if '_AoA_' in base_name:
        parts = base_name.split('_AoA_')
        if len(parts) == 2:
            config = parts[0]
            aoa_number = parts[1]
    else:
        # Format 2: AoA is last numeric segment after dots (e.g., 4.3.1.3.NG.5)
        # Split by dots and find the last purely numeric part
        parts = base_name.split('.')
        
        # Find last numeric part (that's the AoA)
        for i in range(len(parts) - 1, -1, -1):
            if parts[i].lstrip('-').isdigit():
                aoa_number = parts[i]
                config = '.'.join(parts[:i])
                break
        
        # If no numeric AoA found at end, this might be a different format
        if aoa_number is None:
            return None, f"Could not extract AoA from filename: {base_name}"
    
    if not config or aoa_number is None:
        return None, f"Could not parse config/AoA from filename: {base_name}"
    
    # Create standardized AoA string
    aoa = f"AoA_{aoa_number}"
    
    # Parse configuration parts (config string without AoA)
    config_parts = config.split('.')
    
    def safe_get(index, cast_type=int):
        if index is not None and 0 <= index < len(config_parts):
            try:
                return cast_type(config_parts[index])
            except ValueError:
                return config_parts[index]
        return None

    # Extract fields using position map
    geometry_num = safe_get(position_map['geometry'])
    mesh_num = safe_get(position_map['mesh'])
    turbulence_num = safe_get(position_map['turbulence'])
    version_num = safe_get(position_map['version'])
    
    # Handle optional grid field (may be None in 4-part schema)
    grid_index = position_map.get('grid')
    grid_code = safe_get(grid_index, str) if grid_index is not None else None
    
    # Fallback: Check directory name if Grid is not explicit in filename
    if not grid_code or grid_code == 'None':
         folder_name = dirpath.name
         if '.G' in folder_name or '_G' in folder_name or 'With Grid' in folder_name:
             grid_code = 'G'
         elif '.NG' in folder_name or '_NG' in folder_name or 'No Grid' in folder_name:
             grid_code = 'NG'

    # Map to descriptive names
    metadata = {
        'config': config,
        'aoa': aoa,
        'aoa_number': float(aoa_number),
        'version_sort_key': version_num if isinstance(version_num, (int, float)) else 0,
        'geometry': value_mappings.get('geometry', {}).get(geometry_num, f"Geometry_{geometry_num}") if geometry_num else "N/A",
        'mesh': value_mappings.get('mesh', {}).get(mesh_num, f"Mesh_{mesh_num}") if mesh_num else "N/A",
        'turbulence_model': value_mappings.get('turbulence', {}).get(turbulence_num, f"Turbulence_{turbulence_num}") if turbulence_num else "Unknown",
        'version': value_mappings.get('version', {}).get(version_num, f"Version_{version_num}") if version_num else "N/A",
        'grid': value_mappings.get('grid', {}).get(grid_code, f"Grid_{grid_code}") if grid_code else "N/A"
    }
    
    return metadata, None

def load_and_correct_forces(lift_file, drag_file, aoa_degrees):
    """Loads force data and applies AoA correction using vectorization."""
    lift_data = np.array(_read_force_file(lift_file))
    drag_data = np.array(_read_force_file(drag_file))
    
    # Truncate to same length
    min_length = min(len(lift_data), len(drag_data))
    lift_data = lift_data[:min_length]
    drag_data = drag_data[:min_length]
    
    # Apply AoA correction
    aoa_radians = np.radians(aoa_degrees)
    cos_theta = np.cos(aoa_radians)
    sin_theta = np.sin(aoa_radians)
    
    true_lift = lift_data * cos_theta - drag_data * sin_theta
    true_drag = lift_data * sin_theta + drag_data * cos_theta
    
    return true_lift.tolist(), true_drag.tolist()

def load_lift_drag_data(root_dirs, config_extraction_method, position_map, value_mappings, comparison_mode='default'):
    """
    Load lift and drag force data from multiple source directories.
    Implements 'Highest Version Wins' strategy for duplicates, unless comparing versions.
    """
    if not isinstance(root_dirs, list):
        root_dirs = [root_dirs]
        
    validation_report = {
        'total_folders_found': 0,
        'valid_folders_scanned': 0,
        'skipped_folders': 0,
        'versions_suppressed': 0,
        'issues': []
    }
    
    # --- PHASE 1: SCOUTING & IDENTIFICATION ---
    # We group candidates by a unique "Identity" (Geometry + Mesh + Turbulence + Grid + AoA)
    # candidates[identity] = [ { version, path, files... }, ... ]
    candidates = defaultdict(list)
    
    for root_dir in root_dirs:
        root_path = Path(root_dir)
        if not root_path.exists():
            validation_report['issues'].append((str(root_path), "Source directory does not exist"))
            continue

        for dirpath, filenames in find_aoa_folders(root_path):
            validation_report['total_folders_found'] += 1
            
            # 1. Validate folder structure
            is_valid, lift_file, drag_file, case_file, error_msg = validate_aoa_folder(dirpath, filenames)
            
            if not is_valid:
                validation_report['skipped_folders'] += 1
                validation_report['issues'].append((str(dirpath), error_msg))
                continue
                
            # 2. Parse configuration to get Identity + Version
            metadata, error = parse_configuration(
                dirpath, case_file, config_extraction_method, position_map, value_mappings
            )
            
            if error:
                validation_report['skipped_folders'] += 1
                validation_report['issues'].append((str(dirpath), error))
                continue

            # 3. Create Unique Identity (Tuple)
            # Identity: (Geometry, Mesh, Turbulence, Grid, AoA)
            # Matches simulations that are physically identical, regardless of "Version"
            if comparison_mode == 'version':
                # If comparing versions, the version itself is part of the unique identity
                sim_identity = (
                    metadata['geometry'], 
                    metadata['mesh'], 
                    metadata['turbulence_model'], 
                    metadata['grid'], 
                    metadata['aoa'],
                    metadata.get('version_sort_key', 0)
                )
            else:
                sim_identity = (
                    metadata['geometry'], 
                    metadata['mesh'], 
                    metadata['turbulence_model'], 
                    metadata['grid'], 
                    metadata['aoa']
                )
            
            candidates[sim_identity].append({
                'version_key': metadata.get('version_sort_key', 0),
                'dirpath': dirpath,
                'lift_file': lift_file,
                'drag_file': drag_file,
                'metadata': metadata
            })
            
            validation_report['valid_folders_scanned'] += 1

    # --- PHASE 2: SELECTION & LOADING ---
    data_by_config_aoa = defaultdict(lambda: {
        'lift': [], 'drag': [], 'turbulence_model': '', 
        'geometry': '', 'mesh': '', 'version': '', 'grid': ''
    })

    print(f"Scout Phase Complete. Found {len(candidates)} unique simulation scenarios across {validation_report['valid_folders_scanned']} folders.")

    for sim_identity, version_candidates in candidates.items():
        # Sort by version_key descending (Higher is better)
        version_candidates.sort(key=lambda x: x['version_key'], reverse=True)
        
        # Winner is the first one
        winner = version_candidates[0]
        
        # Track losers
        if len(version_candidates) > 1:
            superseded = version_candidates[1:]
            validation_report['versions_suppressed'] += len(superseded)
            for loser in superseded:
                msg = f"Superseded by higher version (v{winner['version_key']} vs v{loser['version_key']}) in {winner['dirpath'].parent.name}"
                print(f"    [WARNING] SUPPRESSED: {loser['dirpath'].name} (v{loser['version_key']}) overridden by {winner['dirpath'].name}")
                validation_report['issues'].append((str(loser['dirpath']), msg))
        
        # Load Data for the Winner only
        try:
            lift_data, drag_data = load_and_correct_forces(
                winner['dirpath'] / winner['lift_file'], 
                winner['dirpath'] / winner['drag_file'], 
                winner['metadata']['aoa_number']
            )
            
            # Store data using the Winner's specific config/AoA name
            # (Or should we normalize the key? Using winner's config string ensures traceability)
            if comparison_mode == 'version':
                 # Ensure unique key for version comparison
                 config_key = f"{winner['metadata']['config']} (v{winner['metadata']['version_sort_key']})"
                 key = (config_key, winner['metadata']['aoa'])
            else:
                 # Ensure config key reflects the unique identity (Grid vs No Grid)
                 # If the config name doesn't already indicate the grid status, append it
                 config_name = winner['metadata']['config']
                 grid_status = winner['metadata'].get('grid', '')
                 
                 if 'With Grid' in grid_status and not ('.G' in config_name or '_G' in config_name):
                     config_name = f"{config_name}.G"
                 elif 'No Grid' in grid_status and not ('.NG' in config_name or '_NG' in config_name):
                     config_name = f"{config_name}.NG"
                 
                 key = (config_name, winner['metadata']['aoa'])
            
            data_by_config_aoa[key]['lift'].extend(lift_data)
            data_by_config_aoa[key]['drag'].extend(drag_data)
            
            # Populate metadata
            from copy import deepcopy
            safe_meta = deepcopy(winner['metadata'])
            
            # Ensure critical keys exist
            if 'turbulence_model' not in safe_meta: 
                safe_meta['turbulence_model'] = 'Unknown'
            if 'grid' not in safe_meta: 
                safe_meta['grid'] = 'N/A'
                
            for field in ['geometry', 'mesh', 'turbulence_model', 'version', 'grid']:
                data_by_config_aoa[key][field] = safe_meta.get(field, 'N/A')
            
            # Store source directory for finding other files (like Cp.xy)
            data_by_config_aoa[key]['source_dir'] = winner['dirpath']
                
        except Exception as e:
            validation_report['skipped_folders'] += 1
            validation_report['issues'].append((str(winner['dirpath']), f"Error loading winner data: {str(e)}"))
    
    return data_by_config_aoa, validation_report


def apply_data_manipulations(all_data, manipulation_rules, value_mappings):
    """Create synthetic data series (e.g., NG/G ratios) according to rules."""
    derived_entries = {}
    reports = []
    if not manipulation_rules:
        return derived_entries, reports

    grid_mapping = value_mappings.get('grid', {}) if value_mappings else {}
    code_to_label = grid_mapping
    label_to_label = {v: v for v in grid_mapping.values()}

    def normalize_grid_label(value):
        if value is None:
            return None
        value = str(value).strip()
        if not value:
            return None
        if value in code_to_label:
            return code_to_label[value]
        return label_to_label.get(value, value)

    for rule in manipulation_rules:
        if not rule or not rule.get('enabled', True):
            continue

        rule_name = rule.get('name', 'derived_series')
        numerator_label = normalize_grid_label(rule.get('numerator_grid'))
        denominator_label = normalize_grid_label(rule.get('denominator_grid'))
        if not numerator_label or not denominator_label:
            reports.append({'name': rule_name, 'created': 0, 'missing_pairs': 0, 'note': 'Invalid grid labels'})
            continue

        group_fields = rule.get('group_by', ['geometry', 'mesh', 'turbulence_model', 'version', 'aoa'])
        operation = rule.get('operation', 'divide').lower()
        suffix = rule.get('output_suffix') or rule_name
        output_grid_label = rule.get('output_grid_label') or f"{rule.get('numerator_grid')} / {rule.get('denominator_grid')}"

        grouped = defaultdict(dict)
        for (config, aoa), data in all_data.items():
            if data.get('is_derived'):
                continue

            group_key = []
            missing_field = False
            for field in group_fields:
                if field == 'config':
                    group_key.append(config)
                elif field == 'aoa':
                    group_key.append(aoa)
                else:
                    if field not in data:
                        missing_field = True
                        break
                    group_key.append(data.get(field))
            if missing_field:
                continue

            grouped[tuple(group_key)][data.get('grid')] = {
                'config': config,
                'aoa': aoa,
                'data': data
            }

        created = 0
        missing_pairs = 0
        for by_grid in grouped.values():
            if numerator_label not in by_grid or denominator_label not in by_grid:
                missing_pairs += 1
                continue

            num_entry = by_grid[numerator_label]
            den_entry = by_grid[denominator_label]

            if num_entry['aoa'] != den_entry['aoa']:
                missing_pairs += 1
                continue

            new_lift = _apply_series_operation(num_entry['data']['lift'], den_entry['data']['lift'], operation)
            new_drag = _apply_series_operation(num_entry['data']['drag'], den_entry['data']['drag'], operation)

            if not new_lift or not new_drag:
                missing_pairs += 1
                continue

            base_parts = num_entry['config'].split('.')
            base_without_grid = '.'.join(base_parts[:-1]) if len(base_parts) > 1 else num_entry['config']
            derived_config_name = f"{base_without_grid}.{suffix}"
            key = (derived_config_name, num_entry['aoa'])

            derived_entries[key] = {
                'lift': new_lift,
                'drag': new_drag,
                'turbulence_model': num_entry['data']['turbulence_model'],
                'geometry': num_entry['data']['geometry'],
                'mesh': num_entry['data']['mesh'],
                'version': num_entry['data']['version'],
                'grid': output_grid_label,
                'source_configs': {
                    'numerator': num_entry['config'],
                    'denominator': den_entry['config'],
                    'operation': operation
                },
                'manipulation_name': rule_name,
                'is_derived': True
            }
            created += 1

        reports.append({'name': rule_name, 'created': created, 'missing_pairs': missing_pairs})

    return derived_entries, reports


def _apply_series_operation(series_a, series_b, operation):
    """Apply element-wise manipulation between two force arrays."""
    arr_a = np.array(series_a, dtype=float)
    arr_b = np.array(series_b, dtype=float)
    length = min(len(arr_a), len(arr_b))
    if length == 0:
        return []

    arr_a = arr_a[:length]
    arr_b = arr_b[:length]

    if operation == 'divide':
        mask = arr_b != 0
        if not np.any(mask):
            return []
        result = arr_a[mask] / arr_b[mask]
    elif operation == 'percent_difference':
        mask = arr_b != 0
        if not np.any(mask):
            return []
        result = ((arr_a[mask] - arr_b[mask]) / arr_b[mask]) * 100.0
    elif operation in ('subtract', 'difference'):
        result = arr_a - arr_b
    elif operation == 'add':
        result = arr_a + arr_b
    else:
        result = arr_a - arr_b

    result = result[np.isfinite(result)]
    return result.tolist()


def _read_force_file(filepath):
    """Read force data from a text file, filtering out NaN values."""
    data = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line or '"' in line or '(' in line:
                continue
            parts = line.split()
            if len(parts) >= 2:
                try:
                    value = float(parts[1])
                    # Skip NaN values
                    if not np.isnan(value):
                        data.append(value)
                except (ValueError, IndexError):
                    continue
    return data

# ==================== STATISTICS FUNCTIONS ====================

def format_sig_figs(value, n=3):
    """Format a value to n significant figures."""
    if value is None or value == 0:
        return "0.00" if value == 0 else ""
    try:
        # Calculate the order of magnitude
        mag = np.floor(np.log10(abs(value)))
        # Scale the value, round it, and scale it back
        factor = 10**(n - 1 - mag)
        rounded = round(value * factor) / factor
        # Formatting to ensure trailing zeros if needed
        # Calculate number of decimals needed: n - (mag + 1)
        decimals = int(max(0, n - (mag + 1)))
        return f"{rounded:.{decimals}f}"
    except (ValueError, OverflowError):
        return str(value)


def compute_statistics(data):
    """Calculate mean and coefficient of variation."""
    mean_val = np.mean(data)
    std_dev = np.std(data)
    cov = (std_dev / mean_val * 100) if mean_val != 0 else 0
    return mean_val, cov


def extract_aoa_number(aoa_string):
    """Extract numeric AoA from string like 'AoA_10'."""
    try:
        return int(aoa_string.split('_')[1])
    except:
        return 0


def get_simulation_family_name(config_string):
    """Collapse config identifiers so variants that only differ by version share one table."""
    # This is legacy/default behavior (groups by Geom.Mesh.Turb.Grid, ignoring Version)
    return get_grouping_key(config_string, mode='default')


def get_grouping_key(config_string, mode='default'):
    """
    Generate a grouping key for the configuration based on the comparison mode.
    
    Modes:
    - 'turbulence': Group by Geom.Mesh (ignores Turb, Grid, Version) -> comparables are Sidebar: Turb/Grid
    - 'grid': Group by Geom.Mesh.Turb (ignores Grid, Version) -> comparables are Grid
    - 'version': Group by Geom.Mesh.Turb.Grid (fully specific) -> comparables are Version
    - 'default': Group by Geom.Mesh.Turb.Grid (ignores Version)
    """
    if not config_string:
        return config_string
        
    parts = config_string.split('.')
    if len(parts) < 3:
        return config_string
        
    # Standard parts: [0]=Geom, [1]=Mesh, [2]=Turb, [3]=Ver, [4]=Grid
    # But usually config string here is pre-processed or raw?
    # Let's assume standard format: 4.3.1.3.NG
    
    geom = parts[0]
    mesh = parts[1] if len(parts) > 1 else '?'
    turb = parts[2] if len(parts) > 2 else '?'
    
    # Handle grid part (sometimes at end)
    grid_part = None
    if len(parts) >= 5:
        grid_part = parts[4]
    elif len(parts) == 4 and parts[3] in {'NG', 'G'}:
        grid_part = parts[3]
    
    if mode == 'turbulence':
        # Group by Geometry + Mesh + Grid
        # This ensures graph titles and comparison families include the grid type (NG/G)
        return f"{geom}.{mesh}.{grid_part}" if grid_part else f"{geom}.{mesh}"
        
    elif mode == 'grid':
        # Group by Geometry + Mesh + Turbulence
        return f"{geom}.{mesh}.{turb}" if len(parts) >= 3 else config_string
    elif mode == 'expanded':
         # Expanded mode: Group by Geom.Mesh ONLY (exclude grid)
         # so that G and NG end up in the same family for pairing
         return f"{geom}.{mesh}"
    elif mode == 'version':
        # Group by everything except version (which is key)
        # But actually for version comparison, we want the table to be "4.3.1.NG" 
        # and columns to be V1, V2, V3.
        # Family is "4.3.1.NG"
        return f"{geom}.{mesh}.{turb}.{grid_part}" if grid_part else f"{geom}.{mesh}.{turb}"
        
    else: # default
        # Default behavior: "4.3.1.NG" (Agostic of version)
        return f"{geom}.{mesh}.{turb}.{grid_part}" if grid_part else f"{geom}.{mesh}.{turb}"


# ==================== CONVERGENCE ANALYSIS FUNCTIONS ====================

def analyze_convergence(data_array, min_trim=0, max_trim=0.5, num_tests=10, min_retention=0.2):
    """
    Analyze convergence with safety checks and confidence scoring.
    
    Args:
        data_array: numpy array of lift or drag values
        min_trim: minimum fraction to remove from start (0 to 1)
        max_trim: maximum fraction to remove from start (0 to 1)
        num_tests: number of trim amounts to test
        min_retention: Safety floor - don't recommend keeping less than this fraction (default 0.2 = 20%)
    
    Returns:
        Dictionary with:
        - Standard metrics: iterations_removed, iterations_used, mean, std_dev, cov
        - New metrics: median, mad, sign_changes, confidence_score
        - Recommendations: trim_recommendation, trim_reason
        - Warnings: issues detected during analysis
    """
    total_iterations = len(data_array)
    trim_fractions = np.linspace(min_trim, max_trim, num_tests)
    
    results = {
        'iterations_removed': [],
        'iterations_used': [],
        'mean': [],
        'std_dev': [],
        'median': [],
        'mad': [],  # Median Absolute Deviation
        'cov': [],
        'sign_changes': [],  # Flag if mean changes sign (oscillating through zero)
        'confidence_score': [],  # 0-1 score: how confident we are in this trim point
        'trim_recommendation': None,
        'trim_reason': None,
        'warnings': []
    }
    
    prev_mean_sign = None
    
    for trim_frac in trim_fractions:
        iterations_to_remove = int(total_iterations * trim_frac)
        trimmed_data = data_array[iterations_to_remove:]
        
        if len(trimmed_data) < 5:
            results['warnings'].append(
                f"Trim {trim_frac:.1%}: Only {len(trimmed_data)} points left (too few for statistics)"
            )
            continue
        
        trimmed_data_np = np.asarray(trimmed_data, dtype=np.float64)
        
        # Calculate statistics
        mean_val = np.mean(trimmed_data_np)
        std_val = np.std(trimmed_data_np)
        median_val = np.median(trimmed_data_np)
        mad_val = np.median(np.abs(trimmed_data_np - median_val))  # Median Absolute Deviation
        
        # Calculate COV (coefficient of variation)
        if mean_val != 0:
            cov_val = (std_val / abs(mean_val) * 100)
        else:
            cov_val = np.inf
            results['warnings'].append(
                f"Trim {trim_frac:.1%}: Mean is zero (stall condition). Using median/MAD instead."
            )
        
        # Detect sign changes (data oscillating through zero)
        current_sign = np.sign(mean_val)
        sign_changed = False
        if prev_mean_sign is not None and prev_mean_sign != current_sign:
            sign_changed = True
            results['warnings'].append(
                f"Trim {trim_frac:.1%}: Mean changed sign (oscillating). Data not fully settled."
            )
        prev_mean_sign = current_sign
        
        # Calculate confidence score (0-1, higher is more confident)
        # Weighted factors: 60% on low COV, 30% on data retention, 10% penalty for oscillation
        data_retention_ratio = len(trimmed_data) / total_iterations
        
        # Normalize COV to 0-1 scale (>50% COV = very uncertain, <2% = very confident)
        cov_normalized = min(max(cov_val / 50, 0), 1) if np.isfinite(cov_val) else 1.0
        
        # Confidence = 60% weight on low COV, 30% on data retention, 10% penalty for oscillation
        oscillation_penalty = 0.5 if sign_changed else 0.0
        confidence = 0.6 * (1 - cov_normalized) + 0.3 * data_retention_ratio - oscillation_penalty
        confidence = max(0, min(1, confidence))  # Clamp to [0, 1]
        
        results['iterations_removed'].append(iterations_to_remove)
        results['iterations_used'].append(len(trimmed_data))
        results['mean'].append(mean_val)
        results['std_dev'].append(std_val)
        results['median'].append(median_val)
        results['mad'].append(mad_val)
        results['cov'].append(cov_val)
        results['sign_changes'].append(sign_changed)
        results['confidence_score'].append(confidence)
    
    # ============ GENERATE RECOMMENDATION ============
    if len(results['cov']) > 0:
        # Find minimum COV (best convergence)
        valid_cov_indices = [i for i, c in enumerate(results['cov']) if np.isfinite(c)]
        
        if valid_cov_indices:
            min_cov_idx = min(valid_cov_indices, key=lambda i: results['cov'][i])
            trim_at_min_cov = results['iterations_used'][min_cov_idx] / total_iterations
            
            # Check if recommendation keeps enough data (safety check)
            if trim_at_min_cov >= min_retention:
                results['trim_recommendation'] = results['iterations_removed'][min_cov_idx]
                confidence_pct = results['confidence_score'][min_cov_idx] * 100
                results['trim_reason'] = (
                    f"Minimum COV ({results['cov'][min_cov_idx]:.2f}%) at "
                    f"{trim_fractions[min_cov_idx]:.1%} trim. Confidence: {confidence_pct:.0f}%"
                )
            else:
                # Over-trimming warning
                results['warnings'].append(
                    f"Best COV ({results['cov'][min_cov_idx]:.2f}%) requires trimming "
                    f"{trim_fractions[min_cov_idx]:.1%} (keeping only {trim_at_min_cov:.1%}). "
                    f"Below safety threshold ({min_retention:.0%}). Using conservative trim instead."
                )
                
                # Find best conservative trim (keeps at least min_retention of data)
                conservative_indices = [
                    i for i in valid_cov_indices 
                    if results['iterations_used'][i] / total_iterations >= min_retention
                ]
                
                if conservative_indices:
                    conservative_best_idx = min(conservative_indices, key=lambda i: results['cov'][i])
                    results['trim_recommendation'] = results['iterations_removed'][conservative_best_idx]
                    confidence_pct = results['confidence_score'][conservative_best_idx] * 100
                    results['trim_reason'] = (
                        f"Conservative trim keeping {results['iterations_used'][conservative_best_idx] / total_iterations:.1%} data. "
                        f"COV: {results['cov'][conservative_best_idx]:.2f}%, Confidence: {confidence_pct:.0f}%"
                    )
    
    return results


def plot_convergence_analysis(config, aoa, lift_data, drag_data, output_dir, max_trim, num_tests):
    """
    Create convergence analysis plots showing how statistics change with data trimming.
    
    Args:
        config: Configuration string
        aoa: Angle of attack string
        lift_data: Array of lift force values
        drag_data: Array of drag force values
        output_dir: Directory to save plots
        max_trim: Maximum fraction of data to trim
        num_tests: Number of trim amounts to test
    
    Returns:
        Tuple of (lift_results, drag_results, plot_path)
    """
    # Analyze both lift and drag
    lift_results = analyze_convergence(np.array(lift_data), min_trim=0, max_trim=max_trim, num_tests=num_tests)
    drag_results = analyze_convergence(np.array(drag_data), min_trim=0, max_trim=max_trim, num_tests=num_tests)
    
    # Create figure with subplots
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
    
    # Plot 1: Lift Mean vs Iterations Removed
    ax1.plot(lift_results['iterations_removed'], lift_results['mean'], 'o-', linewidth=2, markersize=8, color='#1f77b4')
    ax1.set_xlabel('Iterations Removed from Start', fontsize=12)
    ax1.set_ylabel('Lift Mean (N)', fontsize=12)
    ax1.set_title(f'Lift Mean Convergence\n{config} - {aoa}', fontweight='bold', fontsize=14)
    ax1.grid(True, alpha=0.3)
    
    # Plot 2: Lift COV vs Iterations Removed
    ax2.plot(lift_results['iterations_removed'], lift_results['cov'], 'o-', linewidth=2, markersize=8, color='#ff7f0e')
    ax2.set_xlabel('Iterations Removed from Start', fontsize=12)
    ax2.set_ylabel('Lift COV (%)', fontsize=12)
    ax2.set_title(f'Lift COV Convergence\n{config} - {aoa}', fontweight='bold', fontsize=14)
    ax2.grid(True, alpha=0.3)
    
    # Highlight minimum COV point for lift
    min_cov_idx = np.argmin(lift_results['cov'])
    ax2.axvline(x=lift_results['iterations_removed'][min_cov_idx], color='red', linestyle='--', linewidth=2, alpha=0.7)
    ax2.text(lift_results['iterations_removed'][min_cov_idx], max(lift_results['cov']), 
             f"  Min COV\n  Remove: {lift_results['iterations_removed'][min_cov_idx]}\n  Use: {lift_results['iterations_used'][min_cov_idx]}", 
             color='red', fontweight='bold', fontsize=9)
    
    # Plot 3: Drag Mean vs Iterations Removed
    ax3.plot(drag_results['iterations_removed'], drag_results['mean'], 'o-', linewidth=2, markersize=8, color='#2ca02c')
    ax3.set_xlabel('Iterations Removed from Start', fontsize=12)
    ax3.set_ylabel('Drag Mean (N)', fontsize=12)
    ax3.set_title(f'Drag Mean Convergence\n{config} - {aoa}', fontweight='bold', fontsize=14)
    ax3.grid(True, alpha=0.3)
    
    # Plot 4: Drag COV vs Iterations Removed
    ax4.plot(drag_results['iterations_removed'], drag_results['cov'], 'o-', linewidth=2, markersize=8, color='#d62728')
    ax4.set_xlabel('Iterations Removed from Start', fontsize=12)
    ax4.set_ylabel('Drag COV (%)', fontsize=12)
    ax4.set_title(f'Drag COV Convergence\n{config} - {aoa}', fontweight='bold', fontsize=14)
    ax4.grid(True, alpha=0.3)
    
    # Highlight minimum COV point for drag
    min_cov_idx = np.argmin(drag_results['cov'])
    ax4.axvline(x=drag_results['iterations_removed'][min_cov_idx], color='red', linestyle='--', linewidth=2, alpha=0.7)
    ax4.text(drag_results['iterations_removed'][min_cov_idx], max(drag_results['cov']), 
             f"  Min COV\n  Remove: {drag_results['iterations_removed'][min_cov_idx]}\n  Use: {drag_results['iterations_used'][min_cov_idx]}", 
             color='red', fontweight='bold', fontsize=9)
    
    plt.tight_layout()
    
    # Save convergence analysis plot
    convergence_dir = output_dir / "convergence_analysis"
    convergence_dir.mkdir(parents=True, exist_ok=True)
    
    plot_file = convergence_dir / f"convergence_{config}_{aoa}.png"
    try:
        plt.savefig(plot_file, dpi=300) 
    except Exception as e:
        print(f"    ⚠️  Warning: Could not save plot {plot_file.name}: {e}")
    plt.close()
    
    return lift_results, drag_results, str(plot_file)


# ==================== EXCEL EXPORT FUNCTIONS ====================

def create_data_summary_sheet(wb, all_data, num_iterations, convergence_results):
    """Create Data Summary sheet in Excel with formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    group_header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    group_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    row_fill_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Helper function to get optimized data
    def get_optimized_data(config, aoa, data):
        if convergence_results and (config, aoa) in convergence_results:
            conv = convergence_results[(config, aoa)]
            lift_min_cov_idx = np.argmin(conv['lift']['cov'])
            drag_min_cov_idx = np.argmin(conv['drag']['cov'])
            
            optimal_lift_trim = conv['lift']['iterations_removed'][lift_min_cov_idx]
            optimal_drag_trim = conv['drag']['iterations_removed'][drag_min_cov_idx]
            optimal_trim = max(optimal_lift_trim, optimal_drag_trim)
            
            lift_values = data['lift'][optimal_trim:]
            drag_values = data['drag'][optimal_trim:]
        else:
            lift_values = data['lift'][-num_iterations:] if len(data['lift']) >= num_iterations else data['lift']
            drag_values = data['drag'][-num_iterations:] if len(data['drag']) >= num_iterations else data['drag']
        return lift_values, drag_values
    
    # Group data by base configuration
    base_config_data = defaultdict(list)
    for (config, aoa), data in all_data.items():
        base_config = get_simulation_family_name(config)
        
        lift_values, drag_values = get_optimized_data(config, aoa, data)
        
        lift_mean = np.mean(lift_values) if lift_values else 0
        lift_std = np.std(lift_values) if lift_values else 0
        lift_cov = (lift_std / lift_mean * 100) if lift_mean != 0 else 0
        
        drag_mean = np.mean(drag_values) if drag_values else 0
        drag_std = np.std(drag_values) if drag_values else 0
        drag_cov = (drag_std / drag_mean * 100) if drag_mean != 0 else 0
        
        base_config_data[base_config].append({
            'aoa': aoa,
            'aoa_num': extract_aoa_number(aoa),
            'turbulence_model': data['turbulence_model'],
            'num_points': len(lift_values),
            'lift_mean': lift_mean,
            'lift_cov': lift_cov,
            'drag_mean': drag_mean,
            'drag_cov': drag_cov
        })
    
    # Sort each group by AoA
    for base_config in base_config_data:
        base_config_data[base_config].sort(key=lambda x: x['aoa_num'])
    
    # Sort base configs by turbulence model
    model_order = {'SST': 1, 'RNG': 2, 'RSM': 3, 'k-epsilon': 4}
    def get_sort_key(base_config):
        if base_config in base_config_data and base_config_data[base_config]:
            turb_model = base_config_data[base_config][0]['turbulence_model']
            return (model_order.get(turb_model, 99), base_config)
        return (99, base_config)
    sorted_base_configs = sorted(base_config_data.keys(), key=get_sort_key)
    
    # Create worksheet
    ws = wb.active
    ws.title = 'Data Summary'
    
    columns = ['Turbulence Model', 'Angle of Attack', 'Lift Mean (N)', 'Lift COV (%)', 
               'Drag Mean (N)', 'Drag COV (%)', 'Num Points']
    current_row = 1
    
    for base_config in sorted_base_configs:
        data_rows = base_config_data[base_config]
        if not data_rows:
            continue
        
        # Section header
        ws.cell(row=current_row, column=1, value=base_config)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
        header_cell = ws.cell(row=current_row, column=1)
        header_cell.font = group_header_font
        header_cell.fill = group_header_fill
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border_style
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Column headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Data rows
        for row_idx, row_data in enumerate(data_rows):
            fill_color = row_fill_light if row_idx % 2 == 0 else row_fill_white
            
            values = [
                row_data['turbulence_model'],
                row_data['aoa_num'],
                f"{row_data['lift_mean']:.1f}",
                f"{row_data['lift_cov']:.1f}",
                f"{row_data['drag_mean']:.1f}",
                f"{row_data['drag_cov']:.1f}",
                row_data['num_points']
            ]
            
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.alignment = data_alignment
                cell.border = border_style
                cell.fill = fill_color
            
            current_row += 1
        
        current_row += 1  # Blank row
    
    # Autofit columns
    for col_idx in range(1, len(columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_turbulence_comparison_sheet(wb, all_data, num_iterations, convergence_results, comparison_mode='turbulence'):
    """
    Create 'Comparison' sheet depending on mode.
    Mode 'turbulence': Columns are Turbulence Models.
    Mode 'grid': Columns are Grid Types.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    sheet_title = 'Comparison'
    if comparison_mode == 'turbulence':
        sheet_title = 'Turbulence Comparison'
    elif comparison_mode == 'grid':
        sheet_title = 'Grid Comparison'
        
    ws = wb.create_sheet(title=sheet_title)
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    table_title_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    table_title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    row_fill_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Helper function
    def get_optimized_data(config, aoa, data):
        if convergence_results and (config, aoa) in convergence_results:
            conv = convergence_results[(config, aoa)]
            lift_min_cov_idx = np.argmin(conv['lift']['cov'])
            drag_min_cov_idx = np.argmin(conv['drag']['cov'])
            
            optimal_lift_trim = conv['lift']['iterations_removed'][lift_min_cov_idx]
            optimal_drag_trim = conv['drag']['iterations_removed'][drag_min_cov_idx]
            optimal_trim = max(optimal_lift_trim, optimal_drag_trim)
            
            lift_values = data['lift'][optimal_trim:]
            drag_values = data['drag'][optimal_trim:]
        else:
            lift_values = data['lift'][-num_iterations:] if len(data['lift']) >= num_iterations else data['lift']
            drag_values = data['drag'][-num_iterations:] if len(data['drag']) >= num_iterations else data['drag']
        return lift_values, drag_values
    
    # Group by base config and turbulence model
    comparison_data = defaultdict(lambda: defaultdict(dict))
    
    # Determine what is the "Column Header" (Sim Type) and what is the "Row Group" (Family)
    for (config, aoa), data in all_data.items():
        base_config = get_grouping_key(config, mode=comparison_mode)
        
        if comparison_mode == 'turbulence':
            # Label = Turbulence Model (e.g. SST, RNG)
            # Family = Geom + Mesh + Grid (now handled by get_grouping_key)
            sim_type = data['turbulence_model']

        elif comparison_mode == 'grid':
            # Label = Grid (e.g. With Grid, No Grid)
            sim_type = "With Grid" if data['grid'] in ['G', 'With Grid'] else "No Grid"
        else:
            sim_type = data['turbulence_model']
        
        lift_values, drag_values = get_optimized_data(config, aoa, data)
        
        lift_mean = np.mean(lift_values) if len(lift_values) > 0 else 0
        lift_std = np.std(lift_values) if len(lift_values) > 0 else 0
        lift_cov = (lift_std / lift_mean * 100) if lift_mean != 0 else 0
        
        drag_mean = np.mean(drag_values) if len(drag_values) > 0 else 0
        drag_std = np.std(drag_values) if len(drag_values) > 0 else 0
        drag_cov = (drag_std / drag_mean * 100) if drag_mean != 0 else 0
        
        comparison_data[base_config][sim_type][aoa] = {
            'lift_mean': lift_mean,
            'lift_cov': lift_cov,
            'drag_mean': drag_mean,
            'drag_cov': drag_cov,
            'aoa_num': extract_aoa_number(aoa)
        }
    
    # Get sorted AoAs
    all_aoas = set()
    for base_config in comparison_data:
        for sim_type in comparison_data[base_config]:
            all_aoas.update(comparison_data[base_config][sim_type].keys())
    sorted_aoas = sorted(all_aoas, key=extract_aoa_number)
    
    current_row = 1
    
    # Create 4 tables for each base configuration
    for base_config in sorted(comparison_data.keys()):
        models_in_config = comparison_data[base_config]
        
        # Determine all unique turbulence models present for this config
        # Sort them generally, maybe putting SST first if present for baseline reference
        unique_models = sorted(list(models_in_config.keys()))
        if 'SST' in unique_models:
             unique_models.remove('SST')
             unique_models.insert(0, 'SST')
        
        # Table 1: Lift Mean
        ws.cell(row=current_row, column=1, value=f"{base_config} - Lift Mean (N)")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sorted_aoas)+1)
        cell = ws.cell(row=current_row, column=1)
        cell.font = table_title_font
        cell.fill = table_title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Headers
        header_label = 'Turbulence Model' if comparison_mode == 'turbulence' else 'Grid Config'
        ws.cell(row=current_row, column=1, value=header_label).font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).alignment = header_alignment
        ws.cell(row=current_row, column=1).border = border_style
        
        for col_idx, aoa in enumerate(sorted_aoas, 2):
            # Use numeric AoA for header
            period_val = extract_aoa_number(aoa)
            ws.cell(row=current_row, column=col_idx, value=period_val).font = header_font
            ws.cell(row=current_row, column=col_idx).fill = header_fill
            ws.cell(row=current_row, column=col_idx).alignment = header_alignment
            ws.cell(row=current_row, column=col_idx).border = border_style
        current_row += 1
        
        # Data rows
        sst_data = models_in_config.get('SST', {})
        for row_idx, turb_model in enumerate(unique_models):
            if turb_model not in models_in_config:
                continue
            
            fill = row_fill_light if row_idx % 2 == 0 else row_fill_white
            
            ws.cell(row=current_row, column=1, value=turb_model).alignment = data_alignment
            ws.cell(row=current_row, column=1).border = border_style
            ws.cell(row=current_row, column=1).fill = fill
            
            for col_idx, aoa in enumerate(sorted_aoas, 2):
                if aoa in models_in_config[turb_model]:
                    lift_mean = models_in_config[turb_model][aoa]['lift_mean']
                    
                    # Calculate % difference from SST
                    if turb_model != 'SST' and aoa in sst_data:
                        sst_lift = sst_data[aoa]['lift_mean']
                        if sst_lift != 0:
                            pct_diff = ((lift_mean - sst_lift) / sst_lift) * 100
                            value = f"{lift_mean:.1f} ({pct_diff:+.1f}%)"
                        else:
                            value = f"{lift_mean:.1f}"
                    else:
                        value = f"{lift_mean:.1f}"
                    
                    ws.cell(row=current_row, column=col_idx, value=value).alignment = data_alignment
                    ws.cell(row=current_row, column=col_idx).border = border_style
                    ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1
        current_row += 2
        
        # Table 2: Drag Mean
        ws.cell(row=current_row, column=1, value=f"{base_config} - Drag Mean (N)")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sorted_aoas)+1)
        cell = ws.cell(row=current_row, column=1)
        cell.font = table_title_font
        cell.fill = table_title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Headers
        ws.cell(row=current_row, column=1, value='Turbulence Model').font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).alignment = header_alignment
        ws.cell(row=current_row, column=1).border = border_style
        
        for col_idx, aoa in enumerate(sorted_aoas, 2):
            period_val = extract_aoa_number(aoa)
            ws.cell(row=current_row, column=col_idx, value=period_val).font = header_font
            ws.cell(row=current_row, column=col_idx).fill = header_fill
            ws.cell(row=current_row, column=col_idx).alignment = header_alignment
            ws.cell(row=current_row, column=col_idx).border = border_style
        current_row += 1
        
        # Data rows
        for row_idx, turb_model in enumerate(unique_models):
            if turb_model not in models_in_config:
                continue
            
            fill = row_fill_light if row_idx % 2 == 0 else row_fill_white
            
            ws.cell(row=current_row, column=1, value=turb_model).alignment = data_alignment
            ws.cell(row=current_row, column=1).border = border_style
            ws.cell(row=current_row, column=1).fill = fill
            
            for col_idx, aoa in enumerate(sorted_aoas, 2):
                if aoa in models_in_config[turb_model]:
                    drag_mean = models_in_config[turb_model][aoa]['drag_mean']
                    
                    if turb_model != 'SST' and aoa in sst_data:
                        sst_drag = sst_data[aoa]['drag_mean']
                        if sst_drag != 0:
                            pct_diff = ((drag_mean - sst_drag) / sst_drag) * 100
                            value = f"{drag_mean:.1f} ({pct_diff:+.1f}%)"
                        else:
                            value = f"{drag_mean:.1f}"
                    else:
                        value = f"{drag_mean:.1f}"
                    
                    ws.cell(row=current_row, column=col_idx, value=value).alignment = data_alignment
                    ws.cell(row=current_row, column=col_idx).border = border_style
                    ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1
        current_row += 2
        
        # Table 3: Lift COV
        ws.cell(row=current_row, column=1, value=f"{base_config} - Lift COV (%)")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sorted_aoas)+1)
        cell = ws.cell(row=current_row, column=1)
        cell.font = table_title_font
        cell.fill = table_title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Headers
        ws.cell(row=current_row, column=1, value='Turbulence Model').font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).alignment = header_alignment
        ws.cell(row=current_row, column=1).border = border_style
        
        for col_idx, aoa in enumerate(sorted_aoas, 2):
            period_val = extract_aoa_number(aoa)
            ws.cell(row=current_row, column=col_idx, value=period_val).font = header_font
            ws.cell(row=current_row, column=col_idx).fill = header_fill
            ws.cell(row=current_row, column=col_idx).alignment = header_alignment
            ws.cell(row=current_row, column=col_idx).border = border_style
        current_row += 1
        
        # Data rows
        for row_idx, turb_model in enumerate(unique_models):
            if turb_model not in models_in_config:
                continue
            
            fill = row_fill_light if row_idx % 2 == 0 else row_fill_white
            
            ws.cell(row=current_row, column=1, value=turb_model).alignment = data_alignment
            ws.cell(row=current_row, column=1).border = border_style
            ws.cell(row=current_row, column=1).fill = fill
            
            for col_idx, aoa in enumerate(sorted_aoas, 2):
                if aoa in models_in_config[turb_model]:
                    lift_cov = models_in_config[turb_model][aoa]['lift_cov']
                    value = f"{lift_cov:.1f}"
                    
                    ws.cell(row=current_row, column=col_idx, value=value).alignment = data_alignment
                    ws.cell(row=current_row, column=col_idx).border = border_style
                    ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1
        current_row += 2
        
        # Table 4: Drag COV
        ws.cell(row=current_row, column=1, value=f"{base_config} - Drag COV (%)")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sorted_aoas)+1)
        cell = ws.cell(row=current_row, column=1)
        cell.font = table_title_font
        cell.fill = table_title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Headers
        ws.cell(row=current_row, column=1, value='Turbulence Model').font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).alignment = header_alignment
        ws.cell(row=current_row, column=1).border = border_style
        
        for col_idx, aoa in enumerate(sorted_aoas, 2):
            period_val = extract_aoa_number(aoa)
            ws.cell(row=current_row, column=col_idx, value=period_val).font = header_font
            ws.cell(row=current_row, column=col_idx).fill = header_fill
            ws.cell(row=current_row, column=col_idx).alignment = header_alignment
            ws.cell(row=current_row, column=col_idx).border = border_style
        current_row += 1
        
        # Data rows
        for row_idx, turb_model in enumerate(unique_models):
            if turb_model not in models_in_config:
                continue
            
            fill = row_fill_light if row_idx % 2 == 0 else row_fill_white
            
            ws.cell(row=current_row, column=1, value=turb_model).alignment = data_alignment
            ws.cell(row=current_row, column=1).border = border_style
            ws.cell(row=current_row, column=1).fill = fill
            
            for col_idx, aoa in enumerate(sorted_aoas, 2):
                if aoa in models_in_config[turb_model]:
                    drag_cov = models_in_config[turb_model][aoa]['drag_cov']
                    value = f"{drag_cov:.1f}"
                    
                    ws.cell(row=current_row, column=col_idx, value=value).alignment = data_alignment
                    ws.cell(row=current_row, column=col_idx).border = border_style
                    ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1

        current_row += 3
    
    # Autofit columns
    for col_idx in range(1, len(sorted_aoas) + 2):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_version_comparison_sheet(
    wb,
    all_data,
    comparison_configs,
    num_iterations,
    convergence_results,
    q_times_a,
    comparison_mode='default'
):
    """Create a sheet that compares configured version pairs. Returns True if created."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Mode 'version': AUTO generate comparisons for all multiversion families
    if comparison_mode == 'version':
        return _create_auto_version_comparison_sheet(wb, all_data, num_iterations, convergence_results, q_times_a)

    if not comparison_configs:
        return False

    config_lookup = defaultdict(dict)
    for (config, aoa), data in all_data.items():
        config_lookup[config][aoa] = data

    def resolve_config_name(name):
        target = str(name).strip()
        if not target:
            return None
        if target in config_lookup:
            return target
        matches = [cfg for cfg in config_lookup if cfg.startswith(target)]
        if len(matches) == 1:
            return matches[0]
        return None

    def normalize_pairs(entries):
        normalized = []
        if isinstance(entries, dict):
            raw_pairs = entries.get('pairs', [])
            for pair in raw_pairs:
                if isinstance(pair, (list, tuple)) and len(pair) == 2:
                    normalized.append((pair[0], pair[1]))
        else:
            configs = [str(item).strip() for item in (entries or []) if str(item).strip()]
            if len(configs) >= 2:
                baseline = configs[0]
                for target in configs[1:]:
                    normalized.append((baseline, target))
        return normalized

    comparison_plan = {}
    skipped_pairs = []

    for base_key, entries in comparison_configs.items():
        normalized_pairs = normalize_pairs(entries)
        valid_pairs = []
        for raw_a, raw_b in normalized_pairs:
            cfg_a = resolve_config_name(raw_a)
            cfg_b = resolve_config_name(raw_b)
            if cfg_a and cfg_b:
                valid_pairs.append((cfg_a, cfg_b))
            else:
                skipped_pairs.append((base_key, raw_a, raw_b))
                missing = []
                if not cfg_a:
                    missing.append(str(raw_a))
                if not cfg_b:
                    missing.append(str(raw_b))
                if missing:
                    label = ' & '.join(missing)
                    print(f"  ⚠ Version comparison skipped: {label} not found or ambiguous")
        if valid_pairs:
            comparison_plan[base_key] = valid_pairs

    if not comparison_plan:
        if skipped_pairs:
            print("  ⚠ Version comparisons skipped: no matching configuration pairs found")
        return False

    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    subheader_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    row_fill_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    metrics_cache = {}

    def get_metrics(config, aoa):
        key = (config, aoa)
        if key in metrics_cache:
            return metrics_cache[key]

        data = config_lookup.get(config, {}).get(aoa)
        if not data:
            metrics_cache[key] = None
            return None

        if convergence_results and key in convergence_results:
            conv = convergence_results[key]
            lift_idx = np.argmin(conv['lift']['cov'])
            drag_idx = np.argmin(conv['drag']['cov'])
            optimal_trim = max(
                conv['lift']['iterations_removed'][lift_idx],
                conv['drag']['iterations_removed'][drag_idx],
            )
            lift_values = data['lift'][optimal_trim:]
            drag_values = data['drag'][optimal_trim:]
        else:
            lift_values = data['lift'][-num_iterations:] if len(data['lift']) >= num_iterations else data['lift']
            drag_values = data['drag'][-num_iterations:] if len(data['drag']) >= num_iterations else data['drag']

        lift_arr = np.array(lift_values, dtype=float)
        drag_arr = np.array(drag_values, dtype=float)

        lift_mean = float(np.mean(lift_arr)) if lift_arr.size else None
        drag_mean = float(np.mean(drag_arr)) if drag_arr.size else None
        lift_std = float(np.std(lift_arr)) if lift_arr.size else None
        drag_std = float(np.std(drag_arr)) if drag_arr.size else None

        lift_cov = (lift_std / lift_mean * 100) if lift_mean not in (None, 0) and lift_std is not None else None
        drag_cov = (drag_std / drag_mean * 100) if drag_mean not in (None, 0) and drag_std is not None else None

        if q_times_a:
            c_l = (lift_mean / q_times_a) if lift_mean is not None else None
            c_d = (drag_mean / q_times_a) if drag_mean is not None else None
        else:
            c_l = None
            c_d = None

        metrics_cache[key] = {
            'lift_mean': lift_mean,
            'lift_cov': lift_cov,
            'drag_mean': drag_mean,
            'drag_cov': drag_cov,
            'cl': c_l,
            'cd': c_d,
            'count': int(lift_arr.size) if lift_arr.size else 0,
        }
        return metrics_cache[key]

    ws = wb.create_sheet(title='Version_Comparison')
    current_row = 1

    columns = [
        'AoA',
        'Lift A (N)',
        'Lift B (N)',
        'ΔLift (N)',
        'ΔLift (%)',
        'Drag A (N)',
        'Drag B (N)',
        'ΔDrag (N)',
        'ΔDrag (%)',
        'C_L A',
        'C_L B',
        'ΔC_L (%)',
        'C_D A',
        'C_D B',
        'ΔC_D (%)',
        'Points A',
        'Points B',
    ]

    def fmt(value, precision):
        return "" if value is None else f"{value:.{precision}f}"

    def fmt_delta(a_val, b_val, precision, percent=False):
        if a_val is None or b_val is None:
            return ""
        delta = b_val - a_val
        if percent:
            if a_val == 0:
                return ""
            delta = (delta / abs(a_val)) * 100
        sign = "+" if delta >= 0 else ""
        return f"{sign}{delta:.{precision}f}"

    for base_key in sorted(comparison_plan.keys()):
        ws.cell(row=current_row, column=1, value=f"{base_key} Version Comparisons")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
        cell = ws.cell(row=current_row, column=1)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_style
        current_row += 1

        for cfg_a, cfg_b in comparison_plan[base_key]:
            ws.cell(row=current_row, column=1, value=f"{cfg_a} vs {cfg_b}")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
            cell = ws.cell(row=current_row, column=1)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
            current_row += 1

            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style
            current_row += 1

            aoas = set(config_lookup[cfg_a].keys()) | set(config_lookup[cfg_b].keys())
            sorted_aoas = sorted(aoas, key=extract_aoa_number)

            for idx, aoa in enumerate(sorted_aoas):
                metrics_a = get_metrics(cfg_a, aoa)
                metrics_b = get_metrics(cfg_b, aoa)
                fill = row_fill_light if idx % 2 == 0 else row_fill_white

                lift_a = metrics_a['lift_mean'] if metrics_a else None
                lift_b = metrics_b['lift_mean'] if metrics_b else None
                drag_a = metrics_a['drag_mean'] if metrics_a else None
                drag_b = metrics_b['drag_mean'] if metrics_b else None
                cl_a = metrics_a['cl'] if metrics_a else None
                cl_b = metrics_b['cl'] if metrics_b else None
                cd_a = metrics_a['cd'] if metrics_a else None
                cd_b = metrics_b['cd'] if metrics_b else None
                count_a = metrics_a['count'] if metrics_a else 0
                count_b = metrics_b['count'] if metrics_b else 0

                row_values = [
                    extract_aoa_number(aoa),
                    fmt(lift_a, 3),
                    fmt(lift_b, 3),
                    fmt_delta(lift_a, lift_b, 3),
                    fmt_delta(lift_a, lift_b, 2, percent=True),
                    fmt(drag_a, 3),
                    fmt(drag_b, 3),
                    fmt_delta(drag_a, drag_b, 3),
                    fmt_delta(drag_a, drag_b, 2, percent=True),
                    format_sig_figs(cl_a, 3),
                    format_sig_figs(cl_b, 3),
                    fmt_delta(cl_a, cl_b, 2, percent=True),
                    format_sig_figs(cd_a, 3),
                    format_sig_figs(cd_b, 3),
                    fmt_delta(cd_a, cd_b, 2, percent=True),
                    count_a if count_a else "",
                    count_b if count_b else "",
                ]

                for col_idx, val in enumerate(row_values, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.alignment = data_alignment
                    cell.border = border_style
                    cell.fill = fill
                current_row += 1

            current_row += 1  # Blank row between pair tables

        current_row += 1  # Additional spacing between base sections

    # Autofit columns
    for col_idx in range(1, len(columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
        adjusted_width = min(max_length + 2, 26)
        ws.column_dimensions[column_letter].width = adjusted_width

    return True


def _create_auto_version_comparison_sheet(wb, all_data, num_iterations, convergence_results, q_times_a):
    """
    Automatically generate version comparison sheet by grouping all_data by Family.
    Family = Geom.Mesh.Turb.Grid (ignoring version).
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # 1. Group data by Family
    family_groups = defaultdict(list)
    for key, data in all_data.items():
        # key is (config_string, aoa)
        config_str = key[0]
        # For version mode, get_grouping_key gives us the family (ignoring version)
        family = get_grouping_key(config_str, mode='version')
        family_groups[family].append({
            'config': config_str,
            'aoa': key[1],
            'version': data.get('version', 'Unknown'),
            'version_key': data.get('version_sort_key', 0),
            'data': data
        })

    # Filter for families that actually HAVE multiple versions
    multi_version_families = []
    for family, items in family_groups.items():
        # Check if we have different versions for the same AoA?
        # Or just different versions in general?
        # usually we want to compare V1 vs V2 at same AoA.
        
        # Group immediate items by AoA
        aoa_map = defaultdict(list)
        for item in items:
            aoa_map[item['aoa']].append(item)
            
        has_multiple = False
        for aoa, subitems in aoa_map.items():
            if len(subitems) > 1:
                has_multiple = True
                break
        
        if has_multiple:
            multi_version_families.append(family)

    if not multi_version_families:
        print("  ⚠ No multi-version simulations found for auto-comparison.")
        return False

    ws = wb.create_sheet("Version Comparison")
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    table_title_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    table_title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    def get_metrics_val(data):
        # Helper to get mean values
        if convergence_results:
             # Try to find matching convergence result. 
             # Key in convergence_results matches key in all_data
             # We need to reconstruct the key used in all_data.
             # Since we are iterating sub-items which HAVE 'config' which IS the key's config part...
             key = (data['config_key'], data['aoa']) # We stored this in the loop below
             if key in convergence_results:
                 conv = convergence_results[key]
                 lift_idx = np.argmin(conv['lift']['cov'])
                 drag_idx = np.argmin(conv['drag']['cov'])
                 optimal_trim = max(conv['lift']['iterations_removed'][lift_idx], conv['drag']['iterations_removed'][drag_idx])
                 lift = data['data']['lift'][optimal_trim:]
                 drag = data['data']['drag'][optimal_trim:]
                 return np.mean(lift) if lift else 0, np.mean(drag) if drag else 0

        lift = data['data']['lift'][-num_iterations:] if len(data['data']['lift']) >= num_iterations else data['data']['lift']
        drag = data['data']['drag'][-num_iterations:] if len(data['data']['drag']) >= num_iterations else data['data']['drag']
        return np.mean(lift) if lift else 0, np.mean(drag) if drag else 0

    current_row = 1
    
    for family in multi_version_families:
        items = family_groups[family]
        # Group by AoA
        aoa_map = defaultdict(list)
        versions_found = set()
        for item in items:
            aoa_map[item['aoa']].append(item)
            versions_found.add(item['version_key'])
            # Store key for metrics lookup
            # The 'config' field in item is the raw config string from the key?
            # In the loop above: config_str = key[0].
            # And in load_lift_drag_data, key[0] IS the unique string "Config (vX)"
            item['config_key'] = item['config'] 
            
        sorted_versions = sorted(list(versions_found))
        sorted_aoas = sorted(aoa_map.keys(), key=extract_aoa_number)

        # Table Header
        ws.cell(row=current_row, column=1, value=f"{family} - Version Comparison").font = table_title_font
        ws.cell(row=current_row, column=1).fill = table_title_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=1 + len(sorted_versions)*2) # *2 for Lift/Drag
        current_row += 1
        
        # Column Headers
        ws.cell(row=current_row, column=1, value="AoA").font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        
        for i, ver in enumerate(sorted_versions):
            col_start = 2 + i*2
            ws.cell(row=current_row, column=col_start, value=f"v{ver} Lift").font = header_font
            ws.cell(row=current_row, column=col_start).fill = header_fill
            ws.cell(row=current_row, column=col_start+1, value=f"v{ver} Drag").font = header_font
            ws.cell(row=current_row, column=col_start+1).fill = header_fill
        current_row += 1
        
        # Data Rows
        for aoa in sorted_aoas:
            ws.cell(row=current_row, column=1, value=extract_aoa_number(aoa)).alignment = Alignment(horizontal='center')
            
            subitems = aoa_map[aoa]
            # Map version_key to data
            ver_data_map = {item['version_key']: item for item in subitems}
            
            for i, ver in enumerate(sorted_versions):
                col_start = 2 + i*2
                if ver in ver_data_map:
                    l_mean, d_mean = get_metrics_val(ver_data_map[ver])
                    ws.cell(row=current_row, column=col_start, value=l_mean).number_format = '0.0000'
                    ws.cell(row=current_row, column=col_start+1, value=d_mean).number_format = '0.0000'
                else:
                    ws.cell(row=current_row, column=col_start, value="-")
                    ws.cell(row=current_row, column=col_start+1, value="-")
            
            current_row += 1
        
        current_row += 2 # Spacer
        
    return True


def create_coefficients_sheet(wb, all_data, num_iterations, convergence_results, q_times_a):
    """Create Coefficients sheet with C_L and C_D."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    ws = wb.create_sheet(title='Coefficients')
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    row_fill_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Headers
    columns = ['Configuration', 'Turbulence Model', 'AoA', 'C_L', 'C_L COV (%)', 'C_D', 'C_D COV (%)']
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    ws.row_dimensions[1].height = 30
    
    # Data rows
    row = 2
    
    # Sort by Config Family first, then AoA
    sorted_items = sorted(all_data.items(), key=lambda x: (get_simulation_family_name(x[0][0]), extract_aoa_number(x[0][1])))
    
    for (config, aoa), data in sorted_items:
        # Get optimized or fixed iteration data
        if convergence_results and (config, aoa) in convergence_results:
            conv = convergence_results[(config, aoa)]
            lift_min_cov_idx = np.argmin(conv['lift']['cov'])
            drag_min_cov_idx = np.argmin(conv['drag']['cov'])
            
            optimal_lift_trim = conv['lift']['iterations_removed'][lift_min_cov_idx]
            optimal_drag_trim = conv['drag']['iterations_removed'][drag_min_cov_idx]
            optimal_trim = max(optimal_lift_trim, optimal_drag_trim)
            
            lift_values = data['lift'][optimal_trim:]
            drag_values = data['drag'][optimal_trim:]
        else:
            lift_values = data['lift'][-num_iterations:] if len(data['lift']) >= num_iterations else data['lift']
            drag_values = data['drag'][-num_iterations:] if len(data['drag']) >= num_iterations else data['drag']
        
        lift_mean = np.mean(lift_values) if len(lift_values) > 0 else 0
        drag_mean = np.mean(drag_values) if len(drag_values) > 0 else 0
        lift_std = np.std(lift_values) if len(lift_values) > 0 else 0
        drag_std = np.std(drag_values) if len(drag_values) > 0 else 0
        
        C_L = lift_mean / q_times_a if q_times_a != 0 else 0
        C_D = drag_mean / q_times_a if q_times_a != 0 else 0
        C_L_std = lift_std / q_times_a if q_times_a != 0 else 0
        C_D_std = drag_std / q_times_a if q_times_a != 0 else 0
        C_L_cov = (C_L_std / C_L * 100) if C_L != 0 else 0
        C_D_cov = (C_D_std / C_D * 100) if C_D != 0 else 0
        
        fill = row_fill_light if row % 2 == 0 else row_fill_white
        values = [config, data['turbulence_model'], extract_aoa_number(aoa), format_sig_figs(C_L, 3), f"{C_L_cov:.1f}", format_sig_figs(C_D, 3), f"{C_D_cov:.1f}"]
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = data_alignment
            cell.border = border_style
            cell.fill = fill
        row += 1
    
    # Autofit columns
    for col_idx in range(1, len(columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_optimized_statistics_sheet(wb, all_data, convergence_results, num_iterations, q_times_a):
    """Create Optimized Statistics sheet comparing original vs optimized."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    if not convergence_results:
        return
    
    ws = wb.create_sheet(title='Optimized_Statistics')
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    row_fill_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Headers
    columns = ['Turbulence Model', 'Configuration', 'AoA', 'Original Iterations', 'Optimal Trim', 'Optimized Iterations',
               'Lift Mean (Orig)', 'Lift Mean (Opt)', 'Lift COV (Orig)', 'Lift COV (Opt)', 'Lift COV Δ',
               'Drag Mean (Orig)', 'Drag Mean (Opt)', 'Drag COV (Orig)', 'Drag COV (Opt)', 'Drag COV Δ']
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    ws.row_dimensions[1].height = 30
    
    # Data rows
    row = 2
    
    # Sort by Family -> AoA (numerical)
    sorted_items = sorted(
        all_data.items(),
        key=lambda x: (get_simulation_family_name(x[0][0]), extract_aoa_number(x[0][1]))
    )

    for (config, aoa), data in sorted_items:
        if (config, aoa) not in convergence_results:
            continue
        
        conv = convergence_results[(config, aoa)]
        
        # Original stats (last num_iterations)
        orig_lift = data['lift'][-num_iterations:] if len(data['lift']) >= num_iterations else data['lift']
        orig_drag = data['drag'][-num_iterations:] if len(data['drag']) >= num_iterations else data['drag']
        orig_lift_mean, orig_lift_cov = compute_statistics(orig_lift)
        orig_drag_mean, orig_drag_cov = compute_statistics(orig_drag)
        
        # Optimized stats
        lift_min_idx = np.argmin(conv['lift']['cov'])
        drag_min_idx = np.argmin(conv['drag']['cov'])
        optimal_lift_trim = conv['lift']['iterations_removed'][lift_min_idx]
        optimal_drag_trim = conv['drag']['iterations_removed'][drag_min_idx]
        optimal_trim = max(optimal_lift_trim, optimal_drag_trim)
        
        opt_lift = data['lift'][optimal_trim:]
        opt_drag = data['drag'][optimal_trim:]
        opt_lift_mean, opt_lift_cov = compute_statistics(opt_lift)
        opt_drag_mean, opt_drag_cov = compute_statistics(opt_drag)
        
        fill = row_fill_light if row % 2 == 0 else row_fill_white
        values = [
            data['turbulence_model'], config, extract_aoa_number(aoa),
            len(orig_lift), optimal_trim, len(opt_lift),
            f"{orig_lift_mean:.3f}", f"{opt_lift_mean:.3f}",
            f"{orig_lift_cov:.2f}%", f"{opt_lift_cov:.2f}%", f"{(orig_lift_cov - opt_lift_cov):+.2f}%",
            f"{orig_drag_mean:.3f}", f"{opt_drag_mean:.3f}",
            f"{orig_drag_cov:.2f}%", f"{opt_drag_cov:.2f}%", f"{(orig_drag_cov - opt_drag_cov):+.2f}%"
        ]
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = data_alignment
            cell.border = border_style
            cell.fill = fill
        row += 1
    
    # Autofit columns
    for col_idx in range(1, len(columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width


def apply_excel_formatting(excel_file):
    """Legacy function - formatting now done during sheet creation."""
    pass


# ==================== PLOTTING FUNCTIONS ====================

def create_coefficient_graphs(all_data, coefficient_data, output_dir, position_map, value_mappings, comparison_mode='default', max_cov_threshold=None):
    """Create all coefficient graphs organized by comparison family."""
    
    import itertools
    colors = ACADEMIC_COLORS
    markers = ACADEMIC_MARKERS
    
    # Helper to extract grid status from config string
    def get_grid_status(config):
        if '.NG' in config or 'No Grid' in config:
            return 'No Grid'
        elif '.G' in config or 'With Grid' in config:
            return 'With Grid'
        return 'Unknown'
    
    # 1. Group data by "Family" (The Graph Title/Folder)
    #    For expanded/grid modes, we also track grid status
    graphs_data = defaultdict(lambda: defaultdict(list))
    graphs_data_by_grid = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))  # For expanded
    
    for (config, aoa), coeff in coefficient_data.items():
        # Family: The common denominator (e.g., Geometry + Mesh)
        family = get_grouping_key(config, mode=comparison_mode)
        grid_status = get_grid_status(config)
        turb_model = coeff.get('turbulence_model', 'Unknown')
        
        # Attach grid status to coeff for later use
        coeff_with_grid = {**coeff, '_grid_status': grid_status, '_config': config}
        
        # Series Label: What distinguishes this line?
        if comparison_mode == 'turbulence':
            series_label = turb_model
        elif comparison_mode == 'expanded':
            # For expanded, series is turbulence model (primary grouping is by grid)
            series_label = turb_model
        elif comparison_mode == 'grid':
            series_label = grid_status
        elif comparison_mode == 'version':
            org_data = all_data.get((config, aoa), {})
            ver = org_data.get('version', 'V?')
            series_label = f"{ver}"
        else:  # default
            series_label = turb_model

        graphs_data[family][series_label].append(coeff_with_grid)
        
        # For expanded/grid modes, also group by grid status
        if comparison_mode in ('expanded', 'grid'):
            graphs_data_by_grid[family][grid_status][turb_model].append(coeff_with_grid)
    
    # Create graphs for each Family
    for family in sorted(graphs_data.keys()):
        series_dict = graphs_data[family]
        
        # === COMPARISON PLOTS ===
        if comparison_mode == 'expanded':
            # For expanded mode, create multiple comparison sets:
            
            grid_data = graphs_data_by_grid[family]
            
            # 1 & 2: Turbulence comparison per grid status
            for grid_status in ['With Grid', 'No Grid']:
                if grid_status not in grid_data:
                    continue
                    
                safe_grid = grid_status.replace(' ', '_')
                comp_dir = output_dir / "coefficient_graphs" / "Comparison" / f"Turbulence_{safe_grid}" / family
                comp_dir.mkdir(parents=True, exist_ok=True)
                
                plot_items = []
                color_cycle = itertools.cycle(colors)
                marker_cycle = itertools.cycle(markers)
                
                for turb_name in sorted(grid_data[grid_status].keys()):
                    items = sorted(grid_data[grid_status][turb_name], key=lambda x: x['aoa_degrees'])
                    if not items:
                        continue
                        
                    aoa_vals = np.array([x['aoa_degrees'] for x in items])
                    C_L_vals = np.array([x['C_L'] for x in items])
                    C_D_vals = np.array([x['C_D'] for x in items])
                    C_L_std = np.array([x.get('C_L_std', 0) for x in items])
                    C_D_std = np.array([x.get('C_D_std', 0) for x in items])
                    
                    plot_items.append({
                        'aoa': aoa_vals, 'C_L': C_L_vals, 'C_D': C_D_vals,
                        'C_L_std': C_L_std, 'C_D_std': C_D_std,
                        'style': {'label': turb_name, 'color': next(color_cycle), 'marker': next(marker_cycle)}
                    })
                
                if plot_items:
                    _plot_standard_aerodynamics(plot_items, comp_dir, f'{family} ({grid_status})', max_cov_threshold)
            
            # 3: Grid vs No Grid for each turbulence model
            all_turb_models = set()
            for grid_status in grid_data:
                all_turb_models.update(grid_data[grid_status].keys())
                
            for turb_name in sorted(all_turb_models):
                safe_turb = "".join([c for c in turb_name if c.isalnum() or c in (' ', '.', '_', '-')]).strip()
                comp_dir = output_dir / "coefficient_graphs" / "Comparison" / f"Grid_{safe_turb}" / family
                comp_dir.mkdir(parents=True, exist_ok=True)
                
                plot_items = []
                color_cycle = itertools.cycle(colors)
                marker_cycle = itertools.cycle(markers)
                
                for grid_status in ['With Grid', 'No Grid']:
                    if grid_status in grid_data and turb_name in grid_data[grid_status]:
                        items = sorted(grid_data[grid_status][turb_name], key=lambda x: x['aoa_degrees'])
                        if not items:
                            continue
                            
                        # Standard arrays
                        aoa_vals = np.array([x['aoa_degrees'] for x in items])
                        C_L_vals = np.array([x['C_L'] for x in items])
                        C_D_vals = np.array([x['C_D'] for x in items])
                        C_L_std = np.array([x.get('C_L_std', 0) for x in items])
                        C_D_std = np.array([x.get('C_D_std', 0) for x in items])
                        
                        plot_items.append({
                            'aoa': aoa_vals, 'C_L': C_L_vals, 'C_D': C_D_vals,
                            'C_L_std': C_L_std, 'C_D_std': C_D_std,
                            'style': {'label': grid_status, 'color': next(color_cycle), 'marker': next(marker_cycle)}
                        })
                
                if plot_items:
                    _plot_standard_aerodynamics(plot_items, comp_dir, f'{family} - {turb_name}', max_cov_threshold)
        
        else:
            # Standard comparison (all series on one graph)
            comp_dir = output_dir / "coefficient_graphs" / "Comparison" / family
            comp_dir.mkdir(parents=True, exist_ok=True)
            
            plot_items = []
            color_cycle = itertools.cycle(colors)
            marker_cycle = itertools.cycle(markers)
            
            for series_name in sorted(series_dict.keys()):
                items = sorted(series_dict[series_name], key=lambda x: x['aoa_degrees'])
                if not items:
                    continue
                    
                aoa_vals = np.array([x['aoa_degrees'] for x in items])
                C_L_vals = np.array([x['C_L'] for x in items])
                C_D_vals = np.array([x['C_D'] for x in items])
                C_L_std = np.array([x.get('C_L_std', 0) for x in items])
                C_D_std = np.array([x.get('C_D_std', 0) for x in items])
                
                plot_items.append({
                    'aoa': aoa_vals, 'C_L': C_L_vals, 'C_D': C_D_vals,
                    'C_L_std': C_L_std, 'C_D_std': C_D_std,
                    'style': {'label': series_name, 'color': next(color_cycle), 'marker': next(marker_cycle)},
                    # Store derived metrics base if needed, though they are recalculated inside
                    'endurance': items[0].get('endurance', []), # Assuming raw data has it or not
                    # But we also need CL_CD...
                    # Actually, plot_items has dicts with arrays.
                    # We can store raw metadata if we want.
                })
            
            if plot_items:
                # Determine title
                plot_title = family
                if comparison_mode == 'default':
                    # In default mode, 'family' is config string (e.g. 4.3.2.NG).
                    # All series have same turbulence model.
                    # Get it from first item of first series
                    try:
                        first_series_key = sorted(series_dict.keys())[0]
                        first_item = series_dict[first_series_key][0]
                        turb_model = first_item.get('turbulence_model', '')
                        if turb_model and turb_model != 'Unknown':
                            plot_title = f"{family} {turb_model}"
                    except (IndexError, KeyError):
                        pass
                
                # Standard Aerodynamics
                _plot_standard_aerodynamics(plot_items, comp_dir, plot_title, max_cov_threshold)

        # === SINGLE PLOTS (Individual Series) ===
        # For expanded/grid modes, organize by grid status first
        if comparison_mode in ('expanded', 'grid'):
            grid_data = graphs_data_by_grid[family]
            for grid_status in sorted(grid_data.keys()):
                safe_grid = grid_status.replace(' ', '_')
                for turb_name in sorted(grid_data[grid_status].keys()):
                    items = sorted(grid_data[grid_status][turb_name], key=lambda x: x['aoa_degrees'])
                    if not items:
                        continue
                        
                    safe_turb = "".join([c for c in turb_name if c.isalnum() or c in (' ', '.', '_', '-')]).strip()
                    single_dir = output_dir / "coefficient_graphs" / "Single" / safe_grid / safe_turb / family
                    single_dir.mkdir(parents=True, exist_ok=True)
                    
                    aoa_vals = np.array([x['aoa_degrees'] for x in items])
                    C_L_vals = np.array([x['C_L'] for x in items])
                    C_D_vals = np.array([x['C_D'] for x in items])
                    C_L_std = np.array([x.get('C_L_std', 0) for x in items])
                    C_D_std = np.array([x.get('C_D_std', 0) for x in items])
                    
                    # Create single item list
                    color_cycle = itertools.cycle(colors)
                    marker_cycle = itertools.cycle(markers)
                    item = {
                        'aoa': aoa_vals, 'C_L': C_L_vals, 'C_D': C_D_vals,
                        'C_L_std': C_L_std, 'C_D_std': C_D_std,
                        'style': {'label': turb_name, 'color': next(color_cycle), 'marker': next(marker_cycle)}
                    }
                    
                    _plot_standard_aerodynamics([item], single_dir, f'{family} - {turb_name} ({grid_status})', max_cov_threshold)
        else:
            # Standard single plots
            for series_name in sorted(series_dict.keys()):
                items = sorted(series_dict[series_name], key=lambda x: x['aoa_degrees'])
                if not items:
                    continue
                    
                safe_series = "".join([c for c in series_name if c.isalnum() or c in (' ', '.', '_', '-')]).strip()
                single_dir = output_dir / "coefficient_graphs" / "Single" / safe_series / family
                single_dir.mkdir(parents=True, exist_ok=True)
                
                aoa_vals = np.array([x['aoa_degrees'] for x in items])
                C_L_vals = np.array([x['C_L'] for x in items])
                C_D_vals = np.array([x['C_D'] for x in items])
                C_L_std = np.array([x.get('C_L_std', 0) for x in items])
                C_D_std = np.array([x.get('C_D_std', 0) for x in items])
                
                color_cycle = itertools.cycle(colors)
                marker_cycle = itertools.cycle(markers)
                item = {
                    'aoa': aoa_vals, 'C_L': C_L_vals, 'C_D': C_D_vals,
                    'C_L_std': C_L_std, 'C_D_std': C_D_std,
                    'style': {'label': series_name, 'color': next(color_cycle), 'marker': next(marker_cycle)}
                }
                
                _plot_standard_aerodynamics([item], single_dir, f'{family} ({series_name})', max_cov_threshold)

def _plot_multi_series(plot_items, val_key, title, ylabel, output_path, max_cov_threshold=None, x_key='aoa', xlabel='Angle of Attack $\\alpha$ [deg]'):
    """Refactored plotter for multiple series with premium academic style."""
    set_plot_style()
    plt.figure(figsize=(10, 7))
    
    points_removed = 0
    total_points = 0
    
    for item in plot_items:
        x = item[x_key]
        y = item[val_key]
        err = item.get(f'{val_key}_std')
        s = item['style']
        
        # Calculate stats for filtering if needed
        # item has C_L_std, C_L, etc.
        # We need to compute COV: (std / mean) * 100
        
        # Create a mask for valid data (including COV check)
        valid_mask = ~np.isnan(y)
        
        if max_cov_threshold is not None and err is not None:
             # Calculate COV array, handling division by zero
             with np.errstate(divide='ignore', invalid='ignore'):
                 cov_vals = np.abs((err / y) * 100)
             
             # Filter based on threshold
             cov_mask = (cov_vals <= max_cov_threshold) | np.isnan(cov_vals)
             
             points_in_series = np.sum(valid_mask)
             valid_mask = valid_mask & cov_mask
             points_kept = np.sum(valid_mask)
             removed = points_in_series - points_kept
             
             if removed > 0:
                 points_removed += removed
                 print(f"    [INFO]  Filtered {removed} points from {s['label']} (COV > {max_cov_threshold}%)")
        
        if not np.any(valid_mask): continue
        
        plt.errorbar(x[valid_mask], y[valid_mask], yerr=err[valid_mask] if err is not None else None, 
                     label=s['label'], 
                     color=s['color'], marker=s['marker'], 
                     capsize=3, linestyle='-', linewidth=2, markersize=8, alpha=0.9,
                     markeredgecolor='white', markeredgewidth=0.5)
                     
    plt.title(title, pad=20)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    
    # Modern legend and layout
    plt.legend(frameon=True, fancybox=True, framealpha=0.9, edgecolor='0.8', loc='best')
    plt.tight_layout()
    
    try:
        plt.savefig(output_path, bbox_inches='tight')
    except Exception as e:
        print(f"Warning: Failed to save graph {output_path}: {e}")
    plt.close()


def _calculate_derived_aerodynamics(plot_items):
    """Calculates L/D and Endurance Factor for plot items in place."""
    for item in plot_items:
        # Check if already calculated
        if 'L_D' in item: continue
        
        # Calculate derived metrics
        C_L = item['C_L']
        C_D = item['C_D']
        
        # Avoid division by zero
        with np.errstate(divide='ignore', invalid='ignore'):
             L_D = np.where(np.abs(C_D) > 1e-9, C_L / C_D, 0)
             # Endurance Factor: CL^1.5 / CD
             # Handle potential negative Lift (mostly for symmetry) by using abs(CL), though endurance implies positive lift.
             Endurance = np.where(np.abs(C_D) > 1e-9, (np.abs(C_L)**1.5) / C_D, 0)
             
        item['CL_CD'] = L_D
        item['Endurance'] = Endurance

def _plot_standard_aerodynamics(plot_items, output_dir, title_prefix, max_cov_threshold=None):
    """Generates standard set of aerodynamic plots: Lift, Drag, Polar, CL/CD, Endurance."""
    # Ensure derived metrics exist
    _calculate_derived_aerodynamics(plot_items)
    
    # Helper for consistent naming
    def make_name(suffix):
        clean_prefix = "".join([c for c in title_prefix if c.isalnum() or c in (' ', '.', '_', '-')]).strip().replace(' ', '_')
        return f"{clean_prefix}_{suffix}.png"

    # 1. C_L vs AoA
    _plot_multi_series(plot_items, 'C_L', f'{title_prefix} - Lift Coefficient', 
                      'Lift Coefficient ($C_L$)', output_dir / make_name("Lift_Coefficient"), max_cov_threshold)
    # 2. C_D vs AoA
    _plot_multi_series(plot_items, 'C_D', f'{title_prefix} - Drag Coefficient', 
                      'Drag Coefficient ($C_D$)', output_dir / make_name("Drag_Coefficient"), max_cov_threshold)
    # 3. Drag Polar (C_L vs C_D)
    _plot_multi_series(plot_items, 'C_L', f'{title_prefix} - Drag Polar', 
                      'Lift Coefficient ($C_L$)', output_dir / make_name("Drag_Polar"), max_cov_threshold,
                      x_key='C_D', xlabel='Drag Coefficient ($C_D$)')
    # 4. CL/CD vs AoA
    _plot_multi_series(plot_items, 'CL_CD', f'{title_prefix} - Lift-to-Drag Ratio', 
                      'Lift-to-Drag Ratio ($C_L/C_D$)', output_dir / make_name("Lift_to_Drag_Ratio"), max_cov_threshold)
    # 5. Endurance vs AoA
    _plot_multi_series(plot_items, 'Endurance', f'{title_prefix} - Endurance Factor', 
                      'Endurance Factor ($C_L^{1.5} / C_D$)', output_dir / make_name("Endurance_Factor"), max_cov_threshold)
    
    # 6. Combined CL/CD (Dual Axis)
    # Note: _plot_dual_axis_cl_cd generates its own filename currently. We should probably update it too or leave it consistent.
    # It uses "Combined_CL_CD_vs_AoA.png". Let's update it to match pattern inside the function if needed, 
    # but for now let's pass a specific output path if possible? The function doesn't take filename arg.
    # Let's rename the function to accept filename or handle it internally.
    # For now, let's just stick to the main ones.
    _plot_dual_axis_cl_cd(plot_items, output_dir, title_prefix, max_cov_threshold)
    
    # 7. Aerodynamic Summary (Quad Plot)
    _plot_quad_aerodynamics(plot_items, output_dir, title_prefix, max_cov_threshold)


def create_expanded_graphs(coefficient_data, output_dir, comparison_mode='expanded', max_cov_threshold=None):
    """
    Creates 'Expanded Graphs' plotting the ratio of Grid Efficiency / No-Grid Efficiency.
    Ratio = (L/D)_G / (L/D)_NG
    """
    # 1. Organize data for pairing
    # Structure: family_dict[geom_mesh_key][turb_model][aoa] = {'G': val, 'NG': val}
    pair_data = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    
    for (config, aoa), coeff in coefficient_data.items():
        # Identify Geometry + Mesh (Family)
        # Use the provided comparison mode for grouping
        family = get_grouping_key(config, mode=comparison_mode)
        
        # Identify Turbulence Model
        turb = coeff.get('turbulence_model', 'Unknown')
        
        # Identify Grid Status
        grid_meta = coeff.get('grid', 'Unknown')
        is_grid = False
        if '.G' in config or 'With Grid' in grid_meta or 'With Grid' in str(coeff):
            is_grid = True # It is the Grid version
        if '.NG' in config or 'No Grid' in grid_meta or 'No Grid' in str(coeff):
            is_grid = False # It is the No Grid version
            
        # Store
        key = 'G' if is_grid else 'NG'
        pair_data[family][turb][aoa][key] = coeff

    # 2. Calculate Ratios and Build Plot Series
    # We want to plot one line per Turbulence Model for each Family
    
    for family, turb_dict in pair_data.items():
        # Prepare plot items for this family
        plot_items = []
        
        
        # Setup output directory deferred until we have items
        expanded_dir = output_dir / "coefficient_graphs" / "Expanded_Graphs" / family
        
        for turb, aoa_dict in turb_dict.items():
            # For this turbulence model, collect (AoA, Ratio) points
            aoa_list = []
            ratio_list = []
            ratio_std_list = [] # We won't propagate error strictly for now, too complex?
            # Actually, standard error prop for division: R = A/B -> (dR/R)^2 = (dA/A)^2 + (dB/B)^2
            
            # Sort AoAs numerically (keys are strings like 'AoA_10')
            def extract_aoa_num(aoa_str):
                try:
                    return float(aoa_str.replace('AoA_', ''))
                except:
                    return 0
            sorted_aoas = sorted(aoa_dict.keys(), key=extract_aoa_num)
            for aoa in sorted_aoas:
                pair = aoa_dict[aoa]
                if 'G' in pair and 'NG' in pair:
                    g = pair['G']
                    ng = pair['NG']
                    
                    # Calculate Efficiency
                    # Handle near-zero Drag to avoid infinity?
                    if abs(g['C_D']) < 1e-6 or abs(ng['C_D']) < 1e-6:
                        continue
                        
                    eff_g = g['C_L'] / g['C_D']
                    eff_ng = ng['C_L'] / ng['C_D']
                    
                    if abs(eff_ng) < 1e-6:
                        continue
                        
                    ratio = eff_g / eff_ng
                    
                    aoa_list.append(extract_aoa_num(aoa))
                    ratio_list.append(ratio)
                    # ratio_std_list.append(0) # Todo: implement error prop if needed
            
            if not aoa_list:
                continue
                
            # Create series item
            # We need a style
            # Let's grab a color based on turbulence model name hash or cycle?
            # We can't easy cycle here because we loop turb by turb.
            # Let's map turb names to ACADEMIC_COLORS indices if possible, or just cycle.
            # For consistent colors:
            turb_idx = abs(hash(turb)) % len(ACADEMIC_COLORS)
            
            plot_items.append({
                'aoa': np.array(aoa_list),
                'Ratio': np.array(ratio_list),
                'Ratio_std': None, # Skipping error bars for ratio for now
                'style': {
                    'label': turb,
                    'color': ACADEMIC_COLORS[turb_idx],
                    'marker': ACADEMIC_MARKERS[turb_idx]
                }
            })
            
        if not plot_items:
            continue
            
        # 3. Plot
        expanded_dir.mkdir(parents=True, exist_ok=True)
        _plot_multi_series(plot_items, 'Ratio', 
                           f'{family} - Efficiency Improvement', 
                           'Efficiency Ratio ($(C_L/C_D)_G / (C_L/C_D)_{NG}$)',
                           expanded_dir / "Efficiency_Ratio_vs_AoA.png",
                           max_cov_threshold=max_cov_threshold)


def _plot_coefficient_vs_aoa(aoa_vals, coeff_vals, std_vals, style, turb_name, config, ylabel, title, output_path):
    """Helper function to plot coefficient vs AoA."""
    fig, ax = plt.subplots(figsize=(12, 8))
    
    ax.errorbar(aoa_vals, coeff_vals, yerr=std_vals,
                marker=style['marker'], markersize=10, linewidth=2.5, capsize=5,
                color=style['color'], label=turb_name, alpha=0.9)
    
    ax.set_xlabel('Angle of Attack (degrees)', fontsize=14, fontweight='bold')
    ax.set_ylabel(ylabel, fontsize=14, fontweight='bold')
    ax.set_title(f'{title}\n{config}', fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.legend(fontsize=12, loc='best', framealpha=0.9)
    ax.tick_params(labelsize=11)
    
    plt.tight_layout()
    try:
        plt.savefig(output_path, dpi=300)
    except Exception as e:
        print(f"Warning: Could not save graph {output_path}: {e}")
    plt.close()

def _plot_combined(aoa_vals, C_L_vals, C_D_vals, C_L_std_vals, C_D_std_vals, style, turb_name, config, output_path):
    """Helper function to plot combined C_L and C_D."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 8))
    
    # Left: C_L vs AoA
    ax1.errorbar(aoa_vals, C_L_vals, yerr=C_L_std_vals,
                marker=style['marker'], markersize=10, linewidth=2.5, capsize=5,
                color=style['color'], label=turb_name, alpha=0.9)
    ax1.set_xlabel('Angle of Attack (degrees)', fontsize=14, fontweight='bold')
    ax1.set_ylabel('Lift Coefficient ($C_L$)', fontsize=14, fontweight='bold')
    ax1.set_title(f'Lift Coefficient vs AoA\n{config}', fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3, linestyle='--')
    ax1.legend(fontsize=11, loc='best', framealpha=0.9)
    ax1.tick_params(labelsize=11)
    # Right: C_D vs AoA
    ax2.errorbar(aoa_vals, C_D_vals, yerr=C_D_std_vals,
                marker=style['marker'], markersize=10, linewidth=2.5, capsize=5,
                color=style['color'], label=turb_name, alpha=0.9)
    ax2.set_xlabel('Angle of Attack (degrees)', fontsize=14, fontweight='bold')
    ax2.set_ylabel('Drag Coefficient ($C_D$)', fontsize=14, fontweight='bold')
    ax2.set_title(f'Drag Coefficient vs AoA\n{config}', fontsize=14, fontweight='bold')
    ax2.grid(True, alpha=0.3, linestyle='--')
    ax2.legend(fontsize=11, loc='best', framealpha=0.9)
    ax2.tick_params(labelsize=11)
    
    plt.tight_layout()
    try:
        plt.savefig(output_path, dpi=300)
    except Exception as e:
        print(f"Warning: Could not save graph {output_path}: {e}")
    plt.close()

import matplotlib.pyplot as plt
from collections import defaultdict
import numpy as np

def plot_convergence_summary(convergence_results, all_data, output_dir, comparison_mode='turbulence'):
    """
    Plots convergence metrics (COV, Trim amount) vs AoA for different models/grids.
    Grouped by 'Family' (Geometry.Mesh) to allow comparison of Turb/Grid.
    """
    if not convergence_results:
        return

    # 1. Organize data by Family -> SimType -> AoA
    summary_data = defaultdict(lambda: defaultdict(dict))
    
    for (config, aoa), results in convergence_results.items():
        # Get data entry
        data = all_data.get((config, aoa))
        if not data: continue
        
        # Determine Family (Grouping)
        if comparison_mode == 'turbulence':
            # Group by Geom.Mesh.Grid
            # Family should be roughly config minus turbulence model
            # Let's use get_grouping_key but ensure grid is part of it if relevant
            # Or simpler: Geom.Mesh is the base, everything else is variations
            # Actually, standard is Geom.Mesh.Turb.Grid.
            # For turbulence comparison, we want to group by Geom.Mesh.Grid (so Turb varies)
            # The existing get_grouping_key(mode='turbulence') returns Geom.Mesh
            # If configurations differ by Grid (NG vs G), we separate them
            grid_suffix = ".G" if data.get('grid') in ['G', 'With Grid'] else ".NG" if data.get('grid') in ['NG', 'No Grid'] else ""
            base_family = get_grouping_key(config, mode='turbulence') + grid_suffix
            
            sim_type = data['turbulence_model']
            
        elif comparison_mode == 'grid':
            # Group by Geom.Mesh.Turb (so Grid varies)
            base_family = get_grouping_key(config, mode='grid')
            sim_type = "With Grid" if data.get('grid') in ['G', 'With Grid'] else "No Grid"
            
        else:
            base_family = get_grouping_key(config, mode='default')
            sim_type = "Default"

        # Extract metrics
        # Lift COV
        lift_cov_min_idx = np.argmin(results['lift']['cov'])
        lift_cov_min = results['lift']['cov'][lift_cov_min_idx]
        lift_trim_opt = results['lift']['iterations_removed'][lift_cov_min_idx]
        
        # Drag COV
        drag_cov_min_idx = np.argmin(results['drag']['cov'])
        drag_cov_min = results['drag']['cov'][drag_cov_min_idx]
        drag_trim_opt = results['drag']['iterations_removed'][drag_cov_min_idx]
        
        aoa_num = extract_aoa_number(aoa)
        
        summary_data[base_family][sim_type][aoa_num] = {
            'lift_cov': lift_cov_min,
            'lift_trim': lift_trim_opt,
            'drag_cov': drag_cov_min,
            'drag_trim': drag_trim_opt
        }

    # 2. Plot for each Family
    conv_dir = output_dir / "convergence_analysis" / "comparison_summaries"
    conv_dir.mkdir(parents=True, exist_ok=True)
    
    for family, sims in summary_data.items():
        # Setup Figure: 2x2
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        ((ax1, ax2), (ax3, ax4)) = axes
        
        # Colors/Markers
        markers = ['o', 's', '^', 'D', 'v', 'p', 'P', '*', 'h', 'X']
        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
        
        sorted_sim_types = sorted(sims.keys())
        
        legend_handles = []
        legend_labels = []
        
        for idx, sim_type in enumerate(sorted_sim_types):
            color = colors[idx % len(colors)]
            marker = markers[idx % len(markers)]
            
            sim_points = sims[sim_type]
            aoas = sorted(sim_points.keys())
            
            l_cov = [sim_points[a]['lift_cov'] for a in aoas]
            d_cov = [sim_points[a]['drag_cov'] for a in aoas]
            l_trim = [sim_points[a]['lift_trim'] for a in aoas]
            d_trim = [sim_points[a]['drag_trim'] for a in aoas]
            
            # Plot
            line, = ax1.plot(aoas, l_cov, linestyle='-', marker=marker, color=color, label=sim_type, alpha=0.8)
            ax2.plot(aoas, d_cov, linestyle='-', marker=marker, color=color, label=sim_type, alpha=0.8)
            ax3.plot(aoas, l_trim, linestyle='--', marker=marker, color=color, label=sim_type, alpha=0.8)
            ax4.plot(aoas, d_trim, linestyle='--', marker=marker, color=color, label=sim_type, alpha=0.8)
            
            legend_handles.append(line)
            legend_labels.append(sim_type)

        # Styling
        for ax in axes.flat:
            ax.grid(True, alpha=0.3)
            ax.set_xlabel(r'Angle of Attack $\alpha$ [deg]')
        
        ax1.set_ylabel('Min Lift COV (%)')
        ax1.set_title(f'Lift Convergence Stability\n{family}', fontweight='bold')
        
        ax2.set_ylabel('Min Drag COV (%)')
        ax2.set_title(f'Drag Convergence Stability\n{family}', fontweight='bold')
        
        ax3.set_ylabel('Optimal Iteration Trim')
        ax3.set_title(f'Lift Optimal Trim Amount\n{family}', fontweight='bold')
        
        ax4.set_ylabel('Optimal Iteration Trim')
        ax4.set_title(f'Drag Optimal Trim Amount\n{family}', fontweight='bold')
        
        fig.legend(legend_handles, legend_labels, loc='lower center', ncol=min(len(legend_labels), 5), bbox_to_anchor=(0.5, 0.02))
        plt.subplots_adjust(bottom=0.12, hspace=0.3)
        
        # Save
        plot_path = conv_dir / f"Convergence_Summary_{family}.png"
        try:
            plt.savefig(plot_path, dpi=300)
            print(f"    [OK] Generated summary plot: {plot_path.name}")
        except Exception as e:
            print(f"    [WARN] Failed to save summary plot {plot_path.name}: {e}")
        plt.close()

def _plot_dual_axis_cl_cd(plot_items, output_dir, title_prefix, max_cov_threshold=None):
    """Generates a dual-axis plot with CL on left and CD on right."""
    set_plot_style()
    fig, ax1 = plt.subplots(figsize=(10, 7))
    
    ax2 = ax1.twinx()
    
    # Colors for the two axes
    color_cl = '#1f77b4' # Blue
    color_cd = '#d62728' # Red
    
    has_data = False
    
    for item in plot_items:
        x = item['aoa']
        y_cl = item['C_L']
        y_cd = item['C_D']
        
        # Simple validity check
        mask_cl = ~np.isnan(y_cl)
        mask_cd = ~np.isnan(y_cd)
        valid_mask = mask_cl & mask_cd
        
        if not np.any(valid_mask): continue
        has_data = True
        
        label = item['style']['label']
        # If multiple series, use the assigned series color for both markers
        # to distinguish simulations. Otherwise, use hardcoded CL/CD colors.
        use_series_color = len(plot_items) > 1
        c_cl = item['style']['color'] if use_series_color else color_cl
        c_cd = item['style']['color'] if use_series_color else color_cd
        
        if use_series_color:
            cl_label = f"{label} ($C_L$)"
            cd_label = f"{label} ($C_D$)"
        else:
            cl_label = "$C_L$ (Lift)"
            cd_label = "$C_D$ (Drag)"
            
        # Plot CL (Solid)
        ax1.plot(x[valid_mask], y_cl[valid_mask], color=c_cl, marker='o', 
                 linestyle='-', linewidth=2, markersize=6, label=cl_label, alpha=0.9)
        
        # Plot CD (Dashed)
        ax2.plot(x[valid_mask], y_cd[valid_mask], color=c_cd, marker='s', 
                 linestyle='--', linewidth=2, markersize=6, label=cd_label, alpha=0.9)
                 
    if not has_data:
        plt.close()
        return

    ax1.set_xlabel('Angle of Attack $\\alpha$ [deg]')
    ax1.set_ylabel('Lift Coefficient ($C_L$)', color=color_cl, fontweight='bold')
    ax1.tick_params(axis='y', labelcolor=color_cl)
    
    ax2.set_ylabel('Drag Coefficient ($C_D$)', color=color_cd, fontweight='bold')
    ax2.tick_params(axis='y', labelcolor=color_cd)
    
    plt.title(f'{title_prefix} - Combined $C_L$ & $C_D$', pad=20)
    
    # Combine legends
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', frameon=True, fancybox=True, framealpha=0.9)
    
    plt.tight_layout()
    
    clean_prefix = "".join([c for c in title_prefix if c.isalnum() or c in (' ', '.', '_', '-')]).strip().replace(' ', '_')
    plot_path = output_dir / f"{clean_prefix}_Combined_CL_CD.png"
    try:
        plt.savefig(plot_path, bbox_inches='tight')
        # print(f"    [OK] Generated dual-axis plot: {plot_path.name}")
    except Exception as e:
        print(f"    [WARN] Failed to save dual-axis plot {plot_path.name}: {e}")
    plt.close()

def _plot_quad_aerodynamics(plot_items, output_dir, title_prefix, max_cov_threshold=None):
    """Generates a 2x2 summary plot: CL, CD, L/D, and Polar."""
    set_plot_style()
    fig, axs = plt.subplots(2, 2, figsize=(14, 10))
    fig.suptitle(f'{title_prefix} - Aerodynamic Summary', fontsize=16, fontweight='bold', y=0.98)
    
    # Subplot mapping
    ax_cl = axs[0, 0]
    ax_cd = axs[0, 1]
    ax_eff = axs[1, 0]
    ax_polar = axs[1, 1]
    
    # Track if we have any valid data
    has_data = False
    
    # Pre-calculate derived metrics just in case
    # Assuming _calculate_derived_aerodynamics was called before, but safe to check
    _calculate_derived_aerodynamics(plot_items)
    
    # Loop through items and plot on all 4 axes
    for item in plot_items:
        # Get data
        x_aoa = item['aoa']
        y_cl = item['C_L']
        y_cd = item['C_D']
        y_eff = item.get('CL_CD')
        
        # Style
        s = item['style']
        lbl = s['label']
        col = s['color']
        mrk = s['marker']
        
        # Determine valid mask based on COV if needed (or just valid data)
        # Assuming simple valid check for now
        mask_cl = ~np.isnan(y_cl)
        if np.any(mask_cl):
            has_data = True
            ax_cl.plot(x_aoa[mask_cl], y_cl[mask_cl], label=lbl, color=col, marker=mrk,
                       linestyle='-', linewidth=2, markersize=6, alpha=0.8)

        mask_cd = ~np.isnan(y_cd)
        if np.any(mask_cd):
            ax_cd.plot(x_aoa[mask_cd], y_cd[mask_cd], label=lbl, color=col, marker=mrk,
                       linestyle='-', linewidth=2, markersize=6, alpha=0.8)
                       
        if y_eff is not None:
            mask_eff = ~np.isnan(y_eff)
            if np.any(mask_eff):
                ax_eff.plot(x_aoa[mask_eff], y_eff[mask_eff], label=lbl, color=col, marker=mrk,
                           linestyle='-', linewidth=2, markersize=6, alpha=0.8)
                           
        # Drag Polar (CL vs CD)
        mask_polar = (~np.isnan(y_cl)) & (~np.isnan(y_cd))
        if np.any(mask_polar):
            # Sort by CD for cleaner line? Usually Polar is sorted by CL or something monotonic
            # Plotting as scatter + line based on index order (which is AoA order) works for polars usually
            ax_polar.plot(y_cd[mask_polar], y_cl[mask_polar], label=lbl, color=col, marker=mrk,
                          linestyle='-', linewidth=2, markersize=6, alpha=0.8)

    if not has_data:
        plt.close()
        return

    # Titles and Labels
    ax_cl.set_title("Lift Coefficient ($C_L$)", fontweight='bold')
    ax_cl.set_xlabel(r"Angle of Attack $\alpha$ [deg]")
    ax_cl.set_ylabel(r"$C_L$")
    ax_cl.grid(True, linestyle='--', alpha=0.6)
    
    ax_cd.set_title("Drag Coefficient ($C_D$)", fontweight='bold')
    ax_cd.set_xlabel(r"Angle of Attack $\alpha$ [deg]")
    ax_cd.set_ylabel(r"$C_D$")
    ax_cd.grid(True, linestyle='--', alpha=0.6)
    
    ax_eff.set_title("Lift-to-Drag Ratio ($C_L/C_D$)", fontweight='bold')
    ax_eff.set_xlabel(r"Angle of Attack $\alpha$ [deg]")
    ax_eff.set_ylabel(r"$C_L/C_D$")
    ax_eff.grid(True, linestyle='--', alpha=0.6)
    
    ax_polar.set_title("Drag Polar", fontweight='bold')
    ax_polar.set_xlabel(r"Drag Coefficient $C_D$")
    ax_polar.set_ylabel(r"Lift Coefficient $C_L$")
    ax_polar.grid(True, linestyle='--', alpha=0.6)

    # Global Legend (using handles from first plot)
    handles, labels = ax_cl.get_legend_handles_labels()
    if labels:
        fig.legend(handles, labels, loc='lower center', ncol=min(len(labels), 5), 
                   bbox_to_anchor=(0.5, 0.02), frameon=True, fancybox=True, shadow=True)
    
    plt.tight_layout()
    plt.subplots_adjust(bottom=0.15, top=0.92, hspace=0.3, wspace=0.25) # Adjust spacing
    
    clean_prefix = "".join([c for c in title_prefix if c.isalnum() or c in (' ', '.', '_', '-')]).strip().replace(' ', '_')
    plot_path = output_dir / f"{clean_prefix}_Aerodynamic_Summary.png"
    try:
        plt.savefig(plot_path, bbox_inches='tight')
    except Exception as e:
        print(f"    [WARN] Failed to save quad summary plot {plot_path.name}: {e}")
    plt.close()
