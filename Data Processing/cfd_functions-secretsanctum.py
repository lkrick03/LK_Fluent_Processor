"""
CFD Data Processing Functions
Contains all reusable functions for processing CFD simulation data.
"""

import os
import glob
import numpy as np
import pandas as pd
from collections import defaultdict
import pickle
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ==================== DATA VALIDATION FUNCTIONS ====================

def validate_aoa_folder(dirpath, filenames):
    """
    Validate that an AoA folder has required files with no duplicates or issues.
    
    Args:
        dirpath: Path to the AoA folder
        filenames: List of filenames in the folder
    
    Returns:
        Tuple of (is_valid, lift_file, drag_file, case_file, error_msg)
        - is_valid (bool): True if folder passes all checks
        - lift_file (str): Validated lift filename (None if invalid)
        - drag_file (str): Validated drag filename (None if invalid)
        - case_file (str): Validated case filename (None if invalid)
        - error_msg (str): Description of any issues (empty if valid)
    """
    errors = []
    
    # Find lift files
    lift_files = [f for f in filenames if 'lift_force' in f and f.endswith('.txt')]
    if len(lift_files) == 0:
        errors.append("No lift_force_*.txt file found")
    elif len(lift_files) > 1:
        errors.append(f"Multiple lift files found: {lift_files} - SKIPPING")
    
    # Find drag files
    drag_files = [f for f in filenames if 'drag_force' in f and f.endswith('.txt')]
    if len(drag_files) == 0:
        errors.append("No drag_force_*.txt file found")
    elif len(drag_files) > 1:
        errors.append(f"Multiple drag files found: {drag_files} - SKIPPING")
    
    # Find case file
    case_files = [f for f in filenames if f.endswith('.cas') or f.endswith('.cas.h5')]
    if len(case_files) == 0:
        errors.append("No case file (.cas or .cas.h5) found")
    elif len(case_files) > 1:
        errors.append(f"Multiple case files found: {case_files}")
    
    # Compile results
    is_valid = len(errors) == 0
    lift_file = lift_files[0] if len(lift_files) == 1 else None
    drag_file = drag_files[0] if len(drag_files) == 1 else None
    case_file = case_files[0] if len(case_files) >= 1 else None
    error_msg = " | ".join(errors) if errors else ""
    
    return is_valid, lift_file, drag_file, case_file, error_msg


# ==================== DATA LOADING FUNCTIONS ====================

def load_lift_drag_data(root_dir, config_extraction_method, position_map, value_mappings):
    """
    Load lift and drag force data from CFD simulation output files with robust validation.
    
    Args:
        root_dir: Root directory containing simulation data
        config_extraction_method: 'case_file' or 'folder'
        position_map: Dictionary mapping config positions to fields
        value_mappings: Dictionary mapping numeric codes to descriptive names
    
    Returns:
        Tuple of (data_dict, validation_report)
        - data_dict: Dictionary of data organized by (config, aoa) tuples
        - validation_report: Dict with statistics about what was loaded/skipped
    """
    data_by_config_aoa = defaultdict(lambda: {
        'lift': [], 'drag': [], 'turbulence_model': '', 
        'geometry': '', 'mesh': '', 'version': '', 'grid': ''
    })
    
    # Track validation statistics
    validation_report = {
        'total_aoa_folders': 0,
        'valid_aoa_folders': 0,
        'skipped_aoa_folders': 0,
        'issues': []  # List of (aoa_path, issue_description)
    }
    
    for dirpath, _, filenames in os.walk(root_dir):
        # Only process folders that contain AoA data
        if 'AoA_' not in dirpath:
            continue
        
        validation_report['total_aoa_folders'] += 1
        
        # Validate this AoA folder
        is_valid, lift_file, drag_file, case_file, error_msg = validate_aoa_folder(dirpath, filenames)
        
        if not is_valid:
            validation_report['skipped_aoa_folders'] += 1
            validation_report['issues'].append((dirpath, error_msg))
            continue
        
        validation_report['valid_aoa_folders'] += 1
        
        # Extract configuration from case file
        config = None
        if config_extraction_method == 'case_file' and case_file:
            config = case_file.replace('.cas.h5', '').replace('.cas', '')
            if not (config and config[0].isdigit() and '.' in config):
                config = None
        else:
            # Legacy: parse from folder structure
            parts = dirpath.split(os.sep)
            for part in parts:
                if part and part[0].isdigit() and part.count('.') >= 2:
                    config = part
                    break
        
        # Extract AoA value from folder name
        aoa = None
        parts = dirpath.split(os.sep)
        for part in parts:
            if part.startswith('AoA_'):
                aoa = part
                break
        
        if not config or not aoa:
            validation_report['skipped_aoa_folders'] += 1
            validation_report['issues'].append((dirpath, f"Could not extract config/AoA: config={config}, aoa={aoa}"))
            continue
        
        # Extract AoA number
        aoa_number = aoa.split('_')[1]
        
        # Handle old format embedded in filename
        if '_AoA_' in config:
            config = config.replace(f'_AoA_{aoa_number}', f'.{aoa_number}')
        elif config.count('.') == 0:
            config = f"{config}.{aoa_number}"
        
        # Parse configuration
        config_parts = config.split('.')
        positions = position_map
        mappings = value_mappings
        
        # Extract each field based on position
        geometry_num = None
        if positions['geometry'] is not None and len(config_parts) > positions['geometry']:
            try:
                geometry_num = int(config_parts[positions['geometry']])
            except ValueError:
                geometry_num = config_parts[positions['geometry']]
        
        mesh_num = None
        if positions['mesh'] is not None and len(config_parts) > positions['mesh']:
            try:
                mesh_num = int(config_parts[positions['mesh']])
            except ValueError:
                mesh_num = config_parts[positions['mesh']]
        
        turbulence_num = None
        if positions['turbulence'] is not None and len(config_parts) > positions['turbulence']:
            try:
                turbulence_num = int(config_parts[positions['turbulence']])
            except ValueError:
                turbulence_num = config_parts[positions['turbulence']]
        
        version_num = None
        if positions['version'] is not None and len(config_parts) > positions['version']:
            try:
                version_num = int(config_parts[positions['version']])
            except ValueError:
                version_num = config_parts[positions['version']]
        
        grid_code = config_parts[positions['grid']] if positions['grid'] is not None and len(config_parts) > positions['grid'] else None
        
        # Map to descriptive names
        geometry = mappings.get('geometry', {}).get(geometry_num, f"Geometry_{geometry_num}") if geometry_num else "N/A"
        mesh = mappings.get('mesh', {}).get(mesh_num, f"Mesh_{mesh_num}") if mesh_num else "N/A"
        turbulence_model = mappings.get('turbulence', {}).get(turbulence_num, f"Turbulence_{turbulence_num}") if turbulence_num else "Unknown"
        version = mappings.get('version', {}).get(version_num, f"Version_{version_num}") if version_num else "N/A"
        grid = mappings.get('grid', {}).get(grid_code, f"Grid_{grid_code}") if grid_code else "N/A"
        
        # Extract angle of attack
        aoa_degrees = float(aoa_number)
        aoa_radians = np.radians(aoa_degrees)
        
        # Read lift and drag data
        lift_file_path = os.path.join(dirpath, lift_file)
        drag_file_path = os.path.join(dirpath, drag_file)
        
        lift_data = _read_force_file(lift_file_path)
        drag_data = _read_force_file(drag_file_path)
        
        # Apply AoA correction
        cos_theta = np.cos(aoa_radians)
        sin_theta = np.sin(aoa_radians)
        
        true_lift_data = []
        true_drag_data = []
        
        min_length = min(len(lift_data), len(drag_data))
        for i in range(min_length):
            true_lift = lift_data[i] * cos_theta - drag_data[i] * sin_theta
            true_drag = lift_data[i] * sin_theta + drag_data[i] * cos_theta
            true_lift_data.append(true_lift)
            true_drag_data.append(true_drag)
        
        # Store data
        data_by_config_aoa[(config, aoa)]['lift'].extend(true_lift_data)
        data_by_config_aoa[(config, aoa)]['drag'].extend(true_drag_data)
        data_by_config_aoa[(config, aoa)]['geometry'] = geometry
        data_by_config_aoa[(config, aoa)]['mesh'] = mesh
        data_by_config_aoa[(config, aoa)]['turbulence_model'] = turbulence_model
        data_by_config_aoa[(config, aoa)]['version'] = version
        data_by_config_aoa[(config, aoa)]['grid'] = grid
    
    return data_by_config_aoa, validation_report


def _read_force_file(filepath):
    """Read force data from a text file, filtering out NaN values."""
    data = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line or '"' in line or '(' in line:
                continue
            parts = line.split()
            if len(parts) >= 2:
                try:
                    value = float(parts[1])
                    # Skip NaN values
                    if not np.isnan(value):
                        data.append(value)
                except (ValueError, IndexError):
                    continue
    return data


def find_latest_pickle(search_roots, filename='processed_data.pkl'):
    """
    Find the most recently modified pickle file.
    
    Args:
        search_roots: List of root directories to search
        filename: Name of pickle file to find
    
    Returns:
        Path to most recent pickle file, or None if not found
    """
    pickle_files = []
    for root in search_roots:
        if os.path.exists(root):
            search_pattern = os.path.join(root, "**", filename)
            found = glob.glob(search_pattern, recursive=True)
            pickle_files.extend(found)
            if found:
                break
    
    if pickle_files:
        return max(pickle_files, key=os.path.getmtime)
    return None


# ==================== STATISTICS FUNCTIONS ====================

def compute_statistics(data):
    """Calculate mean and coefficient of variation."""
    mean_val = np.mean(data)
    std_dev = np.std(data)
    cov = (std_dev / mean_val * 100) if mean_val != 0 else 0
    return mean_val, cov


def extract_aoa_number(aoa_string):
    """Extract numeric AoA from string like 'AoA_10'."""
    try:
        return int(aoa_string.split('_')[1])
    except:
        return 0


# ==================== CONVERGENCE ANALYSIS FUNCTIONS ====================

def analyze_convergence(data_array, min_trim=0, max_trim=0.5, num_tests=10, min_retention=0.2):
    """
    Analyze convergence with safety checks and confidence scoring.
    
    Args:
        data_array: numpy array of lift or drag values
        min_trim: minimum fraction to remove from start (0 to 1)
        max_trim: maximum fraction to remove from start (0 to 1)
        num_tests: number of trim amounts to test
        min_retention: Safety floor - don't recommend keeping less than this fraction (default 0.2 = 20%)
    
    Returns:
        Dictionary with:
        - Standard metrics: iterations_removed, iterations_used, mean, std_dev, cov
        - New metrics: median, mad, sign_changes, confidence_score
        - Recommendations: trim_recommendation, trim_reason
        - Warnings: issues detected during analysis
    """
    total_iterations = len(data_array)
    trim_fractions = np.linspace(min_trim, max_trim, num_tests)
    
    results = {
        'iterations_removed': [],
        'iterations_used': [],
        'mean': [],
        'std_dev': [],
        'median': [],
        'mad': [],  # Median Absolute Deviation
        'cov': [],
        'sign_changes': [],  # Flag if mean changes sign (oscillating through zero)
        'confidence_score': [],  # 0-1 score: how confident we are in this trim point
        'trim_recommendation': None,
        'trim_reason': None,
        'warnings': []
    }
    
    prev_mean_sign = None
    
    for trim_frac in trim_fractions:
        iterations_to_remove = int(total_iterations * trim_frac)
        trimmed_data = data_array[iterations_to_remove:]
        
        if len(trimmed_data) < 5:
            results['warnings'].append(
                f"Trim {trim_frac:.1%}: Only {len(trimmed_data)} points left (too few for statistics)"
            )
            continue
        
        trimmed_data_np = np.asarray(trimmed_data, dtype=np.float64)
        
        # Calculate statistics
        mean_val = np.mean(trimmed_data_np)
        std_val = np.std(trimmed_data_np)
        median_val = np.median(trimmed_data_np)
        mad_val = np.median(np.abs(trimmed_data_np - median_val))  # Median Absolute Deviation
        
        # Calculate COV (coefficient of variation)
        if mean_val != 0:
            cov_val = (std_val / abs(mean_val) * 100)
        else:
            cov_val = np.inf
            results['warnings'].append(
                f"Trim {trim_frac:.1%}: Mean is zero (stall condition). Using median/MAD instead."
            )
        
        # Detect sign changes (data oscillating through zero)
        current_sign = np.sign(mean_val)
        sign_changed = False
        if prev_mean_sign is not None and prev_mean_sign != current_sign:
            sign_changed = True
            results['warnings'].append(
                f"Trim {trim_frac:.1%}: Mean changed sign (oscillating). Data not fully settled."
            )
        prev_mean_sign = current_sign
        
        # Calculate confidence score (0-1, higher is more confident)
        # Weighted factors: 60% on low COV, 30% on data retention, 10% penalty for oscillation
        data_retention_ratio = len(trimmed_data) / total_iterations
        
        # Normalize COV to 0-1 scale (>50% COV = very uncertain, <2% = very confident)
        cov_normalized = min(max(cov_val / 50, 0), 1) if np.isfinite(cov_val) else 1.0
        
        # Confidence = 60% weight on low COV, 30% on data retention, 10% penalty for oscillation
        oscillation_penalty = 0.5 if sign_changed else 0.0
        confidence = 0.6 * (1 - cov_normalized) + 0.3 * data_retention_ratio - oscillation_penalty
        confidence = max(0, min(1, confidence))  # Clamp to [0, 1]
        
        results['iterations_removed'].append(iterations_to_remove)
        results['iterations_used'].append(len(trimmed_data))
        results['mean'].append(mean_val)
        results['std_dev'].append(std_val)
        results['median'].append(median_val)
        results['mad'].append(mad_val)
        results['cov'].append(cov_val)
        results['sign_changes'].append(sign_changed)
        results['confidence_score'].append(confidence)
    
    # ============ GENERATE RECOMMENDATION ============
    if len(results['cov']) > 0:
        # Find minimum COV (best convergence)
        valid_cov_indices = [i for i, c in enumerate(results['cov']) if np.isfinite(c)]
        
        if valid_cov_indices:
            min_cov_idx = min(valid_cov_indices, key=lambda i: results['cov'][i])
            trim_at_min_cov = results['iterations_used'][min_cov_idx] / total_iterations
            
            # Check if recommendation keeps enough data (safety check)
            if trim_at_min_cov >= min_retention:
                results['trim_recommendation'] = results['iterations_removed'][min_cov_idx]
                confidence_pct = results['confidence_score'][min_cov_idx] * 100
                results['trim_reason'] = (
                    f"Minimum COV ({results['cov'][min_cov_idx]:.2f}%) at "
                    f"{trim_fractions[min_cov_idx]:.1%} trim. Confidence: {confidence_pct:.0f}%"
                )
            else:
                # Over-trimming warning
                results['warnings'].append(
                    f"Best COV ({results['cov'][min_cov_idx]:.2f}%) requires trimming "
                    f"{trim_fractions[min_cov_idx]:.1%} (keeping only {trim_at_min_cov:.1%}). "
                    f"Below safety threshold ({min_retention:.0%}). Using conservative trim instead."
                )
                
                # Find best conservative trim (keeps at least min_retention of data)
                conservative_indices = [
                    i for i in valid_cov_indices 
                    if results['iterations_used'][i] / total_iterations >= min_retention
                ]
                
                if conservative_indices:
                    conservative_best_idx = min(conservative_indices, key=lambda i: results['cov'][i])
                    results['trim_recommendation'] = results['iterations_removed'][conservative_best_idx]
                    confidence_pct = results['confidence_score'][conservative_best_idx] * 100
                    results['trim_reason'] = (
                        f"Conservative trim keeping {results['iterations_used'][conservative_best_idx] / total_iterations:.1%} data. "
                        f"COV: {results['cov'][conservative_best_idx]:.2f}%, Confidence: {confidence_pct:.0f}%"
                    )
    
    return results


def plot_convergence_analysis(config, aoa, lift_data, drag_data, output_dir, max_trim, num_tests):
    """
    Create convergence analysis plots showing how statistics change with data trimming.
    
    Args:
        config: Configuration string
        aoa: Angle of attack string
        lift_data: Array of lift force values
        drag_data: Array of drag force values
        output_dir: Directory to save plots
        max_trim: Maximum fraction of data to trim
        num_tests: Number of trim amounts to test
    
    Returns:
        Tuple of (lift_results, drag_results, plot_path)
    """
    # Analyze both lift and drag
    lift_results = analyze_convergence(np.array(lift_data), min_trim=0, max_trim=max_trim, num_tests=num_tests)
    drag_results = analyze_convergence(np.array(drag_data), min_trim=0, max_trim=max_trim, num_tests=num_tests)
    
    # Create figure with subplots
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
    
    # Plot 1: Lift Mean vs Iterations Removed
    ax1.plot(lift_results['iterations_removed'], lift_results['mean'], 'o-', linewidth=2, markersize=8, color='#1f77b4')
    ax1.set_xlabel('Iterations Removed from Start', fontsize=12)
    ax1.set_ylabel('Lift Mean (N)', fontsize=12)
    ax1.set_title(f'Lift Mean Convergence\n{config} - {aoa}', fontweight='bold', fontsize=14)
    ax1.grid(True, alpha=0.3)
    
    # Plot 2: Lift COV vs Iterations Removed
    ax2.plot(lift_results['iterations_removed'], lift_results['cov'], 'o-', linewidth=2, markersize=8, color='#ff7f0e')
    ax2.set_xlabel('Iterations Removed from Start', fontsize=12)
    ax2.set_ylabel('Lift COV (%)', fontsize=12)
    ax2.set_title(f'Lift COV Convergence\n{config} - {aoa}', fontweight='bold', fontsize=14)
    ax2.grid(True, alpha=0.3)
    
    # Highlight minimum COV point for lift
    min_cov_idx = np.argmin(lift_results['cov'])
    ax2.axvline(x=lift_results['iterations_removed'][min_cov_idx], color='red', linestyle='--', linewidth=2, alpha=0.7)
    ax2.text(lift_results['iterations_removed'][min_cov_idx], max(lift_results['cov']), 
             f"  Min COV\n  Remove: {lift_results['iterations_removed'][min_cov_idx]}\n  Use: {lift_results['iterations_used'][min_cov_idx]}", 
             color='red', fontweight='bold', fontsize=9)
    
    # Plot 3: Drag Mean vs Iterations Removed
    ax3.plot(drag_results['iterations_removed'], drag_results['mean'], 'o-', linewidth=2, markersize=8, color='#2ca02c')
    ax3.set_xlabel('Iterations Removed from Start', fontsize=12)
    ax3.set_ylabel('Drag Mean (N)', fontsize=12)
    ax3.set_title(f'Drag Mean Convergence\n{config} - {aoa}', fontweight='bold', fontsize=14)
    ax3.grid(True, alpha=0.3)
    
    # Plot 4: Drag COV vs Iterations Removed
    ax4.plot(drag_results['iterations_removed'], drag_results['cov'], 'o-', linewidth=2, markersize=8, color='#d62728')
    ax4.set_xlabel('Iterations Removed from Start', fontsize=12)
    ax4.set_ylabel('Drag COV (%)', fontsize=12)
    ax4.set_title(f'Drag COV Convergence\n{config} - {aoa}', fontweight='bold', fontsize=14)
    ax4.grid(True, alpha=0.3)
    
    # Highlight minimum COV point for drag
    min_cov_idx = np.argmin(drag_results['cov'])
    ax4.axvline(x=drag_results['iterations_removed'][min_cov_idx], color='red', linestyle='--', linewidth=2, alpha=0.7)
    ax4.text(drag_results['iterations_removed'][min_cov_idx], max(drag_results['cov']), 
             f"  Min COV\n  Remove: {drag_results['iterations_removed'][min_cov_idx]}\n  Use: {drag_results['iterations_used'][min_cov_idx]}", 
             color='red', fontweight='bold', fontsize=9)
    
    plt.tight_layout()
    
    # Save convergence analysis plot
    convergence_dir = os.path.join(output_dir, "convergence_analysis")
    os.makedirs(convergence_dir, exist_ok=True)
    
    plot_file = os.path.join(convergence_dir, f"convergence_{config}_{aoa}.png")
    plt.savefig(plot_file, dpi=300, bbox_inches='tight')
    plt.close()
    
    return lift_results, drag_results, plot_file


# ==================== EXCEL EXPORT FUNCTIONS ====================

def create_data_summary_sheet(wb, all_data, num_iterations, convergence_results):
    """Create Data Summary sheet in Excel with formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    group_header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    group_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    row_fill_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Helper function to get optimized data
    def get_optimized_data(config, aoa, data):
        if convergence_results and (config, aoa) in convergence_results:
            conv = convergence_results[(config, aoa)]
            lift_min_cov_idx = np.argmin(conv['lift']['cov'])
            drag_min_cov_idx = np.argmin(conv['drag']['cov'])
            
            optimal_lift_trim = conv['lift']['iterations_removed'][lift_min_cov_idx]
            optimal_drag_trim = conv['drag']['iterations_removed'][drag_min_cov_idx]
            optimal_trim = max(optimal_lift_trim, optimal_drag_trim)
            
            lift_values = data['lift'][optimal_trim:]
            drag_values = data['drag'][optimal_trim:]
        else:
            lift_values = data['lift'][-num_iterations:] if len(data['lift']) >= num_iterations else data['lift']
            drag_values = data['drag'][-num_iterations:] if len(data['drag']) >= num_iterations else data['drag']
        return lift_values, drag_values
    
    # Group data by base configuration
    base_config_data = defaultdict(list)
    for (config, aoa), data in all_data.items():
        config_parts = config.split('.')
        if len(config_parts) > 1:
            base_config = '.'.join(config_parts[:-1])
        else:
            base_config = config
        
        lift_values, drag_values = get_optimized_data(config, aoa, data)
        
        lift_mean = np.mean(lift_values) if lift_values else 0
        lift_std = np.std(lift_values) if lift_values else 0
        lift_cov = (lift_std / lift_mean * 100) if lift_mean != 0 else 0
        
        drag_mean = np.mean(drag_values) if drag_values else 0
        drag_std = np.std(drag_values) if drag_values else 0
        drag_cov = (drag_std / drag_mean * 100) if drag_mean != 0 else 0
        
        base_config_data[base_config].append({
            'aoa': aoa,
            'aoa_num': extract_aoa_number(aoa),
            'turbulence_model': data['turbulence_model'],
            'num_points': len(lift_values),
            'lift_mean': lift_mean,
            'lift_cov': lift_cov,
            'drag_mean': drag_mean,
            'drag_cov': drag_cov
        })
    
    # Sort each group by AoA
    for base_config in base_config_data:
        base_config_data[base_config].sort(key=lambda x: x['aoa_num'])
    
    # Sort base configs by turbulence model
    model_order = {'SST': 1, 'RNG': 2, 'RSM': 3, 'k-epsilon': 4}
    def get_sort_key(base_config):
        if base_config in base_config_data and base_config_data[base_config]:
            turb_model = base_config_data[base_config][0]['turbulence_model']
            return (model_order.get(turb_model, 99), base_config)
        return (99, base_config)
    sorted_base_configs = sorted(base_config_data.keys(), key=get_sort_key)
    
    # Create worksheet
    ws = wb.active
    ws.title = 'Data Summary'
    
    columns = ['Turbulence Model', 'Angle of Attack', 'Lift Mean (N)', 'Lift COV (%)', 
               'Drag Mean (N)', 'Drag COV (%)', 'Num Points']
    current_row = 1
    
    for base_config in sorted_base_configs:
        data_rows = base_config_data[base_config]
        if not data_rows:
            continue
        
        # Section header
        ws.cell(row=current_row, column=1, value=base_config)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
        header_cell = ws.cell(row=current_row, column=1)
        header_cell.font = group_header_font
        header_cell.fill = group_header_fill
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border_style
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Column headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Data rows
        for row_idx, row_data in enumerate(data_rows):
            fill_color = row_fill_light if row_idx % 2 == 0 else row_fill_white
            
            values = [
                row_data['turbulence_model'],
                row_data['aoa'],
                f"{row_data['lift_mean']:.1f}",
                f"{row_data['lift_cov']:.1f}",
                f"{row_data['drag_mean']:.1f}",
                f"{row_data['drag_cov']:.1f}",
                row_data['num_points']
            ]
            
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.alignment = data_alignment
                cell.border = border_style
                cell.fill = fill_color
            
            current_row += 1
        
        current_row += 1  # Blank row
    
    # Autofit columns
    for col_idx in range(1, len(columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_turbulence_comparison_sheet(wb, all_data, num_iterations, convergence_results):
    """Create Turbulence Comparison sheet with 4 tables."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    ws = wb.create_sheet(title='Turbulence Comparison')
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    table_title_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    table_title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    row_fill_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Helper function
    def get_optimized_data(config, aoa, data):
        if convergence_results and (config, aoa) in convergence_results:
            conv = convergence_results[(config, aoa)]
            lift_min_cov_idx = np.argmin(conv['lift']['cov'])
            drag_min_cov_idx = np.argmin(conv['drag']['cov'])
            
            optimal_lift_trim = conv['lift']['iterations_removed'][lift_min_cov_idx]
            optimal_drag_trim = conv['drag']['iterations_removed'][drag_min_cov_idx]
            optimal_trim = max(optimal_lift_trim, optimal_drag_trim)
            
            lift_values = data['lift'][optimal_trim:]
            drag_values = data['drag'][optimal_trim:]
        else:
            lift_values = data['lift'][-num_iterations:] if len(data['lift']) >= num_iterations else data['lift']
            drag_values = data['drag'][-num_iterations:] if len(data['drag']) >= num_iterations else data['drag']
        return lift_values, drag_values
    
    # Group by base config and turbulence model
    turbulence_data = defaultdict(lambda: defaultdict(dict))
    
    for (config, aoa), data in all_data.items():
        config_parts = config.split('.')
        base_config = '.'.join(config_parts[:-1]) if len(config_parts) > 1 else config
        turb_model = data['turbulence_model']
        
        lift_values, drag_values = get_optimized_data(config, aoa, data)
        
        lift_mean = np.mean(lift_values) if len(lift_values) > 0 else 0
        lift_std = np.std(lift_values) if len(lift_values) > 0 else 0
        lift_cov = (lift_std / lift_mean * 100) if lift_mean != 0 else 0
        
        drag_mean = np.mean(drag_values) if len(drag_values) > 0 else 0
        drag_std = np.std(drag_values) if len(drag_values) > 0 else 0
        drag_cov = (drag_std / drag_mean * 100) if drag_mean != 0 else 0
        
        turbulence_data[base_config][turb_model][aoa] = {
            'lift_mean': lift_mean,
            'lift_cov': lift_cov,
            'drag_mean': drag_mean,
            'drag_cov': drag_cov,
            'aoa_num': extract_aoa_number(aoa)
        }
    
    # Get sorted AoAs
    all_aoas = set()
    for base_config in turbulence_data:
        for turb_model in turbulence_data[base_config]:
            all_aoas.update(turbulence_data[base_config][turb_model].keys())
    sorted_aoas = sorted(all_aoas, key=extract_aoa_number)
    
    current_row = 1
    
    # Create 4 tables for each base configuration
    for base_config in sorted(turbulence_data.keys()):
        models_in_config = turbulence_data[base_config]
        
        # Table 1: Lift Mean
        ws.cell(row=current_row, column=1, value=f"{base_config} - Lift Mean (N)")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sorted_aoas)+1)
        cell = ws.cell(row=current_row, column=1)
        cell.font = table_title_font
        cell.fill = table_title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Headers
        ws.cell(row=current_row, column=1, value='Turbulence Model').font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).alignment = header_alignment
        ws.cell(row=current_row, column=1).border = border_style
        
        for col_idx, aoa in enumerate(sorted_aoas, 2):
            ws.cell(row=current_row, column=col_idx, value=aoa).font = header_font
            ws.cell(row=current_row, column=col_idx).fill = header_fill
            ws.cell(row=current_row, column=col_idx).alignment = header_alignment
            ws.cell(row=current_row, column=col_idx).border = border_style
        current_row += 1
        
        # Data rows
        sst_data = models_in_config.get('SST', {})
        for row_idx, turb_model in enumerate(['SST', 'RNG', 'RSM', 'k-epsilon']):
            if turb_model not in models_in_config:
                continue
            
            fill = row_fill_light if row_idx % 2 == 0 else row_fill_white
            
            ws.cell(row=current_row, column=1, value=turb_model).alignment = data_alignment
            ws.cell(row=current_row, column=1).border = border_style
            ws.cell(row=current_row, column=1).fill = fill
            
            for col_idx, aoa in enumerate(sorted_aoas, 2):
                if aoa in models_in_config[turb_model]:
                    lift_mean = models_in_config[turb_model][aoa]['lift_mean']
                    
                    # Calculate % difference from SST
                    if turb_model != 'SST' and aoa in sst_data:
                        sst_lift = sst_data[aoa]['lift_mean']
                        if sst_lift != 0:
                            pct_diff = ((lift_mean - sst_lift) / sst_lift) * 100
                            value = f"{lift_mean:.1f} ({pct_diff:+.1f}%)"
                        else:
                            value = f"{lift_mean:.1f}"
                    else:
                        value = f"{lift_mean:.1f}"
                    
                    ws.cell(row=current_row, column=col_idx, value=value).alignment = data_alignment
                    ws.cell(row=current_row, column=col_idx).border = border_style
                    ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1
        current_row += 2
        
        # Table 2: Drag Mean
        ws.cell(row=current_row, column=1, value=f"{base_config} - Drag Mean (N)")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sorted_aoas)+1)
        cell = ws.cell(row=current_row, column=1)
        cell.font = table_title_font
        cell.fill = table_title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Headers
        ws.cell(row=current_row, column=1, value='Turbulence Model').font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).alignment = header_alignment
        ws.cell(row=current_row, column=1).border = border_style
        
        for col_idx, aoa in enumerate(sorted_aoas, 2):
            ws.cell(row=current_row, column=col_idx, value=aoa).font = header_font
            ws.cell(row=current_row, column=col_idx).fill = header_fill
            ws.cell(row=current_row, column=col_idx).alignment = header_alignment
            ws.cell(row=current_row, column=col_idx).border = border_style
        current_row += 1
        
        # Data rows
        for row_idx, turb_model in enumerate(['SST', 'RNG', 'RSM', 'k-epsilon']):
            if turb_model not in models_in_config:
                continue
            
            fill = row_fill_light if row_idx % 2 == 0 else row_fill_white
            
            ws.cell(row=current_row, column=1, value=turb_model).alignment = data_alignment
            ws.cell(row=current_row, column=1).border = border_style
            ws.cell(row=current_row, column=1).fill = fill
            
            for col_idx, aoa in enumerate(sorted_aoas, 2):
                if aoa in models_in_config[turb_model]:
                    drag_mean = models_in_config[turb_model][aoa]['drag_mean']
                    
                    if turb_model != 'SST' and aoa in sst_data:
                        sst_drag = sst_data[aoa]['drag_mean']
                        if sst_drag != 0:
                            pct_diff = ((drag_mean - sst_drag) / sst_drag) * 100
                            value = f"{drag_mean:.1f} ({pct_diff:+.1f}%)"
                        else:
                            value = f"{drag_mean:.1f}"
                    else:
                        value = f"{drag_mean:.1f}"
                    
                    ws.cell(row=current_row, column=col_idx, value=value).alignment = data_alignment
                    ws.cell(row=current_row, column=col_idx).border = border_style
                    ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1
        current_row += 2
        
        # Table 3: Lift COV
        ws.cell(row=current_row, column=1, value=f"{base_config} - Lift COV (%)")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sorted_aoas)+1)
        cell = ws.cell(row=current_row, column=1)
        cell.font = table_title_font
        cell.fill = table_title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Headers
        ws.cell(row=current_row, column=1, value='Turbulence Model').font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).alignment = header_alignment
        ws.cell(row=current_row, column=1).border = border_style
        
        for col_idx, aoa in enumerate(sorted_aoas, 2):
            ws.cell(row=current_row, column=col_idx, value=aoa).font = header_font
            ws.cell(row=current_row, column=col_idx).fill = header_fill
            ws.cell(row=current_row, column=col_idx).alignment = header_alignment
            ws.cell(row=current_row, column=col_idx).border = border_style
        current_row += 1
        
        # Data rows
        for row_idx, turb_model in enumerate(['SST', 'RNG', 'RSM', 'k-epsilon']):
            if turb_model not in models_in_config:
                continue
            
            fill = row_fill_light if row_idx % 2 == 0 else row_fill_white
            
            ws.cell(row=current_row, column=1, value=turb_model).alignment = data_alignment
            ws.cell(row=current_row, column=1).border = border_style
            ws.cell(row=current_row, column=1).fill = fill
            
            for col_idx, aoa in enumerate(sorted_aoas, 2):
                if aoa in models_in_config[turb_model]:
                    lift_cov = models_in_config[turb_model][aoa]['lift_cov']
                    value = f"{lift_cov:.1f}"
                    
                    ws.cell(row=current_row, column=col_idx, value=value).alignment = data_alignment
                    ws.cell(row=current_row, column=col_idx).border = border_style
                    ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1
        current_row += 2
        
        # Table 4: Drag COV
        ws.cell(row=current_row, column=1, value=f"{base_config} - Drag COV (%)")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sorted_aoas)+1)
        cell = ws.cell(row=current_row, column=1)
        cell.font = table_title_font
        cell.fill = table_title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Headers
        ws.cell(row=current_row, column=1, value='Turbulence Model').font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).alignment = header_alignment
        ws.cell(row=current_row, column=1).border = border_style
        
        for col_idx, aoa in enumerate(sorted_aoas, 2):
            ws.cell(row=current_row, column=col_idx, value=aoa).font = header_font
            ws.cell(row=current_row, column=col_idx).fill = header_fill
            ws.cell(row=current_row, column=col_idx).alignment = header_alignment
            ws.cell(row=current_row, column=col_idx).border = border_style
        current_row += 1
        
        # Data rows
        for row_idx, turb_model in enumerate(['SST', 'RNG', 'RSM', 'k-epsilon']):
            if turb_model not in models_in_config:
                continue
            
            fill = row_fill_light if row_idx % 2 == 0 else row_fill_white
            
            ws.cell(row=current_row, column=1, value=turb_model).alignment = data_alignment
            ws.cell(row=current_row, column=1).border = border_style
            ws.cell(row=current_row, column=1).fill = fill
            
            for col_idx, aoa in enumerate(sorted_aoas, 2):
                if aoa in models_in_config[turb_model]:
                    drag_cov = models_in_config[turb_model][aoa]['drag_cov']
                    value = f"{drag_cov:.1f}"
                    
                    ws.cell(row=current_row, column=col_idx, value=value).alignment = data_alignment
                    ws.cell(row=current_row, column=col_idx).border = border_style
                    ws.cell(row=current_row, column=col_idx).fill = fill
            current_row += 1
        current_row += 3
    
    # Autofit columns
    for col_idx in range(1, len(sorted_aoas) + 2):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_coefficients_sheet(wb, all_data, num_iterations, convergence_results, q_times_a):
    """Create Coefficients sheet with C_L and C_D."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    ws = wb.create_sheet(title='Coefficients')
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    row_fill_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Headers
    columns = ['Configuration', 'Turbulence Model', 'AoA', 'C_L', 'C_L COV (%)', 'C_D', 'C_D COV (%)']
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    ws.row_dimensions[1].height = 30
    
    # Data rows
    row = 2
    for (config, aoa), data in sorted(all_data.items()):
        # Get optimized or fixed iteration data
        if convergence_results and (config, aoa) in convergence_results:
            conv = convergence_results[(config, aoa)]
            lift_min_cov_idx = np.argmin(conv['lift']['cov'])
            drag_min_cov_idx = np.argmin(conv['drag']['cov'])
            
            optimal_lift_trim = conv['lift']['iterations_removed'][lift_min_cov_idx]
            optimal_drag_trim = conv['drag']['iterations_removed'][drag_min_cov_idx]
            optimal_trim = max(optimal_lift_trim, optimal_drag_trim)
            
            lift_values = data['lift'][optimal_trim:]
            drag_values = data['drag'][optimal_trim:]
        else:
            lift_values = data['lift'][-num_iterations:] if len(data['lift']) >= num_iterations else data['lift']
            drag_values = data['drag'][-num_iterations:] if len(data['drag']) >= num_iterations else data['drag']
        
        lift_mean = np.mean(lift_values) if len(lift_values) > 0 else 0
        drag_mean = np.mean(drag_values) if len(drag_values) > 0 else 0
        lift_std = np.std(lift_values) if len(lift_values) > 0 else 0
        drag_std = np.std(drag_values) if len(drag_values) > 0 else 0
        
        C_L = lift_mean / q_times_a if q_times_a != 0 else 0
        C_D = drag_mean / q_times_a if q_times_a != 0 else 0
        C_L_std = lift_std / q_times_a if q_times_a != 0 else 0
        C_D_std = drag_std / q_times_a if q_times_a != 0 else 0
        C_L_cov = (C_L_std / C_L * 100) if C_L != 0 else 0
        C_D_cov = (C_D_std / C_D * 100) if C_D != 0 else 0
        
        fill = row_fill_light if row % 2 == 0 else row_fill_white
        values = [config, data['turbulence_model'], aoa, f"{C_L:.6f}", f"{C_L_cov:.1f}", f"{C_D:.6f}", f"{C_D_cov:.1f}"]
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = data_alignment
            cell.border = border_style
            cell.fill = fill
        row += 1
    
    # Autofit columns
    for col_idx in range(1, len(columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_optimized_statistics_sheet(wb, all_data, convergence_results, num_iterations, q_times_a):
    """Create Optimized Statistics sheet comparing original vs optimized."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    if not convergence_results:
        return
    
    ws = wb.create_sheet(title='Optimized_Statistics')
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_alignment = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    row_fill_light = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    row_fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Headers
    columns = ['Turbulence Model', 'Configuration', 'AoA', 'Original Iterations', 'Optimal Trim', 'Optimized Iterations',
               'Lift Mean (Orig)', 'Lift Mean (Opt)', 'Lift COV (Orig)', 'Lift COV (Opt)', 'Lift COV Δ',
               'Drag Mean (Orig)', 'Drag Mean (Opt)', 'Drag COV (Orig)', 'Drag COV (Opt)', 'Drag COV Δ']
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    ws.row_dimensions[1].height = 30
    
    # Data rows
    row = 2
    for (config, aoa), data in sorted(all_data.items()):
        if (config, aoa) not in convergence_results:
            continue
        
        conv = convergence_results[(config, aoa)]
        
        # Original stats (last num_iterations)
        orig_lift = data['lift'][-num_iterations:] if len(data['lift']) >= num_iterations else data['lift']
        orig_drag = data['drag'][-num_iterations:] if len(data['drag']) >= num_iterations else data['drag']
        orig_lift_mean, orig_lift_cov = compute_statistics(orig_lift)
        orig_drag_mean, orig_drag_cov = compute_statistics(orig_drag)
        
        # Optimized stats
        lift_min_idx = np.argmin(conv['lift']['cov'])
        drag_min_idx = np.argmin(conv['drag']['cov'])
        optimal_lift_trim = conv['lift']['iterations_removed'][lift_min_idx]
        optimal_drag_trim = conv['drag']['iterations_removed'][drag_min_idx]
        optimal_trim = max(optimal_lift_trim, optimal_drag_trim)
        
        opt_lift = data['lift'][optimal_trim:]
        opt_drag = data['drag'][optimal_trim:]
        opt_lift_mean, opt_lift_cov = compute_statistics(opt_lift)
        opt_drag_mean, opt_drag_cov = compute_statistics(opt_drag)
        
        fill = row_fill_light if row % 2 == 0 else row_fill_white
        values = [
            data['turbulence_model'], config, aoa,
            len(orig_lift), optimal_trim, len(opt_lift),
            f"{orig_lift_mean:.3f}", f"{opt_lift_mean:.3f}",
            f"{orig_lift_cov:.2f}%", f"{opt_lift_cov:.2f}%", f"{(orig_lift_cov - opt_lift_cov):+.2f}%",
            f"{orig_drag_mean:.3f}", f"{opt_drag_mean:.3f}",
            f"{orig_drag_cov:.2f}%", f"{opt_drag_cov:.2f}%", f"{(orig_drag_cov - opt_drag_cov):+.2f}%"
        ]
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = data_alignment
            cell.border = border_style
            cell.fill = fill
        row += 1
    
    # Autofit columns
    for col_idx in range(1, len(columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width


def apply_excel_formatting(excel_file):
    """Legacy function - formatting now done during sheet creation."""
    pass


# ==================== PLOTTING FUNCTIONS ====================

def create_coefficient_graphs(all_data, coefficient_data, output_dir, position_map, value_mappings):
    """Create all coefficient graphs organized by turbulence model and configuration."""
    
    # Group configurations by turbulence model and base config
    configs_by_turbulence = defaultdict(lambda: defaultdict(list))
    
    for (config, aoa), coeff in coefficient_data.items():
        turb_model = coeff['turbulence_model']
        parts = config.split('.')
        base_config = '.'.join(parts[:4]) if len(parts) >= 4 else config
        configs_by_turbulence[turb_model][base_config].append((config, aoa))
    
    # Define styles
    model_styles = {
        'SST': {'color': '#1f77b4', 'marker': 'o', 'label': 'SST'},
        'RNG': {'color': '#ff7f0e', 'marker': 's', 'label': 'RNG k-ε'},
        'RSM': {'color': '#2ca02c', 'marker': '^', 'label': 'RSM'},
        'k-epsilon': {'color': '#d62728', 'marker': 'D', 'label': 'k-ε'}
    }
    
    # Create graphs
    for turb_model in sorted(configs_by_turbulence.keys()):
        for base_config in sorted(configs_by_turbulence[turb_model].keys()):
            config_graphs_dir = os.path.join(output_dir, "coefficient_graphs", turb_model, base_config)
            os.makedirs(config_graphs_dir, exist_ok=True)
            
            # Get coefficient data for this base config
            config_keys = configs_by_turbulence[turb_model][base_config]
            config_coeff_data = {key: coefficient_data[key] for key in config_keys}
            
            if not config_coeff_data:
                continue
            
            # Organize data
            aoa_list, C_L_list, C_D_list, C_L_std_list, C_D_std_list = [], [], [], [], []
            for (config, aoa), coeff in config_coeff_data.items():
                aoa_list.append(coeff['aoa_degrees'])
                C_L_list.append(coeff['C_L'])
                C_D_list.append(coeff['C_D'])
                C_L_std_list.append(coeff['C_L_std'])
                C_D_std_list.append(coeff['C_D_std'])
            
            # Sort by AoA
            combined = list(zip(aoa_list, C_L_list, C_D_list, C_L_std_list, C_D_std_list))
            combined.sort(key=lambda x: x[0])
            
            aoa_vals = np.array([x[0] for x in combined])
            C_L_vals = np.array([x[1] for x in combined])
            C_D_vals = np.array([x[2] for x in combined])
            C_L_std_vals = np.array([x[3] for x in combined])
            C_D_std_vals = np.array([x[4] for x in combined])
            
            style = model_styles.get(turb_model, {'color': '#1f77b4', 'marker': 'o', 'label': turb_model})
            
            # Plot 1: C_L vs AoA
            _plot_coefficient_vs_aoa(aoa_vals, C_L_vals, C_L_std_vals, style, turb_model, base_config,
                                    'Lift Coefficient ($C_L$)', 'Lift Coefficient vs Angle of Attack',
                                    os.path.join(config_graphs_dir, "C_L_vs_AoA.png"))
            
            # Plot 2: C_D vs AoA
            _plot_coefficient_vs_aoa(aoa_vals, C_D_vals, C_D_std_vals, style, turb_model, base_config,
                                    'Drag Coefficient ($C_D$)', 'Drag Coefficient vs Angle of Attack',
                                    os.path.join(config_graphs_dir, "C_D_vs_AoA.png"))
            
            # Plot 3: Combined
            _plot_combined(aoa_vals, C_L_vals, C_D_vals, C_L_std_vals, C_D_std_vals, style, turb_model, base_config,
                         os.path.join(config_graphs_dir, "C_L_C_D_Combined.png"))


def _plot_coefficient_vs_aoa(aoa_vals, coeff_vals, std_vals, style, turb_name, config, ylabel, title, output_path):
    """Helper function to plot coefficient vs AoA."""
    fig, ax = plt.subplots(figsize=(12, 8))
    
    ax.errorbar(aoa_vals, coeff_vals, yerr=std_vals,
                marker=style['marker'], markersize=10, linewidth=2.5, capsize=5,
                color=style['color'], label=turb_name, alpha=0.9)
    
    ax.set_xlabel('Angle of Attack (degrees)', fontsize=14, fontweight='bold')
    ax.set_ylabel(ylabel, fontsize=14, fontweight='bold')
    ax.set_title(f'{title}\n{config}', fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.legend(fontsize=12, loc='best', framealpha=0.9)
    ax.tick_params(labelsize=11)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

def _plot_combined(aoa_vals, C_L_vals, C_D_vals, C_L_std_vals, C_D_std_vals, style, turb_name, config, output_path):
    """Helper function to plot combined C_L and C_D."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 8))
    
    # Left: C_L vs AoA
    ax1.errorbar(aoa_vals, C_L_vals, yerr=C_L_std_vals,
                marker=style['marker'], markersize=10, linewidth=2.5, capsize=5,
                color=style['color'], label=turb_name, alpha=0.9)
    ax1.set_xlabel('Angle of Attack (degrees)', fontsize=14, fontweight='bold')
    ax1.set_ylabel('Lift Coefficient ($C_L$)', fontsize=14, fontweight='bold')
    ax1.set_title(f'Lift Coefficient vs AoA\n{config}', fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3, linestyle='--')
    ax1.legend(fontsize=11, loc='best', framealpha=0.9)
    ax1.tick_params(labelsize=11)
    
    # Right: C_D vs AoA
    ax2.errorbar(aoa_vals, C_D_vals, yerr=C_D_std_vals,
                marker=style['marker'], markersize=10, linewidth=2.5, capsize=5,
                color=style['color'], label=turb_name, alpha=0.9)
    ax2.set_xlabel('Angle of Attack (degrees)', fontsize=14, fontweight='bold')
    ax2.set_ylabel('Drag Coefficient ($C_D$)', fontsize=14, fontweight='bold')
    ax2.set_title(f'Drag Coefficient vs AoA\n{config}', fontsize=14, fontweight='bold')
    ax2.grid(True, alpha=0.3, linestyle='--')
    ax2.legend(fontsize=11, loc='best', framealpha=0.9)
    ax2.tick_params(labelsize=11)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
